"""
scan_omr.py
-----------
Production OMR reader for the NMMS-style 40-question / 4-option sheet.

Pipeline per sheet (no fresh circle detection ever happens here):
  1. align_sheet()      -> warp photo into the canonical frame using the
                            outer table border as anchor points, with
                            blur/border-confidence/orientation checks.
  2. load template.json -> 160 fixed (x, y) bubble centers, built once.
  3. normalize illumination, then sample fill-intensity AND fill-ratio at
     each of the 160 known coordinates.
  4. compute a PER-SHEET adaptive threshold from that sheet's own
     lightest/darkest bubbles (never a hardcoded global constant),
     cross-validated between Otsu and k-means so a bad threshold can be
     detected rather than trusted blindly.
  5. classify each question as a clean single answer, BLANK, MULTI, or
     REVIEW (ambiguous / too close to the threshold to trust), with a
     0-100 confidence score built from multiple independent metrics. The
     code never silently guesses on a close call.

Usage:
    python3 scan_omr.py sheet1.jpg sheet2.jpg ... --out omr_results.xlsx
"""
import argparse
import json
import logging
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill

from align import align_sheet, load_image, blur_score
from calibrate_template import detect_grid_lines, PARAM_SWEEP, HEADER_ROWS

TEMPLATE_PATH = "template.json"
N_QUESTIONS = 40
N_OPTIONS = 4

# How close two candidate darkness scores must be to call it MULTI vs a
# genuine single mark, and how close a single score must be to the
# threshold to be untrustworthy and get flagged REVIEW instead of guessed.
MULTI_MARGIN = 0.12     # fraction of the sheet's fill-value range
REVIEW_MARGIN = 0.06    # fraction of the sheet's fill-value range
MIN_FILL_RATIO = 0.22   # a genuine fill should cover at least ~22% of the ROI
GOOD_FILL_RATIO = 0.55  # coverage at/above this is treated as fully confident
MIN_SCORE_RANGE = 20.0  # below this, the sheet's own contrast is too flat to trust

# --- Image-quality-aware thresholds (WhatsApp-compressed / low-res / blurry
# mobile photos need more lenient, evidence-based handling than a crisp
# flatbed scan - see estimate_image_quality() and classify_question()). ---
LOW_RES_LONG_EDGE = 1600     # source photo long edge below this -> "low_res"
HEAVY_COMPRESSION_BLOCKINESS = 6.0  # 8x8-grid edge-energy ratio above this -> likely heavy JPEG recompression

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("scan_omr")


def load_template(path=TEMPLATE_PATH):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def normalize_illumination(gray, sigma=35):
    """
    Flattens uneven phone-camera lighting/shadow gradients across the page
    by dividing out a heavily-blurred estimate of the local background,
    so a bubble in a shadowed corner isn't systematically scored darker
    than an identical bubble in a bright corner.
    """
    bg = cv2.GaussianBlur(gray, (0, 0), sigmaX=sigma)
    norm = cv2.divide(gray.astype(np.float32), bg.astype(np.float32) + 1e-6)
    norm = np.clip(norm * 160.0, 0, 255).astype(np.uint8)
    return norm


def enhance_for_detection(norm_gray):
    """
    Recovers local contrast lost to JPEG recompression, mobile-camera
    sensor noise, and mild blur - all of which flatten the difference
    between a lightly-filled bubble and blank paper without changing
    geometry at all (this runs strictly after normalize_illumination and
    never touches pixel coordinates).

    Returns TWO images, deliberately kept separate:
      denoised    - mild fastNlMeansDenoising only. Used wherever pixel-
                     level ink/paper SEGMENTATION happens (connected-blob
                     detection, local-background contrast), because
                     running per-bubble segmentation on top of an already
                     locally-contrast-maximized image double-amplifies
                     noise: CLAHE stretches every small tile to use the
                     full 0-255 range, which makes even blank paper's
                     faint grain/print-screen texture look artificially
                     bimodal to a local Otsu split.
      clahe_enhanced - denoised, then CLAHE (contrast-limited adaptive
                     histogram equalization) on top. Used only for the
                     AVERAGE-darkness features (mean/center-weighted),
                     where a consistent contrast boost shared across all
                     4 options of the same question is genuinely helpful
                     for pulling a faint-but-real mark's average further
                     from blank paper's, without needing pixel-level
                     segmentation to still work correctly on top of it.
    """
    denoised = cv2.fastNlMeansDenoising(norm_gray, h=6, templateWindowSize=7, searchWindowSize=21)
    clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(16, 16))
    clahe_enhanced = clahe.apply(denoised)
    return denoised, clahe_enhanced


def estimate_image_quality(orig_img, align_quality):
    """
    Flags the specific real-world degradations called out for mobile/
    WhatsApp photos - low resolution, heavy recompression blockiness, and
    blur (reusing align_sheet's own blur_score rather than recomputing) -
    so classify_question() can relax its evidence bar in a principled,
    documented way instead of either over-trusting noisy pixels or
    dumping every borderline mark into REVIEW on a photo that was always
    going to be noisy.
    """
    h, w = orig_img.shape[:2]
    long_edge = max(h, w)
    low_res = long_edge < LOW_RES_LONG_EDGE

    gray = cv2.cvtColor(orig_img, cv2.COLOR_BGR2GRAY) if orig_img.ndim == 3 else orig_img
    # Blockiness heuristic: JPEG recompression concentrates gradient energy
    # at 8-pixel grid boundaries. Ratio of on-grid to off-grid horizontal/
    # vertical gradient energy well above 1 indicates heavy recompression
    # (typical of images that have been through WhatsApp's transcoder,
    # sometimes more than once).
    gx = np.abs(np.diff(gray.astype(np.float32), axis=1))
    gy = np.abs(np.diff(gray.astype(np.float32), axis=0))
    on_grid_x = gx[:, 7::8].mean() if gx.shape[1] > 7 else 0.0
    off_grid_x = np.delete(gx, np.arange(7, gx.shape[1], 8), axis=1).mean() if gx.shape[1] > 7 else 1.0
    on_grid_y = gy[7::8, :].mean() if gy.shape[0] > 7 else 0.0
    off_grid_y = np.delete(gy, np.arange(7, gy.shape[0], 8), axis=0).mean() if gy.shape[0] > 7 else 1.0
    blockiness = float(((on_grid_x + 1e-6) / (off_grid_x + 1e-6) + (on_grid_y + 1e-6) / (off_grid_y + 1e-6)) / 2.0)
    heavy_compression = blockiness > HEAVY_COMPRESSION_BLOCKINESS

    n_flags = int(low_res) + int(heavy_compression) + int(not align_quality.get("blur_ok", True))

    return {
        "source_width": int(w),
        "source_height": int(h),
        "low_res": bool(low_res),
        "blockiness": round(blockiness, 2),
        "heavy_compression": bool(heavy_compression),
        "blur_score": align_quality.get("blur_score"),
        "degraded_flags": n_flags,
    }


def _circle_mask(shape, cx, cy, r):
    mask = np.zeros(shape, dtype=np.uint8)
    cv2.circle(mask, (cx, cy), r, 255, -1)
    return mask


def _annulus_mask(shape, cx, cy, r_in, r_out):
    mask = np.zeros(shape, dtype=np.uint8)
    cv2.circle(mask, (cx, cy), r_out, 255, -1)
    cv2.circle(mask, (cx, cy), r_in, 0, -1)
    return mask


def extract_bubble_features(seg_gray, clahe_gray, x, y, radius):
    """
    Builds a small feature vector per bubble instead of a single average-
    intensity score, so a faint-but-real mark (common on low-res/WhatsApp
    photos) can still be told apart from JPEG noise or a shadow.

    seg_gray   - denoised-only image (see enhance_for_detection), used for
                 anything that segments ink pixels from paper pixels.
    clahe_gray - denoised+CLAHE image, used only for the average-darkness
                 features, where a shared contrast boost across a
                 question's 4 options helps a faint mark stand out without
                 needing per-pixel segmentation to survive on top of it.

    Sampling uses a patch sized to the full printed radius (not just the
    inner ROI) so there's genuine paper visible OUTSIDE the inner ROI
    circle for the local-background annulus - the previous version sized
    the patch to exactly the inner ROI, leaving zero margin, which made
    local_bg_contrast silently always 0.

      mean_darkness     - inverted mean intensity inside the inner ROI
                           (kept for continuity with the sheet-level prior).
      std_darkness       - spread of darkness inside the ROI.
      center_weighted    - darkness weighted by a Gaussian centred on the
                           bubble, so a genuine fill (denser toward the
                           middle) outweighs a thin stray arc of ink near
                           the printed outline picking up noise.
      connected_frac     - fraction of the inner ROI covered by the
                           LARGEST single connected blob of "dark" pixels,
                           where "dark" is relative to THIS bubble's own
                           local background annulus (not a fixed or
                           globally-Otsu'd cut). A real pencil/pen fill is
                           one solid blob well below the paper's own
                           local darkness; scattered JPEG speckle or
                           CLAHE-amplified paper grain isn't.
      local_bg_contrast  - difference between the ROI's dark-pixel mean
                           and the darkness of a thin annulus of paper
                           just outside the printed circle. Captures
                           contrast against the immediate local
                           background, which a page-wide illumination
                           correction can still miss on a photo with a
                           sharp local shadow edge (e.g. a hand or phone
                           holding the page down).
      fill_ratio_local   - fraction of ROI pixels darker than the
                           local-background-relative cut (local analogue
                           of the old sheet-threshold fill ratio).
    """
    r_inner = max(4, int(radius * 0.72))
    r_outer = max(r_inner + 6, int(radius * 1.05))
    x0, x1 = max(0, x - r_outer), min(seg_gray.shape[1], x + r_outer)
    y0, y1 = max(0, y - r_outer), min(seg_gray.shape[0], y + r_outer)
    seg_patch = seg_gray[y0:y1, x0:x1]
    clahe_patch = clahe_gray[y0:y1, x0:x1]
    empty = {
        "mean_darkness": 0.0, "std_darkness": 0.0, "center_weighted": 0.0,
        "connected_frac": 0.0, "local_bg_contrast": 0.0, "fill_ratio_local": 0.0,
    }
    if seg_patch.size == 0:
        return empty

    cx, cy = seg_patch.shape[1] // 2, seg_patch.shape[0] // 2
    roi_mask = _circle_mask(seg_patch.shape, cx, cy, r_inner)
    roi_vals = seg_patch[roi_mask == 255]
    if roi_vals.size == 0:
        return empty

    # --- average-darkness features: from the CLAHE-enhanced image ---
    clahe_darkness = 255.0 - clahe_patch.astype(np.float32)
    clahe_roi_darkness = clahe_darkness[roi_mask == 255]
    mean_darkness = float(clahe_roi_darkness.mean())
    std_darkness = float(clahe_roi_darkness.std())

    yy, xx = np.mgrid[0:seg_patch.shape[0], 0:seg_patch.shape[1]]
    sigma = max(r_inner * 0.55, 1.0)
    gweight = np.exp(-((xx - cx) ** 2 + (yy - cy) ** 2) / (2 * sigma ** 2))
    gweight = gweight * (roi_mask == 255)
    wsum = gweight.sum()
    center_weighted = float((clahe_darkness * gweight).sum() / wsum) if wsum > 1e-6 else mean_darkness

    # --- segmentation features: from the denoised-only image, using a
    # local-background-relative cut instead of a per-ROI Otsu split ---
    ann_r_in = r_inner + 1
    ann_r_out = min(r_outer, min(seg_patch.shape) // 2)
    seg_darkness = 255.0 - seg_patch.astype(np.float32)
    seg_roi_darkness = seg_darkness[roi_mask == 255]

    connected_frac = 0.0
    fill_ratio_local = 0.0
    local_bg_contrast = 0.0
    if ann_r_out > ann_r_in:
        ann_mask = _annulus_mask(seg_patch.shape, cx, cy, ann_r_in, ann_r_out)
        ann_vals = seg_darkness[ann_mask == 255]
        if ann_vals.size >= 6:
            bg_mean = float(ann_vals.mean())
            bg_std = float(ann_vals.std())
            local_bg_contrast = float(seg_roi_darkness.mean() - bg_mean)
            # A pixel counts as "ink" only if it's meaningfully darker than
            # THIS bubble's own local paper, not the page average - a
            # fixed floor keeps this from firing on pure sensor noise when
            # the local background is already very flat/low-std.
            dark_cut = bg_mean + max(12.0, 2.2 * bg_std)
            dark_mask = ((seg_darkness > dark_cut) & (roi_mask == 255)).astype(np.uint8)
            if dark_mask.any():
                fill_ratio_local = float(dark_mask.sum()) / float((roi_mask == 255).sum())
                n_labels, labels, stats, _ = cv2.connectedComponentsWithStats(dark_mask, connectivity=8)
                if n_labels > 1:
                    largest = stats[1:, cv2.CC_STAT_AREA].max()
                    connected_frac = float(largest) / float((roi_mask == 255).sum())

    return {
        "mean_darkness": mean_darkness,
        "std_darkness": std_darkness,
        "center_weighted": center_weighted,
        "connected_frac": connected_frac,
        "local_bg_contrast": local_bg_contrast,
        "fill_ratio_local": fill_ratio_local,
    }



def per_sheet_threshold(all_scores):
    """
    Derive a fill/no-fill threshold from THIS sheet's own score
    distribution rather than a global constant, so lighting/pen/scan
    differences between sheets don't need separate tuning.

    Cross-validates Otsu against a 2-means clustering of the same scores;
    if they disagree substantially, the sheet's contrast is treated as
    unreliable (surfaced via "agreement": False) so callers can be more
    conservative rather than trusting a single noisy estimate.

    This remains a SHEET-LEVEL prior used for blank-page sanity checks and
    as one signal among several in classify_question() - the primary
    per-question call is now made by comparing each question's own 4
    options against each other (see classify_question()), which is far
    more robust to faint marks on low-quality photos than a single global
    cut ever can be.
    """
    scores = np.array(all_scores, dtype=np.float32)
    scores_8u = np.clip(scores, 0, 255).astype(np.uint8)
    otsu_val, _ = cv2.threshold(scores_8u, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    km_scores = scores.reshape(-1, 1).astype(np.float32)
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 20, 0.5)
    try:
        _, labels, centers = cv2.kmeans(km_scores, 2, None, criteria, 5, cv2.KMEANS_PP_CENTERS)
        centers = centers.flatten()
        lo_c, hi_c = float(min(centers)), float(max(centers))
        kmeans_val = (lo_c + hi_c) / 2.0
        lo_std = float(scores[labels.flatten() == np.argmin(centers)].std()) if scores.size else 0.0
        hi_std = float(scores[labels.flatten() == np.argmax(centers)].std()) if scores.size else 0.0
    except cv2.error:
        kmeans_val = otsu_val
        lo_std = hi_std = float(scores.std())

    rng = max(float(scores.max() - scores.min()), 1e-6)
    # Otsu weights its split by class population, so on this domain
    # (~25% filled bubbles vs ~75% blank) it systematically drifts toward
    # the minority (filled) cluster - that drift is expected, not a
    # quality problem. The meaningful check is whether Otsu still lands
    # somewhere between the two k-means cluster centers (i.e. it agrees
    # there IS a blank cluster and a filled cluster, just not exactly
    # where the midpoint is). If Otsu falls outside that range, the
    # sheet's distribution is genuinely not behaving as expected.
    agreement = lo_c <= otsu_val <= hi_c

    # k-means' cluster-center midpoint is the principled statistic for
    # "where the two classes meet" and is used as the working threshold.
    # If Otsu disagrees badly (outside the cluster-center range), lean
    # conservative (higher) since the sheet's bimodality is unreliable.
    threshold = kmeans_val if agreement else max(otsu_val, kmeans_val)

    return {
        "threshold": float(threshold),
        "score_min": float(scores.min()),
        "score_max": float(scores.max()),
        "otsu": float(otsu_val),
        "kmeans": float(kmeans_val),
        "agreement": bool(agreement),
        "blank_cluster_std": lo_std,
        "fill_cluster_std": hi_std,
    }


def _robust_zscore(values, clip=3.0):
    """z-score using median/MAD instead of mean/std - resistant to a single
    outlier (the actual filled option) skewing the very spread it's being
    measured against, which mean/std would be prone to on a 4-sample set.

    Clipped to +/-clip: when the reference cluster (the other 3 options)
    is near-identical, MAD collapses toward 0 and even a small absolute
    difference blows up into a huge z-score - which would then dominate a
    weighted combination with better-behaved features regardless of how
    small its weight is. Clipping keeps any one feature's vote bounded."""
    values = np.array(values, dtype=np.float64)
    med = np.median(values)
    mad = np.median(np.abs(values - med))
    scale = mad * 1.4826 if mad > 1e-6 else (values.std() + 1e-6)
    z = (values - med) / max(scale, 1e-6)
    return np.clip(z, -clip, clip)


def classify_question(features, threshold, score_range, image_quality=None):
    """
    features: list of 4 per-option feature dicts from extract_bubble_features().
    threshold / score_range: this SHEET's own per_sheet_threshold() output,
      used as a secondary sanity prior (e.g. "is this whole question just
      flat/blank") rather than the primary fill/no-fill decision.
    image_quality: optional estimate_image_quality() dict; when the photo
      is flagged degraded (low-res/heavily-compressed/blurry), the
      evidence bar for trusting a borderline mark as OK is relaxed a
      little rather than reflexively demoting every faint-but-real mark
      on a known-noisy photo to REVIEW.

    Returns (selected_option_or_None, status, confidence_0_100, note) -
    same contract as before.

    Decision rule (per question, comparing its own 4 options horizontally):
      - If one option is clearly the darkest of the four -> that's the
        answer (OK), even if it's faint in absolute terms. A real mark
        that's fainter than typical is still a real mark, not a reason to
        second-guess the call.
      - If none of the four stands out from the others -> BLANK.
      - If two options are both meaningfully dark and neither is a clear
        winner over the other -> MULTI (can't tell which was intended).
      - Only the genuinely ambiguous boundary right between "darkest" and
        "blank" (too close to confidently call either way) still gets
        REVIEW, as a narrow safety net against pure noise - not as a
        general low-confidence catch-all.

    Each option is scored by how it compares to the OTHER THREE options in
    the SAME question (a per-question, per-photo-local "adaptive
    threshold"), using several features z-scored ACROSS the question's 4
    options independently first (median/MAD, so a single real outlier
    doesn't get diluted by mixing raw feature magnitudes - which live on
    very different natural scales - into one composite before any
    statistical comparison happens), then combined. This matters most on
    genuinely faint marks: a lightly filled bubble often shows a clear
    signal in ONE feature (e.g. center-weighted darkness) while a noisier
    feature (e.g. connected-ink blob size, when the mark is more a smudge
    than a solid fill) stays flat - z-scoring first means the clear signal
    still wins instead of being averaged away by the flat one.
    """
    rng = max(score_range, 1e-6)
    mean_scores = [f["mean_darkness"] for f in features]
    center_scores = [f["center_weighted"] for f in features]
    connected = [f["connected_frac"] for f in features]
    fill_local = [f["fill_ratio_local"] for f in features]
    bg_contrast = [f["local_bg_contrast"] for f in features]

    z_center = _robust_zscore(center_scores)
    z_connected = _robust_zscore(connected)
    z_filllocal = _robust_zscore(fill_local)
    z_bg = _robust_zscore(bg_contrast)

    # Center-weighted darkness and local-background contrast get the most
    # trust: both are CONTINUOUS measures that still move cleanly on a
    # genuinely faint/smudged mark. connected_frac/fill_local depend on a
    # binary per-pixel "dark enough" cut, which can legitimately land on
    # exactly 0 for a soft, diffuse mark even when it's clearly the
    # darkest option overall - so they're kept as smaller corroborating
    # votes rather than co-equal with the continuous features.
    combined_z = 0.50 * z_center + 0.25 * z_bg + 0.15 * z_connected + 0.10 * z_filllocal

    best_idx = int(np.argmax(combined_z))
    best_z = float(combined_z[best_idx])
    others_z = sorted([float(v) for i, v in enumerate(combined_z) if i != best_idx], reverse=True)
    second_z = others_z[0] if others_z else -3.0

    degraded = bool(image_quality and image_quality.get("degraded_flags", 0) > 0)
    # A degraded photo gets a little more benefit of the doubt before a
    # genuine mark is trusted as intentional - the ink is real, it's just
    # harder to see through compression/blur, not more likely to be noise.
    # blank_bar is deliberately the ONLY line between "no option stands
    # out" (BLANK) and "there IS a darkest option" (OK): once an option
    # clears it, it's the answer, however faint - a real mark shouldn't
    # be second-guessed just because it's a light fill. It stays well
    # above pure-noise levels (a handful of stray dark pixels from JPEG
    # artifacting or sensor noise shouldn't clear it on their own).
    blank_bar = 0.40 if not degraded else 0.30
    fill_floor = MIN_FILL_RATIO if not degraded else MIN_FILL_RATIO * 0.7

    # No option stands out at all relative to the other three -> BLANK.
    if best_z < blank_bar:
        return None, "BLANK", 95.0, ""

    # Two (or more) options both look meaningfully darker than the rest,
    # and aren't clearly separated from each other -> MULTI. The second
    # option has to independently clear the SAME bar used to call "there
    # is a dark option" (not just be a little above zero) - otherwise a
    # single weak-but-real mark next to ordinary background noise in
    # another option would get miscalled as a tie instead of resolving
    # to its one genuine answer.
    z_gap = best_z - second_z
    if second_z > blank_bar and z_gap < 0.9:
        confidence = 100.0 * (1.0 - np.clip(z_gap / 0.9, 0.0, 1.0))
        return None, "MULTI", round(confidence, 1), "two or more bubbles darker than the rest of the row with no clear winner"

    fill_best = max(connected[best_idx], fill_local[best_idx])
    conn_conf = np.clip(z_connected[best_idx] / 2.0, 0.0, 1.0)
    z_conf = np.clip((best_z - blank_bar) / 2.0, 0.0, 1.0)
    gap_conf = np.clip(z_gap / 1.5, 0.0, 1.0)
    bg_conf = np.clip(z_bg[best_idx] / 2.0, 0.0, 1.0)
    confidence = round(100.0 * (0.35 * z_conf + 0.25 * conn_conf + 0.25 * gap_conf + 0.15 * bg_conf), 1)

    # Secondary sanity note against the sheet-wide threshold - kept in the
    # confidence score and notes for transparency, but no longer used to
    # demote a genuine per-question winner to REVIEW: once an option is
    # clearly the darkest of its own row (cleared the BLANK bar above)
    # and isn't tied with a second option (cleared the MULTI check above),
    # it IS the answer - a faint mark is still a real mark.
    sheet_margin = (mean_scores[best_idx] - threshold) / rng
    note = ""
    if fill_best < fill_floor and sheet_margin < -0.15:
        note = "faint relative to the sheet as a whole, but clearly the darkest option within its own row"

    return best_idx + 1, "OK", confidence, note





def _detect_rows_via_column_strips(gray, cols, strip_half_width=6, min_gap=25):
    """Supplementary row-line detector used when the standard binarize-
    then-erode approach (detect_grid_lines) can't recover every internal
    row divider - e.g. on a photo where the data-row cells are printed on
    a light gray screened background rather than plain white, which some
    rows' thin divider ink doesn't binarize/erode cleanly against.

    Rather than threshold the whole page, this sums raw (unthresholded)
    darkness only in narrow vertical strips centred on each ALREADY
    detected column line (`cols`, from the normal, reliable column
    detection). Those strips sit on blank line-only territory - no bubble
    ink, no digits - so a horizontal divider shows up as a clean periodic
    peak in the summed darkness even when the surrounding cell shading
    would confuse a global binarize step.
    """
    H, W = gray.shape
    darkness = 255.0 - gray.astype(np.float64)
    proj = np.zeros(H)
    for x in cols:
        lo = max(0, int(x) - strip_half_width)
        hi = min(W, int(x) + strip_half_width + 1)
        if hi > lo:
            proj += darkness[:, lo:hi].sum(axis=1)
    kernel = np.ones(5) / 5.0
    proj_s = np.convolve(proj, kernel, mode="same")
    baseline = np.median(proj_s)
    thresh = baseline + 0.3 * (proj_s.max() - baseline)
    peaks = []
    for i in range(1, H - 1):
        if proj_s[i] > thresh and proj_s[i] >= proj_s[i - 1] and proj_s[i] >= proj_s[i + 1]:
            peaks.append(i)
    dedup = []
    for p in peaks:
        if not dedup or p - dedup[-1] > min_gap:
            dedup.append(p)
        else:
            dedup[-1] = (dedup[-1] + p) // 2
    return dedup


def grid_correct(gray, template):
    """
    The outer-border homography (align.py) gets a photo roughly into the
    canonical frame, but on some real photos (uneven table-vs-page contrast,
    a stray notch where a divider line meets the border, etc.) the border
    contour itself is a few percent off in scale/position - enough to
    silently shift every one of the 160 bubble coordinates by a consistent
    amount and turn genuine fills into false BLANKs.

    The fix: the sheet's own printed gridlines are ground truth (they're
    fixed by the print layout), so re-detect them on THIS warped sheet with
    the same method used at calibration time, and fit a simple per-axis
    linear correction (scale + offset) from this sheet's line positions to
    the reference sheet's line positions stored in template.json. This is
    strictly more reliable than trusting the border homography alone,
    because it's anchored to the exact same printed structure the bubble
    coordinates were derived from.

    Note on fit direction: the correction must map a bubble's REFERENCE
    (template) coordinate to the coordinate it actually occupies on THIS
    sheet's own warped canvas - i.e. ref-space -> detected-space. The fit
    below is therefore built as polyfit(ref_positions, detected_positions),
    not the other way around, so it can be applied directly (ax*x_ref+bx)
    without an inversion step at call time.

    Three refinements on top of a single global fit:

    - COLUMN correction is fit INDEPENDENTLY per question-block (left
      Q1-20 / right Q21-40), not as one line across the full table width.
      The two blocks are visually separate halves of the sheet, so
      non-uniform distortion (lens distortion, a slight page curl through
      the middle of the photo) can leave one half under-corrected while a
      single global x-fit over-corrects the other. Block boundaries come
      from "col_block_ranges" in template.json (written by
      calibrate_template.py); an older template without that field falls
      back to one block spanning the whole width, i.e. the previous
      behaviour.
    - Both the row fit and each column-block fit give extra weight to the
      OUTER BIG BORDER line at each end (the thick printed table edge, or
      for an inner block edge, the shared centre divider) versus the thin
      internal option/row dividers. The outer border is the least
      ambiguous, highest-contrast line on the sheet, so it should anchor
      scale/offset more than a noisier internal line rather than being
      averaged in as just one line among many.
    - The ROW axis and COLUMN axis are searched INDEPENDENTLY across the
      adaptive-threshold parameter sweep, instead of both being required
      to match on the very same (block_size, C) attempt. In practice the
      11 column lines detect cleanly under almost any parameter choice,
      but on some real photos (e.g. a grayish screened cell background)
      the thin internal row dividers binarize inconsistently and break up
      under erosion even though they're clearly visible in the raw pixel
      values - while the columns are still fine. Tying both axes to one
      shared attempt would throw away a perfectly good column fit just
      because that particular attempt's row count came up short.
    - If NO parameter attempt recovers every internal row line, rows first
      try a full match against just the DATA rows (skipping the two
      header-only lines, which no bubble is ever placed against), then a
      supplementary column-strip-based re-detection (using the already-
      reliable column positions to sample cleanly and dodge a low-
      contrast cell background), and only as a last resort fall back to a
      2-point fit anchored on just the outermost detected line pair - the
      OUTER BIG BORDER. That border is by far the thickest, highest-
      contrast, most reliably detected line on the sheet (and, since
      align_sheet's own perspective warp targets it directly, it also
      tends to sit close to the canvas edges already) so anchoring row
      scale/offset on just those two points is far more trustworthy than
      no correction at all - but it can't fix non-uniform local spacing
      the way a full per-line match can, so it's tried last.

    Returns (correct_xy, matched: bool). If neither axis can establish
    even the outer-border-anchor fallback, correct_xy is the identity
    function and matched is False, so callers can fall back to the
    border-only alignment (with a lower-confidence note).
    """
    ref_rows = template.get("ref_row_lines")
    ref_cols = template.get("ref_col_lines")
    if not ref_rows or not ref_cols:
        return (lambda x, y: (x, y)), False

    n_rows, n_cols = len(ref_rows), len(ref_cols)
    ref_rows_arr = np.array(ref_rows)
    ref_cols_arr = np.array(ref_cols)

    # Column blocks to fit independently. Falls back to a single block
    # covering every column line (old global-fit behaviour) if the
    # template predates this field.
    col_block_ranges = template.get("col_block_ranges") or [[0, n_cols]]

    def _weighted_border_fit(ref_positions, detected_positions):
        """polyfit(ref -> detected) with the two OUTER endpoints (the big
        printed border on each side) weighted more than interior lines,
        so the least ambiguous, highest-contrast line on the sheet
        anchors scale/offset more than a noisier internal divider."""
        w = np.ones(len(ref_positions))
        w[0] *= 4.0
        w[-1] *= 4.0
        a, b = np.polyfit(ref_positions, detected_positions, 1, w=w)
        return a, b

    def _fit_axis(detected_arr, ref_arr):
        """Tier 1: full weighted fit if every reference line was matched
        one-for-one. Tier 2 (fallback): a robust 2-point fit anchored on
        just the outermost detected line pair - the OUTER BIG BORDER -
        used when the internal lines couldn't all be recovered. Returns
        (a, b, is_full_quality) or None if neither tier is trustworthy.
        """
        if len(detected_arr) == len(ref_arr) and len(detected_arr) >= 2:
            a, b = _weighted_border_fit(ref_arr, detected_arr)
            if 0.85 <= a <= 1.15:
                resid = np.abs(a * ref_arr + b - detected_arr)
                if resid.max() <= 12.0:
                    return a, b, True
        if len(detected_arr) >= 2:
            lo, hi = float(detected_arr.min()), float(detected_arr.max())
            ref_lo, ref_hi = float(ref_arr[0]), float(ref_arr[-1])
            ref_span = ref_hi - ref_lo
            # Only trust this pair as "the outer border" if the overall
            # span it covers is plausibly close to the reference span -
            # guards against locking onto two unrelated interior lines.
            if ref_span > 1e-6 and abs((hi - lo) - ref_span) < 0.25 * ref_span:
                a = (hi - lo) / ref_span
                b = lo - a * ref_lo
                if 0.85 <= a <= 1.15:
                    return a, b, False
        return None

    def _fit_rows(detected_arr):
        """Same tiers as _fit_axis, plus a middle tier: a full match of
        every DATA-row line, skipping the two header-only lines at the
        very top (the "options 1-4" label rows). Those two lines aren't
        used to place any bubble (see calibrate_template.py's
        data_row_boundaries = row_lines[HEADER_ROWS:]), so a fit that
        recovers all 21 data-row lines but misses the 2 header ones is
        just as accurate for correction purposes as recovering all 23 -
        it shouldn't be discarded in favour of a cruder 2-point fallback.
        """
        full = _fit_axis(detected_arr, ref_rows_arr)
        if full is not None and full[2]:
            return full
        ref_data = ref_rows_arr[HEADER_ROWS:]
        data_only = _fit_axis(detected_arr, ref_data)
        if data_only is not None and data_only[2]:
            return data_only
        # Neither full-quality tier worked; keep whichever anchor-only
        # (tier 2) result is available, preferring the one against the
        # complete reference span.
        return full or data_only

    best_row_fit = None      # (ay, by, is_full)
    best_col_fits = None     # [(ref_lo, ref_hi, ax, bx), ...] per block

    for block_size, c_val in PARAM_SWEEP:
        rows, cols = detect_grid_lines(gray, block_size, c_val)
        rows_arr = np.array(rows)

        row_fit = _fit_rows(rows_arr)
        if row_fit is not None:
            if best_row_fit is None or (row_fit[2] and not best_row_fit[2]):
                best_row_fit = row_fit

        # If the standard detector didn't land a full-quality row fit but
        # DID find all the column lines, use those reliable column
        # positions to re-detect rows via the column-strip method - much
        # more robust against a low-contrast/screened row background.
        if (best_row_fit is None or not best_row_fit[2]) and len(cols) == n_cols:
            alt_rows = np.array(_detect_rows_via_column_strips(gray, cols))
            alt_fit = _fit_rows(alt_rows)
            if alt_fit is not None:
                if best_row_fit is None or (alt_fit[2] and not best_row_fit[2]):
                    best_row_fit = alt_fit

        if best_col_fits is None and len(cols) == n_cols:
            cols_arr = np.array(cols)
            fits = []
            blocks_ok = True
            for lo, hi in col_block_ranges:
                c_slice = cols_arr[lo:hi]
                r_slice = ref_cols_arr[lo:hi]
                if len(c_slice) < 2:
                    blocks_ok = False
                    break
                ax, bx = _weighted_border_fit(r_slice, c_slice)
                if not (0.85 <= ax <= 1.15):
                    blocks_ok = False
                    break
                resid = np.abs(ax * r_slice + bx - c_slice)
                if resid.max() > 12.0:
                    blocks_ok = False
                    break
                fits.append((float(r_slice[0]), float(r_slice[-1]), ax, bx))
            if blocks_ok:
                best_col_fits = fits

        # Once rows have a full-quality fit and columns are solved, further
        # sweeping can't improve on either - stop early.
        if best_row_fit is not None and best_row_fit[2] and best_col_fits is not None:
            break

    if best_row_fit is None or best_col_fits is None:
        return (lambda x, y: (x, y)), False

    ay, by, _row_full = best_row_fit
    col_fits = best_col_fits

    def correct_xy(x, y, ay=ay, by=by, col_fits=col_fits):
        # Bubble template x-coordinates are column CENTRES, so they
        # always fall strictly inside exactly one block's reference
        # span - pick that block's own fit rather than a global one.
        for lo_ref, hi_ref, ax, bx in col_fits:
            if lo_ref - 1e-3 <= x <= hi_ref + 1e-3:
                return ax * x + bx, ay * y + by
        # Defensive fallback (shouldn't trigger for a valid template):
        # nearest block by distance from its reference span.
        ax, bx = min(
            col_fits,
            key=lambda f: min(abs(x - f[0]), abs(x - f[1])),
        )[2:4]
        return ax * x + bx, ay * y + by

    return correct_xy, True


def scan_sheet(image_path, template, debug_dir=None):
    img = load_image(image_path)
    if img is None:
        raise FileNotFoundError(image_path)

    warped, align_quality = align_sheet(img, out_size=(template["canon_w"], template["canon_h"]))
    gray = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
    norm_gray = normalize_illumination(gray)
    # Recover local contrast lost to JPEG recompression/blur/sensor noise
    # (mobile/WhatsApp photos especially) before any bubble is sampled -
    # purely a pixel-value operation, no effect on geometry/coordinates.
    enh_gray, clahe_gray = enhance_for_detection(norm_gray)
    radius = template["radius"]

    image_quality = estimate_image_quality(img, align_quality)

    # 0. Cross-check the border-based warp against this sheet's own
    # printed gridlines and correct any residual scale/offset drift before
    # sampling - see grid_correct() for why this matters.
    correct_xy, grid_matched = grid_correct(gray, template)
    corrected_bubbles = {
        key: dict(zip(("x", "y"), correct_xy(c["x"], c["y"])))
        for key, c in template["bubbles"].items()
    }

    # 1. extract the full per-bubble feature set (mean/center-weighted
    # darkness, largest connected ink blob, local background contrast,
    # local fill ratio) on the enhanced image for all 160 bubbles.
    features = {}
    for key, c in corrected_bubbles.items():
        features[key] = extract_bubble_features(enh_gray, clahe_gray, int(round(c["x"])), int(round(c["y"])), radius)

    # 2. per-sheet adaptive threshold, cross-validated (Otsu vs k-means).
    # Kept as a secondary sanity prior for classify_question() (e.g.
    # catching a wholly blank/washed-out sheet) - the primary per-question
    # decision now compares each question's own 4 options against each
    # other, which is far more robust to faint marks than one global cut.
    mean_scores = {k: f["mean_darkness"] for k, f in features.items()}
    thr_info = per_sheet_threshold(list(mean_scores.values()))
    threshold = thr_info["threshold"]
    score_range = thr_info["score_max"] - thr_info["score_min"]

    sheet_notes = []
    if not align_quality["blur_ok"]:
        sheet_notes.append(f"image may be blurry (focus score {align_quality['blur_score']})")
    if align_quality["border_confidence"] != "high":
        sheet_notes.append(f"sheet border detected with {align_quality['border_confidence']} confidence "
                            f"(method={align_quality['border_method']})")
    if not thr_info["agreement"]:
        sheet_notes.append("fill/no-fill threshold estimators disagreed; using conservative estimate")
    low_contrast = score_range < MIN_SCORE_RANGE
    if low_contrast:
        sheet_notes.append(f"very low contrast across bubbles (range={score_range:.1f}); sheet may be blank/washed out")
    if not grid_matched:
        sheet_notes.append("could not verify alignment against printed gridlines; relying on border detection only")
    if image_quality["low_res"]:
        sheet_notes.append(f"low-resolution source photo ({image_quality['source_width']}x{image_quality['source_height']}); "
                            f"relaxed confidence cutoff applied")
    if image_quality["heavy_compression"]:
        sheet_notes.append(f"heavy JPEG recompression detected (blockiness={image_quality['blockiness']}); "
                            f"typical of WhatsApp-forwarded images")

    # NOTE (fixed): previously this only escalated to REVIEW on a grid-line
    # mismatch when border_confidence was ALSO "low" - but a "high"-
    # confidence border contour can still be off by enough pixels to
    # mis-sample every bubble (that's the exact scenario grid_correct()
    # exists to catch, per its own docstring). Verified against a
    # synthetic sheet: with grid_matched=False and border_confidence=
    # "high", a fully BLANK sheet was silently scored with several
    # confident MULTI/OK calls instead of being flagged - i.e. the
    # geometry safety net had a gap whenever the secondary (gridline)
    # verification failed but the primary (border) detection looked fine.
    # grid_matched=False now always forces REVIEW, matching the same
    # standalone treatment already given to blur_ok/low_contrast, so a
    # sheet whose bubble coordinates were never cross-verified is never
    # silently trusted.
    force_review = (
        not align_quality["blur_ok"]
        or not grid_matched
        or low_contrast
    )

    # 3. classify each question using the per-question relative comparison
    # (see classify_question() docstring for why this replaces a single
    # global fill/no-fill cut as the primary decision).
    rows = []
    for q in range(1, N_QUESTIONS + 1):
        q_features = [features[f"{q}_{opt}"] for opt in range(1, N_OPTIONS + 1)]
        selected, status, confidence, note = classify_question(q_features, threshold, score_range, image_quality)

        if force_review and status != "REVIEW":
            # A severe sheet-level quality issue (blur, unlocatable border,
            # or a near-flat score distribution) means the per-sheet
            # threshold itself can't be trusted, so BLANK/MULTI/OK calls
            # are all equally suspect - not just the "OK" ones.
            status = "REVIEW"
            confidence = min(confidence, 40.0)
            note = (note + "; " if note else "") + "sheet-level quality flag"

        rows.append({
            "Question": q,
            "Selected_Option": selected if selected is not None else "",
            "Status": status,
            "Confidence": confidence,
            "Notes": note,
        })

    if debug_dir:
        dbg = warped.copy()
        for q in range(1, N_QUESTIONS + 1):
            row = rows[q - 1]
            for opt in range(1, N_OPTIONS + 1):
                c = corrected_bubbles[f"{q}_{opt}"]
                x, y = int(c["x"]), int(c["y"])
                if row["Status"] == "MULTI":
                    color = (0, 0, 255)
                elif row["Status"] == "REVIEW":
                    color = (0, 165, 255)
                elif row["Selected_Option"] == opt:
                    color = (0, 200, 0)
                else:
                    color = (200, 200, 200)
                cv2.circle(dbg, (x, y), radius, color, 2)
        Path(debug_dir).mkdir(parents=True, exist_ok=True)
        cv2.imwrite(str(Path(debug_dir) / f"{Path(image_path).stem}_debug.jpg"), dbg)

    meta = {
        "threshold": threshold,
        "score_min": thr_info["score_min"],
        "score_max": thr_info["score_max"],
        "align_quality": align_quality,
        "grid_matched": grid_matched,
        "sheet_notes": sheet_notes,
        "image_quality": image_quality,
    }
    return rows, meta


def _scan_one(args):
    """Picklable worker wrapper for ProcessPoolExecutor."""
    img_path, template, debug_dir = args
    name = Path(img_path).stem
    try:
        rows, meta = scan_sheet(img_path, template, debug_dir=debug_dir)
        return name, rows, meta, None
    except Exception as e:  # noqa: BLE001 - report and continue the batch
        return name, None, None, str(e)


STATUS_FILLS = {
    "MULTI": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "REVIEW": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "BLANK": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


def _write_excel(out_path, summary_rows, all_sheets):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        for name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            status_col = df.columns.get_loc("Status") + 1
            for row_idx, status in enumerate(df["Status"], start=2):
                fill = STATUS_FILLS.get(status)
                if fill:
                    ws.cell(row=row_idx, column=status_col).fill = fill


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("images", nargs="+", help="Sheet photo(s) to scan")
    ap.add_argument("--out", default="omr_results.xlsx")
    ap.add_argument("--debug-dir", default="debug_overlays")
    ap.add_argument("--workers", type=int, default=1,
                     help="Parallel worker processes for large batches (default: 1)")
    args = ap.parse_args()

    template = load_template()

    all_sheets = {}
    summary_rows = []
    tasks = [(p, template, args.debug_dir) for p in args.images]

    def handle_result(name, rows, meta, error):
        if error:
            log.error("[FAILED] %s: %s", name, error)
            summary_rows.append({"Sheet": name, "Status": f"FAILED: {error}"})
            return
        df = pd.DataFrame(rows)
        all_sheets[name[:31]] = df  # Excel sheet name char limit

        n_review = (df["Status"] == "REVIEW").sum()
        n_multi = (df["Status"] == "MULTI").sum()
        n_blank = (df["Status"] == "BLANK").sum()
        n_ok = (df["Status"] == "OK").sum()
        avg_conf = round(df["Confidence"].mean(), 1)
        notes = "; ".join(meta["sheet_notes"]) if meta["sheet_notes"] else ""
        summary_rows.append({
            "Sheet": name, "OK": n_ok, "Blank": n_blank,
            "Multi": n_multi, "Review": n_review,
            "Avg_Confidence": avg_conf,
            "Threshold": round(meta["threshold"], 1),
            "Sheet_Notes": notes,
        })
        log.info("%s: OK=%d BLANK=%d MULTI=%d REVIEW=%d avg_conf=%.1f%s",
                  name, n_ok, n_blank, n_multi, n_review, avg_conf,
                  f" [{notes}]" if notes else "")

    if args.workers > 1:
        with ProcessPoolExecutor(max_workers=args.workers) as pool:
            futures = [pool.submit(_scan_one, t) for t in tasks]
            for fut in as_completed(futures):
                handle_result(*fut.result())
    else:
        for t in tasks:
            handle_result(*_scan_one(t))

    _write_excel(args.out, summary_rows, all_sheets)

    log.info("Wrote %s", args.out)
    if any(r.get("Review", 0) or r.get("Multi", 0) for r in summary_rows):
        log.warning("Sheets contain REVIEW/MULTI rows that need manual eyes - "
                     "check %s/ overlays before trusting scores.", args.debug_dir)


if __name__ == "__main__":
    main()
