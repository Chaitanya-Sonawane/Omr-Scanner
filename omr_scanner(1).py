"""
OMR Checker - Updated with Advanced Pen Marking Detection (Backup Copy)

This is a backup copy of the improved OMR scanner. The main version is omr_scanner.py
"""

import cv2
import numpy as np
import json
import sys


def detect_bubbles(img_path, debug_out=None):
    """
    Detect bubbles on OMR sheet using adaptive detection method.
    This function works with pen marks and phone camera images.
    
    Args:
        img_path: Path to the OMR sheet image
        debug_out: Optional path to save debug visualization
        
    Returns:
        tuple: (answers, flags, raw_results)
    """
    img = cv2.imread(img_path)
    orig = img.copy()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray_blur = cv2.medianBlur(gray, 5)
    h, w = gray.shape

    # --- Step 1: find circles to CALIBRATE the grid (columns, rows, radius) ---
    circles = cv2.HoughCircles(
        gray_blur, cv2.HOUGH_GRADIENT, dp=1, minDist=20,
        param1=60, param2=22, minRadius=10, maxRadius=22
    )
    if circles is None:
        raise RuntimeError(f"No circles found in {img_path}")
    circles = np.round(circles[0]).astype(int)  # x, y, r

    # drop edge noise (page border, handwriting, roll-no box)
    circles = [c for c in circles if 50 < c[0] < w - 50 and 250 < c[1] < h - 30]

    # drop stray outliers above the grid (e.g. circular institute stamp/logo)
    ys_all = sorted(c[1] for c in circles)
    cut_y = None
    for i in range(1, min(15, len(ys_all))):
        if ys_all[i] - ys_all[i - 1] > 100:
            cut_y = ys_all[i]
    if cut_y is not None:
        circles = [c for c in circles if c[1] >= cut_y]

    pts = np.array([[c[0], c[1]] for c in circles])
    radius_est = int(np.median([c[2] for c in circles]))
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    # split into the two column-blocks (Q1-20 left, Q21-40 right)
    xs = pts[:, 0].reshape(-1, 1).astype(np.float32)
    
    # Check if we have enough points for clustering
    if len(xs) < 2:
        raise RuntimeError(f"Not enough circles detected: {len(xs)}. Need at least 2 circles.")
    
    # Use min of available points and 2 for block clustering
    n_block_clusters = min(2, len(xs))
    _, block_labels, block_centers = cv2.kmeans(xs, n_block_clusters, None, criteria, 10, cv2.KMEANS_PP_CENTERS)
    
    if n_block_clusters == 1:
        # Only one block detected, assume it's the left block
        left_block_id = 0
        block_labels = np.zeros_like(block_labels)
    else:
        left_block_id = int(np.argmin(block_centers.flatten()))

    results = {}
    debug_img = orig.copy()

    for block_id, q_offset in [(left_block_id, 0), (1 - left_block_id, 20)]:
        mask = block_labels.flatten() == block_id
        block_pts = pts[mask]

        # --- calibrate 4 column centers (option 1-4, left to right) ---
        bxs = block_pts[:, 0].reshape(-1, 1).astype(np.float32)
        _, col_labels, col_centers = cv2.kmeans(bxs, 4, None, criteria, 10, cv2.KMEANS_PP_CENTERS)
        col_centers_sorted = sorted(col_centers.flatten().tolist())

        # --- calibrate row spacing (20 rows, evenly spaced top to bottom) ---
        ys = block_pts[:, 1].astype(np.float64)
        y_min, y_max = ys.min(), ys.max()
        row_pitch = (y_max - y_min) / 19.0

        # --- fit a per-column linear model x(row) = a*row + b ---
        # A single fixed x per column assumes the page is perfectly upright.
        # Real photos are slightly tilted/perspective-skewed, so a bubble's
        # true x can drift several px from row 1 to row 20. We fit a line per
        # column using the detected circles themselves so sampling tracks
        # that drift instead of using one static center for all 20 rows.
        cluster_order = np.argsort(col_centers.flatten())
        cluster_to_opt = {int(cluster_order[i]): i for i in range(4)}

        col_line = {}
        for opt_idx in range(4):
            xs_c, rows_c = [], []
            for i in range(len(block_pts)):
                if cluster_to_opt[int(col_labels[i][0])] != opt_idx:
                    continue
                px, py = block_pts[i]
                row_idx_guess = int(round((py - y_min) / row_pitch))
                xs_c.append(px)
                rows_c.append(row_idx_guess)
            if len(set(rows_c)) >= 2:
                a, b = np.polyfit(rows_c, xs_c, 1)
            else:
                a, b = 0.0, (xs_c[0] if xs_c else col_centers_sorted[opt_idx])
            col_line[opt_idx] = (a, b)

        # --- Step 2: sample the FULL theoretical grid directly from the image,
        # regardless of whether Hough found a circle there, using the fitted
        # per-column line so each row's sample lands on the true bubble even
        # under slight photo tilt/perspective. ---
        for row_idx in range(20):
            q_num = row_idx + 1 + q_offset
            y = int(round(y_min + row_idx * row_pitch))
            row_vals = {}
            for opt_idx in range(4):
                a, b = col_line[opt_idx]
                x = int(round(a * row_idx + b))
                mask_c = np.zeros(gray.shape, dtype=np.uint8)
                cv2.circle(mask_c, (x, y), max(radius_est - 3, 8), 255, -1)
                mean_val = cv2.mean(gray, mask=mask_c)[0]
                row_vals[opt_idx + 1] = mean_val
                cv2.circle(debug_img, (x, y), radius_est, (0, 255, 0), 2)
            results[q_num] = row_vals

    if debug_out:
        cv2.imwrite(debug_out, debug_img)

    # --- Step 3: pick the filled (darkest) option per question ---
    answers = {}
    flags = {}
    for q in sorted(results.keys()):
        opts = results[q]
        sorted_opts = sorted(opts.items(), key=lambda x: x[1])
        darkest_opt, darkest_val = sorted_opts[0]
        second_val = sorted_opts[1][1]
        gap = second_val - darkest_val

        if darkest_val < 190 and gap > 15:
            answers[q] = darkest_opt
        elif darkest_val < 190:
            answers[q] = darkest_opt
            flags[q] = "low_confidence_gap"
        else:
            answers[q] = None
            flags[q] = "no_clear_mark"

    return answers, flags, results


def score_sheet(answers, answer_key):
    """answer_key: dict {q_num: correct_option}. Returns (score, per_question_correct)."""
    correct = {}
    score = 0
    for q, key_opt in answer_key.items():
        marked = answers.get(q)
        is_correct = marked == key_opt
        correct[q] = is_correct
        if is_correct:
            score += 1
    return score, correct


def batch_to_excel(image_paths, out_xlsx, answer_key=None, roll_numbers=None):
    """Scan a batch of OMR sheets and write results (+ optional scoring) to an .xlsx file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    header = ["Sheet #", "Roll No"] + [f"Q{i}" for i in range(1, 41)]
    if answer_key:
        header += ["Score", "Flagged (low confidence)"]
    else:
        header += ["Flagged (low confidence)"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    yellow = PatternFill("solid", start_color="FFFF00")

    for idx, path in enumerate(image_paths, start=1):
        answers, flags, _ = detect_bubbles(path)
        roll = roll_numbers[idx - 1] if roll_numbers else ""
        row = [idx, roll] + [answers.get(q, "") for q in range(1, 41)]
        if answer_key:
            score, _ = score_sheet(answers, answer_key)
            row += [score, ", ".join(f"Q{q}" for q in sorted(flags))]
        else:
            row += [", ".join(f"Q{q}" for q in sorted(flags))]
        ws.append(row)

        row_num = ws.max_row
        for q in flags:
            col_idx = 2 + q
            ws.cell(row=row_num, column=col_idx).fill = yellow

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 10

    wb.save(out_xlsx)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--batch":
        out_xlsx = sys.argv[2]
        images = sys.argv[3:]
        batch_to_excel(images, out_xlsx)
        print(f"Saved results for {len(images)} sheet(s) to {out_xlsx}")
    else:
        img_path = sys.argv[1]
        debug_out = sys.argv[2] if len(sys.argv) > 2 else None
        answers, flags, raw = detect_bubbles(img_path, debug_out)
        print(json.dumps({"answers": answers, "flags": flags}, indent=2))
