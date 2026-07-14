#!/usr/bin/env python3
"""
backend-2/api_server.py
FastAPI backend — files(9) production pipeline.

Pipeline (scan_omr.py):
  align_sheet -> grid_correct -> extract_bubble_features (multi-metric)
  -> per_sheet_threshold (Otsu x k-means cross-validated)
  -> classify_question (per-question z-score relative comparison)

Key capabilities:
  - No per-sheet circle detection. Template loaded once at startup.
  - Adaptive per-sheet threshold, never a hardcoded global constant.
  - Image-quality estimation (low-res / heavy-JPEG / blur) with
    principled evidence-bar relaxation for mobile/WhatsApp photos.
  - Answers: OK / BLANK / MULTI / REVIEW — never silently guessed.
  - /api/calibrate to regenerate template.json from a new reference sheet.

Run:
    uvicorn api_server:app --reload --port 8001
"""

import asyncio
import json
import os
import tempfile
import time
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path
from threading import Lock
from typing import Dict, List, Optional

import openpyxl
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    HRFlowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

import numpy as np

from omr_engine import load_template, scan_image


# ── JSON sanitizer for numpy types ───────────────────────────────────────────
# The template scanner (omr_engine.scan_image) returns numpy scalar types
# (np.bool_, np.integer, np.floating) inside the per-question results and the
# align/image-quality metadata. FastAPI's default response serialization
# (jsonable_encoder) cannot encode these - e.g. np.bool_ raises
# "'numpy.bool_' object is not iterable" - which surfaces as an opaque HTTP 500
# AFTER the scan already succeeded. Convert everything to native Python types
# before returning so the mobile frontend always gets clean JSON.
def _json_safe(obj):
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(v) for v in obj]
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, np.floating):
        return float(obj)
    if isinstance(obj, np.ndarray):
        return _json_safe(obj.tolist())
    return obj


# ── App setup ──────────────────────────────────────────────────────────────
app = FastAPI(title="OMR Scanner API v2 (template-based)", version="2.0")

_raw_origins = os.environ.get(
    "ALLOWED_ORIGINS",
    "http://localhost:3000,http://127.0.0.1:3000"
)
_allow_origins = [o.strip() for o in _raw_origins.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allow_origins,
    allow_origin_regex=r"https://.*\.netlify\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Template (loaded once at startup, re-loadable via /api/calibrate) ──────
_TEMPLATE_PATH = Path(os.environ.get("TEMPLATE_PATH",
                                     Path(__file__).parent / "template.json"))
_template: dict = {}

@app.on_event("startup")
def _load_template_on_startup():
    global _template
    if _TEMPLATE_PATH.exists():
        _template = load_template(_TEMPLATE_PATH)
        print(f"✅ Template loaded: {len(_template.get('bubbles', {}))} bubbles "
              f"from {_TEMPLATE_PATH}")
    else:
        print(f"⚠️  template.json not found at {_TEMPLATE_PATH}. "
              "POST /api/calibrate to generate one before scanning.")


# ── Persistent answer key ─────────────────────────────────────────────────
_DATA_DIR = Path(os.environ.get("DATA_DIR", Path(__file__).parent / "data"))
_DATA_DIR.mkdir(parents=True, exist_ok=True)
_KEY_FILE = _DATA_DIR / "saved_answer_key.json"
_saved_key: dict = {}

if _KEY_FILE.exists():
    try:
        _saved_key = json.loads(_KEY_FILE.read_text())
        print(f"✅ Saved answer key loaded ({len(_saved_key)} questions)")
    except Exception as e:
        print(f"⚠️  Could not load saved answer key: {e}")

# ── Mobile scan directories ─────────────────────────────────────────────────
_MOBILE_FRONTEND_DIR = Path(__file__).parent.parent / "frontend" / "mobile"
_MOBILE_UPLOAD_DIR = Path(__file__).parent / "mobile_uploads"
_MOBILE_UPLOAD_DIR.mkdir(exist_ok=True)

# ── Mobile scan session store ───────────────────────────────────────────────
class MobileSessionStore:
    def __init__(self):
        self._lock = Lock()
        self._sessions: Dict[str, Dict] = {}

    def create(self) -> str:
        sid = uuid.uuid4().hex[:12]
        with self._lock:
            self._sessions[sid] = {
                "id": sid,
                "created_at": time.time(),
                "sheets": [],
            }
        return sid

    def get(self, sid: str) -> Dict:
        with self._lock:
            s = self._sessions.get(sid)
        if s is None:
            raise HTTPException(404, f"Unknown session {sid}")
        return s

    def add_result(self, sid: str, summary: Dict):
        with self._lock:
            session = self._sessions.setdefault(
                sid, {"id": sid, "created_at": time.time(), "sheets": []}
            )
            session["sheets"].append(summary)

    def stats(self, sid: str) -> Dict:
        session = self.get(sid)
        sheets = session["sheets"]
        n = len(sheets)
        if n == 0:
            return {"session_id": sid, "scanned": 0}
        avg_conf = sum(s.get("avg_confidence", 0) for s in sheets) / n
        needs_retake = sum(1 for s in sheets if s.get("retake_recommended", False))
        total_flagged = sum(s.get("flagged_count", 0) for s in sheets)
        return {
            "session_id": sid,
            "scanned": n,
            "avg_confidence": round(avg_conf, 1),
            "sheets_needing_retake": needs_retake,
            "total_flagged_questions": total_flagged,
            "sheets": sheets,
        }

mobile_sessions = MobileSessionStore()

OPT_MAP = {"1": "A", "2": "B", "3": "C", "4": "D",
           1: "A", 2: "B", 3: "C", 4: "D"}


def _normalize_key(answers: dict) -> dict:
    """Normalize to {str_q_num: letter} e.g. {"1": "A", "2": "B"}."""
    out = {}
    for k, v in answers.items():
        q = str(k).lstrip("qQ")
        out[q] = OPT_MAP.get(v, str(v).upper())
    return out


def _persist_key(answers: dict):
    _saved_key.clear()
    _saved_key.update(answers)
    try:
        _KEY_FILE.write_text(json.dumps(answers))
    except Exception as e:
        print(f"⚠️  Could not persist answer key: {e}")


# ── In-memory session store ───────────────────────────────────────────────
class Session:
    def __init__(self, sid: str):
        self.id = sid
        self.answer_key: Optional[dict] = None
        self.sheets: List[dict] = []
        self.results: List[dict] = []
        self.status = "created"
        self.created_at = datetime.now()


_sessions: Dict[str, Session] = {}


def _get_session(sid: str) -> Session:
    if sid not in _sessions:
        raise HTTPException(404, "Session not found")
    return _sessions[sid]


# ── Helper: score one sheet against the answer key ────────────────────────
def _score(questions: list, answer_key: dict) -> tuple:
    """Returns (score, total, detailed_answers dict, flags dict)."""
    score = 0
    detailed = {}
    flags = {}
    total = len(answer_key)

    for item in questions:
        q = str(item["q"])
        status = item["status"]
        option = item["option"]
        correct = answer_key.get(q)
        if correct is None:
            continue

        marked = ""
        flag = None
        is_correct = False

        if status == "MULTI":
            marked = "MULTI"
            flag = "multi_mark"
        elif status == "REVIEW":
            marked = OPT_MAP.get(option, "") if option else ""
            flag = "review"
        elif status == "BLANK":
            marked = ""
        else:  # OK
            marked = OPT_MAP.get(option, "") if option else ""
            is_correct = bool(marked) and marked == correct

        if is_correct:
            score += 1
        if flag:
            flags[q] = flag

        detailed[q] = {
            "marked": marked,
            "correct": correct,
            "is_correct": is_correct,
            "status": status,
            "flag": flag,
        }

    return score, total, detailed, flags


# ─────────────────────────────────────────────────────────────────────────
#  Routes
# ─────────────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"message": "OMR Scanner API v2", "status": "running",
            "template_loaded": bool(_template)}


@app.get("/health")
def health():
    return {"status": "ok", "template_loaded": bool(_template)}


# ── Template management ───────────────────────────────────────────────────

@app.post("/api/calibrate")
async def calibrate(file: UploadFile = File(...)):
    """
    Upload a clean reference sheet to regenerate template.json.
    Runs calibrate_template.py logic inline so the server doesn't need to
    shell out. The new template is loaded immediately into memory.
    """
    # Lazy import — calibrate_template lives in the same package
    from calibrate_template import build_template  # noqa: PLC0415

    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        new_template = build_template(
            tmp_path,
            out_path=str(_TEMPLATE_PATH),
            debug_path=str(_DATA_DIR / "debug_template_overlay.jpg"),
        )
        global _template
        _template = new_template
        return {
            "status": "ok",
            "bubbles": len(_template["bubbles"]),
            "radius": _template["radius"],
        }
    except Exception as e:
        raise HTTPException(400, str(e))
    finally:
        os.unlink(tmp_path)


@app.get("/api/template")
def get_template():
    if not _template:
        raise HTTPException(404, "No template loaded. POST /api/calibrate first.")
    return {
        "canon_w": _template["canon_w"],
        "canon_h": _template["canon_h"],
        "radius": _template["radius"],
        "n_bubbles": len(_template["bubbles"]),
    }


# ── Answer key ───────────────────────────────────────────────────────────

@app.get("/api/answer-key")
def get_saved_key():
    if not _saved_key:
        raise HTTPException(404, "No saved answer key")
    frontend = {f"q{k}": v for k, v in _saved_key.items()}
    return {"answers": frontend}


@app.post("/api/answer-key/save-manual")
def save_key_global(payload: Dict):
    answers = _normalize_key(payload.get("answers", {}))
    if not answers:
        raise HTTPException(400, "No answers provided")
    _persist_key(answers)
    return {"status": "saved", "total_questions": len(answers)}


@app.post("/api/answer-key/scan")
async def scan_key_from_image(file: UploadFile = File(...)):
    """Extract answer key from a scanned reference sheet."""
    if not _template:
        raise HTTPException(503, "Template not loaded. POST /api/calibrate first.")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        result = scan_image(tmp_path, _template)
    except Exception as e:
        raise HTTPException(400, str(e))
    finally:
        os.unlink(tmp_path)

    # Only keep OK answers; flag anything else for review
    answers = {}
    flags = {}
    for item in result["questions"]:
        q = str(item["q"])
        if item["status"] == "OK" and item["option"] is not None:
            answers[q] = OPT_MAP[item["option"]]
        else:
            flags[q] = item["status"]

    _persist_key(answers)
    return {"status": "ok", "answers": answers, "flags": flags,
            "meta": result["meta"]}


# ── Sessions ──────────────────────────────────────────────────────────────

@app.post("/api/session")
def create_session():
    sid = f"sess_{uuid.uuid4().hex[:8]}"
    _sessions[sid] = Session(sid)
    return {"session_id": sid, "status": "created"}


@app.get("/api/session/{sid}/status")
def session_status(sid: str):
    s = _get_session(sid)
    return {
        "status": s.status,
        "total_sheets": len(s.sheets),
        "processed": sum(1 for x in s.sheets if x["status"] == "DONE"),
        "queued":    sum(1 for x in s.sheets if x["status"] == "QUEUED"),
        "processing":sum(1 for x in s.sheets if x["status"] == "PROCESSING"),
        "errors":    sum(1 for x in s.sheets if x["status"] == "ERROR"),
    }


# ── Answer key per session ────────────────────────────────────────────────

@app.post("/api/session/{sid}/answer-key/manual")
def set_key_manual(sid: str, answer_key: Dict):
    s = _get_session(sid)
    answers = _normalize_key(answer_key.get("answers", {}))
    s.answer_key = answers
    s.status = "answer_key_set"
    _persist_key(answers)
    return {"status": "success", "answers": answers}


@app.post("/api/session/{sid}/answer-key/use-saved")
def use_saved_key(sid: str):
    s = _get_session(sid)
    if not _saved_key:
        raise HTTPException(404, "No saved answer key")
    s.answer_key = dict(_saved_key)
    s.status = "answer_key_set"
    frontend = {f"q{k}": v for k, v in s.answer_key.items()}
    return {"status": "success", "answers": frontend}


@app.post("/api/session/{sid}/answer-key")
async def set_key_from_image(sid: str, file: UploadFile = File(...)):
    """Extract answer key by scanning an uploaded reference sheet image."""
    s = _get_session(sid)
    if not _template:
        raise HTTPException(503, "Template not loaded. POST /api/calibrate first.")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        result = scan_image(tmp_path, _template)
    except Exception as e:
        raise HTTPException(400, str(e))
    finally:
        os.unlink(tmp_path)

    answers = {}
    flags = {}
    for item in result["questions"]:
        q = str(item["q"])
        if item["status"] == "OK" and item["option"] is not None:
            answers[q] = OPT_MAP[item["option"]]
        else:
            flags[q] = item["status"]

    s.answer_key = answers
    s.status = "answer_key_set"
    _persist_key(answers)
    return {"status": "success", "answers": answers, "flags": flags}


@app.get("/api/session/{sid}/answer-key")
def get_session_key(sid: str):
    s = _get_session(sid)
    if not s.answer_key:
        raise HTTPException(404, "Answer key not set")
    return {"answers": s.answer_key}


# ── Sheet upload ──────────────────────────────────────────────────────────

@app.post("/api/session/{sid}/sheets")
async def upload_sheets(
    sid: str,
    files: List[UploadFile] = File(...),
    names: Optional[str] = Form(None),
):
    s = _get_session(sid)
    if not s.answer_key:
        raise HTTPException(400, "Answer key not set")
    if not _template:
        raise HTTPException(503, "Template not loaded. POST /api/calibrate first.")

    name_list = [n.strip() for n in names.split(",")] if names else []
    sheet_ids = []

    for idx, f in enumerate(files):
        sheet_id = f"sheet_{uuid.uuid4().hex[:8]}"
        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
            tmp.write(await f.read())
            tmp_path = tmp.name

        student = (name_list[idx] if idx < len(name_list) and name_list[idx]
                   else f"Student_{len(s.sheets) + 1}")

        s.sheets.append({
            "id": sheet_id,
            "filename": f.filename,
            "path": tmp_path,
            "name": student,
            "status": "QUEUED",
        })
        sheet_ids.append(sheet_id)

    s.status = "processing"
    return {"status": "success", "sheet_ids": sheet_ids, "total_sheets": len(sheet_ids)}


# ── SSE processing stream ─────────────────────────────────────────────────

@app.get("/api/session/{sid}/progress")
async def process_sheets(sid: str):
    s = _get_session(sid)

    async def stream():
        pending = [sh for sh in s.sheets if sh["status"] not in ("DONE", "ERROR")]
        if not pending:
            yield f"data: {json.dumps({'type': 'BATCH_COMPLETE', 'totalProcessed': len(s.results)})}\n\n"
            return

        for sheet in pending:
            sheet_id = sheet["id"]
            sheet["status"] = "PROCESSING"
            yield f"data: {json.dumps({'type': 'PROCESSING', 'sheet_id': sheet_id, 'student_id': sheet['name']})}\n\n"

            try:
                result = scan_image(sheet["path"], _template)
                score, total, detailed, flags = _score(result["questions"], s.answer_key)

                confidence = max(0.0, 1.0 - len(flags) * 0.1)
                pct = (score / total * 100) if total else 0

                rec = {
                    "id": sheet["name"],
                    "name": sheet["name"],
                    "filename": sheet["filename"],
                    "score": score,
                    "total": total,
                    "percentage": pct,
                    "answers": detailed,
                    "flags": flags,
                    "confidence": confidence,
                    "omr_meta": result["meta"],
                    "timestamp": datetime.now().isoformat(),
                }
                s.results.append(rec)
                sheet["status"] = "DONE"
                sheet["result"] = rec

                yield f"data: {json.dumps({'type': 'DONE', 'sheet_id': sheet_id, 'status': 'DONE', 'score': {'correct': score, 'total': total}, 'confidence': confidence, 'student_id': sheet['name']})}\n\n"

            except Exception as e:
                sheet["status"] = "ERROR"
                sheet["error"] = str(e)
                yield f"data: {json.dumps({'type': 'ERROR', 'sheet_id': sheet_id, 'error': str(e)})}\n\n"

            await asyncio.sleep(0.2)

        s.status = "complete"
        yield f"data: {json.dumps({'type': 'BATCH_COMPLETE', 'totalProcessed': len(s.results)})}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


# ── Results ───────────────────────────────────────────────────────────────

@app.get("/api/session/{sid}/results")
def get_results(sid: str):
    s = _get_session(sid)
    if not s.results:
        return {"results": [], "statistics": {}, "status": s.status}

    scores = [r["score"] for r in s.results]
    total_q = s.results[0]["total"]
    passing = total_q // 2

    stats = {
        "total_students": len(s.results),
        "average_score": round(sum(scores) / len(scores), 2),
        "highest_score": max(scores),
        "lowest_score": min(scores),
        "pass_count": sum(1 for sc in scores if sc >= passing),
        "fail_count": sum(1 for sc in scores if sc < passing),
        "pass_rate": round(sum(1 for sc in scores if sc >= passing) / len(scores) * 100, 1),
    }
    return {"results": s.results, "statistics": stats,
            "total_questions": total_q, "status": s.status}


@app.get("/api/session/{sid}/detection/{sheet_id}")
def get_detection(sid: str, sheet_id: str):
    s = _get_session(sid)
    sheet = next((x for x in s.sheets if x["id"] == sheet_id), None)
    if not sheet:
        raise HTTPException(404, "Sheet not found")
    if "result" not in sheet:
        raise HTTPException(400, "Sheet not processed yet")
    result = sheet["result"]
    questions = sorted(
        [
            {"q_no": int(q), **detail}
            for q, detail in result["answers"].items()
        ],
        key=lambda x: x["q_no"],
    )
    return {
        "student_id": result["name"],
        "filename": result["filename"],
        "total_score": result["score"],
        "out_of": result["total"],
        "omr_meta": result.get("omr_meta", {}),
        "questions": questions,
    }


# ── Excel export ──────────────────────────────────────────────────────────

@app.get("/api/session/{sid}/export/excel")
def export_excel(sid: str):
    s = _get_session(sid)
    if not s.results:
        raise HTTPException(400, "No results to export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ["Name", "Score", "Total", "Percentage", "Pass/Fail",
               "Multi-marks", "Reviews"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    total_q = s.results[0]["total"]
    passing = total_q // 2

    for row, r in enumerate(s.results, 2):
        pct = f"{r['percentage']:.1f}%"
        passed = r["score"] >= passing
        n_multi  = sum(1 for d in r["answers"].values() if d.get("flag") == "multi_mark")
        n_review = sum(1 for d in r["answers"].values() if d.get("flag") == "review")
        ws.cell(row, 1, r["name"])
        ws.cell(row, 2, r["score"])
        ws.cell(row, 3, total_q)
        ws.cell(row, 4, pct)
        cell = ws.cell(row, 5, "Pass" if passed else "Fail")
        cell.font = Font(color="006100" if passed else "9C0006")
        ws.cell(row, 6, n_multi)
        ws.cell(row, 7, n_review)

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"omr_results_{sid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── PDF export ────────────────────────────────────────────────────────────

_BLUE  = colors.HexColor('#2563EB')
_LBLUE = colors.HexColor('#DBEAFE')
_GREEN = colors.HexColor('#16A34A')
_RED   = colors.HexColor('#DC2626')
_LGREY = colors.HexColor('#F3F4F6')
_GOLD  = colors.HexColor('#F59E0B')

_SUBJECTS = [
    ("Intelligence",   1,  10),
    ("Science",       11,  20),
    ("Social Science",21,  30),
    ("Mathematics",   31,  40),
]


def _subject_score(answers: dict, start: int, end: int) -> int:
    return sum(
        1 for q in range(start, end + 1)
        if answers.get(str(q), {}).get("is_correct", False)
    )


def _student_pdf_elements(result: dict, session_id: str, styles) -> list:
    """Return ReportLab flowables for one student's page."""
    name    = result.get("name", "Unknown")
    answers = result.get("answers", {})
    total   = result.get("score", 0)
    out_of  = result.get("total", 40)

    elems = []
    elems.append(Paragraph(f"<b>{name}</b>", styles["Title"]))
    elems.append(Paragraph(
        f"Session: {session_id} &nbsp;&nbsp; Date: {datetime.now().strftime('%Y-%m-%d')}",
        styles["Normal"],
    ))
    elems.append(HRFlowable(width="100%", thickness=1, color=_BLUE))
    elems.append(Spacer(1, 0.15 * inch))

    # Subject breakdown
    subj_data = [["Subject", "Score", "Out of"]]
    for subj, s, e in _SUBJECTS:
        subj_data.append([subj, str(_subject_score(answers, s, e)), "10"])
    subj_data.append(["TOTAL", str(total), str(out_of)])

    t = Table(subj_data, colWidths=[2.8 * inch, 1 * inch, 1 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  _BLUE),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("BACKGROUND",   (0, -1), (-1, -1), _LBLUE),
        ("FONTNAME",     (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN",        (1, 0), (-1, -1),  "CENTER"),
        ("GRID",         (0, 0), (-1, -1),  0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, _LGREY]),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 0.2 * inch))

    elems.append(Paragraph("<b>Question-wise Responses</b>", styles["Normal"]))
    elems.append(Spacer(1, 0.08 * inch))

    for subj, s, e in _SUBJECTS:
        elems.append(Paragraph(f"<b>{subj} (Q{s}–Q{e})</b>", styles["Normal"]))
        header_row  = ["Q"]      + [str(q) for q in range(s, e + 1)]
        marked_row  = ["Marked"] + [answers.get(str(q), {}).get("marked", "-") or "-" for q in range(s, e + 1)]
        correct_row = ["Correct"]+ [answers.get(str(q), {}).get("correct", "-") or "-" for q in range(s, e + 1)]
        result_row  = ["Result"] + ["✓" if answers.get(str(q), {}).get("is_correct") else "✗" for q in range(s, e + 1)]

        qt = Table(
            [header_row, marked_row, correct_row, result_row],
            colWidths=[0.85 * inch] + [0.55 * inch] * 10,
        )
        tstyle = [
            ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), _LGREY),
            ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ]
        for col_i, q in enumerate(range(s, e + 1), start=1):
            c = _GREEN if answers.get(str(q), {}).get("is_correct") else _RED
            tstyle.append(("TEXTCOLOR", (col_i, 3), (col_i, 3), c))
            tstyle.append(("FONTNAME",  (col_i, 3), (col_i, 3), "Helvetica-Bold"))
        qt.setStyle(TableStyle(tstyle))
        elems.append(qt)
        elems.append(Spacer(1, 0.12 * inch))

    elems.append(PageBreak())
    return elems


@app.get("/api/session/{sid}/export/pdf")
def export_pdf(sid: str):
    """Per-student PDF with subject breakdown + question-wise responses."""
    s = _get_session(sid)
    if not s.results:
        raise HTTPException(400, "No results to export")

    styles = getSampleStyleSheet()
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
    )
    elements = []

    for result in s.results:
        elements.extend(_student_pdf_elements(result, sid, styles))

    # Summary page
    elements.append(Paragraph("<b>Session Results Summary</b>", styles["Title"]))
    elements.append(HRFlowable(width="100%", thickness=1, color=_BLUE))
    elements.append(Spacer(1, 0.2 * inch))

    hdr = ["Name"] + [s for s, _, _ in _SUBJECTS] + ["Total /40"]
    rows = [hdr]
    for r in s.results:
        row = [r.get("name", "")]
        for _, start, end in _SUBJECTS:
            row.append(str(_subject_score(r.get("answers", {}), start, end)))
        row.append(str(r.get("score", 0)))
        rows.append(row)

    sum_t = Table(rows, colWidths=[2 * inch, 0.9 * inch, 0.7 * inch, 1 * inch, 0.9 * inch, 0.9 * inch])
    sum_t.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), _BLUE),
        ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
        ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",          (1, 0), (-1, -1), "CENTER"),
        ("GRID",           (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _LGREY]),
        ("FONTSIZE",       (0, 0), (-1, -1), 9),
    ]))
    elements.append(sum_t)

    doc.build(elements)
    output.seek(0)
    fname = f"omr_results_{sid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.get("/api/export/all-students/pdf")
def export_all_students_pdf():
    """Single ranked table of all students across all active sessions."""
    all_results = [r for sess in _sessions.values() for r in sess.results]
    if not all_results:
        raise HTTPException(400, "No results available")

    styles = getSampleStyleSheet()
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
    )
    elements = [
        Paragraph("<b>All Students Results</b>", styles["Title"]),
        Paragraph(
            f"Total: {len(all_results)} &nbsp;&nbsp; Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Normal"],
        ),
        HRFlowable(width="100%", thickness=1, color=_BLUE),
        Spacer(1, 0.2 * inch),
    ]

    header = ["#", "Name", "Intel\n/10", "Sci\n/10", "Soc.Sci\n/10", "Maths\n/10", "Total\n/40"]
    sorted_r = sorted(all_results, key=lambda r: r.get("score", 0), reverse=True)
    table_data = [header]
    for i, r in enumerate(sorted_r, 1):
        ans = r.get("answers", {})
        row = [str(i), r.get("name", "")]
        for _, s, e in _SUBJECTS:
            row.append(str(_subject_score(ans, s, e)))
        row.append(str(r.get("score", 0)))
        table_data.append(row)

    tbl = Table(table_data,
                colWidths=[0.4*inch, 2.4*inch, 0.8*inch, 0.7*inch, 0.9*inch, 0.8*inch, 0.8*inch],
                repeatRows=1)
    tstyle = [
        ("BACKGROUND",     (0, 0), (-1, 0), _BLUE),
        ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
        ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("ALIGN",          (1, 0), (1, -1), "LEFT"),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTSIZE",       (0, 0), (-1, -1), 8.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _LGREY]),
    ]
    for rank_row in range(1, min(4, len(table_data))):
        tstyle.append(("BACKGROUND", (0, rank_row), (-1, rank_row), colors.HexColor("#FEF9C3")))
        tstyle.append(("FONTNAME",   (0, rank_row), (-1, rank_row), "Helvetica-Bold"))
    tbl.setStyle(TableStyle(tstyle))
    elements.append(tbl)

    doc.build(elements)
    output.seek(0)
    fname = f"all_students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── Mobile Scan Endpoints ─────────────────────────────────────────────────────

@app.post("/api/mobile/session")
def create_mobile_session():
    """Create a mobile scan session."""
    return {"session_id": mobile_sessions.create()}


@app.get("/api/mobile/session/{session_id}/stats")
def get_mobile_session_stats(session_id: str):
    """Get mobile session statistics."""
    return mobile_sessions.stats(session_id)


@app.post("/api/mobile/scan")
async def mobile_scan(
    image: UploadFile = File(...),
    session_id: Optional[str] = Form(None),
    sheet_label: Optional[str] = Form(None),
):
    """
    Mobile scan endpoint - processes a pre-validated capture from the mobile frontend.
    Uses template-based scanner.
    """
    if image.content_type not in ("image/jpeg", "image/png", "image/webp"):
        raise HTTPException(400, f"Unsupported content type: {image.content_type}")

    sheet_id = uuid.uuid4().hex[:10]
    suffix = Path(image.filename or "sheet.jpg").suffix or ".jpg"
    dest = _MOBILE_UPLOAD_DIR / f"{sheet_id}{suffix}"
    data = await image.read()
    if len(data) < 1024:
        raise HTTPException(400, "Uploaded image is empty or too small")
    dest.write_bytes(data)

    try:
        # Use template-based scanner
        result = scan_image(str(dest), _template)
        
        # Build summary for mobile frontend
        confidences = [q.get("confidence", 0) for q in result["questions"]]
        avg_conf = sum(confidences) / len(confidences) if confidences else 0.0
        
        flagged = [
            {
                "question": q.get("q"),
                "status": q.get("status"),
                "confidence": q.get("confidence"),
                "notes": q.get("notes"),
            }
            for q in result["questions"]
            if q.get("status") in ("REVIEW", "MULTI") or q.get("confidence", 0) < 50.0
        ]
        
        retake_recommended = avg_conf < 55.0 or len(flagged) > 10
        
        summary = {
            "avg_confidence": round(avg_conf, 1),
            "flagged_count": len(flagged),
            "flagged_questions": flagged,
            "retake_recommended": retake_recommended,
            "sheet_notes": result["meta"].get("sheet_notes", []),
            "align_quality": result["meta"].get("align_quality", {}),
            "image_quality": result["meta"].get("image_quality", {}),
        }
        
        mobile_result = _json_safe({
            "sheet_id": sheet_id,
            "sheet_label": sheet_label,
            "questions": result["questions"],
            **summary,
        })
        
    except Exception as exc:
        raise HTTPException(422, f"Could not process sheet: {exc}") from exc

    if session_id:
        mobile_sessions.add_result(
            session_id,
            _json_safe({
                "sheet_id": sheet_id,
                "sheet_label": sheet_label,
                "avg_confidence": mobile_result.get("avg_confidence", 0),
                "flagged_count": mobile_result.get("flagged_count", 0),
                "retake_recommended": mobile_result.get("retake_recommended", False),
            }),
        )

    return mobile_result


@app.get("/api/mobile/config")
def mobile_config():
    """Expose template configuration for mobile frontend."""
    return {
        "canon_w": _template.get("canon_w", 1400),
        "canon_h": _template.get("canon_h", 2200),
        "target_aspect": _template.get("canon_w", 1400) / _template.get("canon_h", 2200),
        "questions": 40,
        "options_per_question": 4,
        "template_available": True,
    }


@app.get("/api/mobile/health")
def mobile_health():
    """Health check for mobile frontend."""
    return {
        "status": "ok",
        "template_available": True,
        "questions": 40,
    }


# ── Static File Serving ─────────────────────────────────────────────────────────

# Serve mobile frontend
if _MOBILE_FRONTEND_DIR.exists():
    app.mount("/mobile", StaticFiles(directory=str(_MOBILE_FRONTEND_DIR), html=True), name="mobile")


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8001))
    uvicorn.run("api_server:app", host="0.0.0.0", port=port, reload=False)
