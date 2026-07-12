"""
grade.py
--------
Grades student OMR sheets against an answer key photo, using the EXACT
SAME engine as scan_omr.py (align.py -> template.json -> scan_sheet()).
There is no separate detection mechanism for the key: it's the same
physical form, so it's read the same way.

Grading policy for ambiguous student answers (no response received to the
clarifying question, so this is the sensible default - change the
constants below if you want different behavior):
  - BLANK / MULTI  -> counted WRONG (0 marks). Both are unambiguous under
    standard OMR exam rules: no answer, or more than one answer, is not a
    valid response.
  - REVIEW         -> graded using the system's best guess (the darkest
    bubble), but the question is flagged "Needs Verification" rather than
    silently trusted, and is NOT silently counted as correct or wrong
    without that flag. A human should glance at flagged rows before
    finalizing scores.

Usage:
    # 1. One-time: read and validate the answer key
    python3 grade.py --key answer_key_photo.jpg --check-key

    # 2. Grade a batch of student sheets against that key
    python3 grade.py --key answer_key_photo.jpg student1.jpg student2.jpg ... \\
        --out graded_results.xlsx --workers 4
"""
import argparse
import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

from scan_omr import (
    load_template, scan_sheet, _scan_one, N_QUESTIONS,
)
from concurrent.futures import ProcessPoolExecutor, as_completed

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("grade")

RESULT_FILLS = {
    "Wrong": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Needs Verification": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Correct": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}


def read_answer_key(key_image_path, template, debug_dir=None):
    """
    Reads the answer key sheet with the identical engine used for student
    sheets, then enforces the one extra rule a key must satisfy: EVERY
    question must resolve to a clean OK. There is no valid "correct answer"
    to grade against if the key itself is BLANK/MULTI/REVIEW somewhere -
    that's a hard stop, not a guess.

    Returns (key_dict, meta). Raises ValueError listing the exact
    questions that failed validation, so you know precisely what to fix
    (retake the photo, or darken/clarify those specific bubbles) before
    grading can proceed.
    """
    rows, meta = scan_sheet(key_image_path, template, debug_dir=debug_dir)
    bad = [r for r in rows if r["Status"] != "OK"]
    if bad:
        bad_qs = ", ".join(f"Q{r['Question']} ({r['Status']})" for r in bad)
        raise ValueError(
            f"Answer key is not fully resolvable - cannot grade against an ambiguous key.\n"
            f"Problem question(s): {bad_qs}\n"
            f"Retake the key photo (clearer/darker marks, better lighting) or fix those "
            f"bubbles, then re-run --check-key before grading students."
        )
    key = {r["Question"]: r["Selected_Option"] for r in rows}
    log.info("Answer key OK: all %d questions resolved cleanly.", N_QUESTIONS)
    return key, meta


def grade_rows(rows, key):
    """
    Diffs one student's scanned rows against the answer key.
    Returns (graded_rows, score, needs_verification_count).
    """
    graded = []
    score = 0
    needs_verification = 0
    for r in rows:
        q = r["Question"]
        correct_answer = key.get(q)
        status = r["Status"]

        if status in ("BLANK", "MULTI"):
            result = "Wrong"
        elif status == "REVIEW":
            result = "Correct" if r["Selected_Option"] == correct_answer else "Wrong"
            result = f"{result} (Needs Verification)"
            needs_verification += 1
        else:  # OK
            result = "Correct" if r["Selected_Option"] == correct_answer else "Wrong"

        if result.startswith("Correct"):
            score += 1

        graded.append({
            "Question": q,
            "Selected_Option": r["Selected_Option"],
            "Correct_Answer": correct_answer,
            "Result": result,
            "Confidence": r["Confidence"],
            "Notes": r["Notes"],
        })
    return graded, score, needs_verification


def _fill_for_result(result):
    if result.startswith("Wrong"):
        return RESULT_FILLS["Wrong"]
    if "Needs Verification" in result:
        return RESULT_FILLS["Needs Verification"]
    return RESULT_FILLS["Correct"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--key", required=True, help="Answer key sheet photo")
    ap.add_argument("students", nargs="*", help="Student sheet photo(s) to grade")
    ap.add_argument("--out", default="graded_results.xlsx")
    ap.add_argument("--debug-dir", default="grade_debug")
    ap.add_argument("--workers", type=int, default=1)
    ap.add_argument("--check-key", action="store_true",
                     help="Only validate the answer key, don't grade anything")
    args = ap.parse_args()

    template = load_template()

    try:
        key, key_meta = read_answer_key(args.key, template, debug_dir=args.debug_dir)
    except ValueError as e:
        log.error(str(e))
        return

    log.info("Answer key: %s", key)
    if args.check_key or not args.students:
        return

    tasks = [(p, template, args.debug_dir) for p in args.students]
    student_results = {}

    def handle(name, rows, meta, error):
        if error:
            log.error("[FAILED] %s: %s", name, error)
            student_results[name] = {"error": error}
            return
        graded, score, needs_verification = grade_rows(rows, key)
        student_results[name] = {
            "graded": graded, "score": score,
            "needs_verification": needs_verification, "meta": meta,
        }
        pct = round(100.0 * score / N_QUESTIONS, 1)
        flag = f" [{needs_verification} need manual verification]" if needs_verification else ""
        log.info("%s: %d/%d (%.1f%%)%s", name, score, N_QUESTIONS, pct, flag)

    if args.workers > 1:
        with ProcessPoolExecutor(max_workers=args.workers) as pool:
            futures = [pool.submit(_scan_one, t) for t in tasks]
            for fut in as_completed(futures):
                name, rows, meta, error = fut.result()
                handle(name, rows, meta, error)
    else:
        for t in tasks:
            name, rows, meta, error = _scan_one(t)
            handle(name, rows, meta, error)

    _write_excel(args.out, key, student_results)
    log.info("Wrote %s", args.out)


def _write_excel(out_path, key, student_results):
    summary_rows = []
    per_question_correct = {q: 0 for q in range(1, N_QUESTIONS + 1)}
    n_graded = 0

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, res in student_results.items():
            if "error" in res:
                summary_rows.append({"Sheet": name, "Score": f"FAILED: {res['error']}"})
                continue
            n_graded += 1
            for g in res["graded"]:
                if g["Result"].startswith("Correct"):
                    per_question_correct[g["Question"]] += 1

            pct = round(100.0 * res["score"] / N_QUESTIONS, 1)
            summary_rows.append({
                "Sheet": name, "Score": res["score"], "Total": N_QUESTIONS,
                "Percentage": pct, "Needs_Verification": res["needs_verification"],
                "Sheet_Notes": "; ".join(res["meta"]["sheet_notes"]),
            })

            df = pd.DataFrame(res["graded"])
            sheet_name = name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            result_col = df.columns.get_loc("Result") + 1
            for row_idx, result in enumerate(df["Result"], start=2):
                ws.cell(row=row_idx, column=result_col).fill = _fill_for_result(result)

        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)

        # Item analysis: how many students got each question right - useful
        # for spotting a badly-worded question or a wrong answer key entry.
        if n_graded:
            item_rows = [
                {"Question": q, "Correct_Answer": key[q],
                 "Num_Correct": per_question_correct[q], "Num_Graded": n_graded,
                 "Pct_Correct": round(100.0 * per_question_correct[q] / n_graded, 1)}
                for q in range(1, N_QUESTIONS + 1)
            ]
            pd.DataFrame(item_rows).to_excel(writer, sheet_name="Item Analysis", index=False)


if __name__ == "__main__":
    main()
