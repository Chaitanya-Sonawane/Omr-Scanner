"""
FastAPI backend for OMR Evaluation Web Application.
"""
import asyncio
import json
import os
import sys
import time
import uuid
from pathlib import Path
from threading import Lock
from typing import Any, Dict, Optional

import aiofiles
import numpy as np
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

import session as sess_store
import answer_key_store
import results_store

# Add omr directory to path for importing template-based scanner
sys.path.insert(0, str(Path(__file__).parent / "omr"))

# Try to import template-based scanner before other omr imports that may have dependency issues
try:
    from scan_omr import load_template, scan_sheet, N_QUESTIONS
    TEMPLATE_PATH = Path(__file__).parent / "omr" / "template.json"
    _template = load_template(str(TEMPLATE_PATH))
    TEMPLATE_AVAILABLE = True
except (ImportError, Exception) as e:
    print(f"Template-based scanner not available: {e}")
    TEMPLATE_AVAILABLE = False
    _template = None
    N_QUESTIONS = 40

# Import other omr modules (may have torch/torchvision dependencies)
try:
    from omr.omr_scanner import detect_bubbles
    from omr.report import generate_report, generate_summary_report
    from queue_processor import enqueue_sheets, subscribe_sse, unsubscribe_sse
    OMR_MODULES_AVAILABLE = True
except Exception as e:
    print(f"OMR modules not available (torch/torchvision dependency issue): {e}")
    OMR_MODULES_AVAILABLE = False

app = FastAPI(title="OMR Evaluator API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

ALLOWED_EXTS = {".png", ".jpg", ".jpeg", ".pdf"}
MAX_SHEETS = 50

# Mobile scan directories
MOBILE_FRONTEND_DIR = Path(__file__).parent.parent / "frontend" / "mobile"
MOBILE_UPLOAD_DIR = Path(__file__).parent / "mobile_uploads"
MOBILE_UPLOAD_DIR.mkdir(exist_ok=True)

# Mobile scan session store
class MobileSessionStore:
    def __init__(self):
        self._lock = Lock()
        self._sessions: dict[str, dict] = {}

    def create(self) -> str:
        sid = uuid.uuid4().hex[:12]
        with self._lock:
            self._sessions[sid] = {
                "id": sid,
                "created_at": time.time(),
                "sheets": [],
            }
        return sid

    def get(self, sid: str) -> dict:
        with self._lock:
            s = self._sessions.get(sid)
        if s is None:
            raise HTTPException(404, f"Unknown session {sid}")
        return s

    def add_result(self, sid: str, summary: dict):
        with self._lock:
            session = self._sessions.setdefault(
                sid, {"id": sid, "created_at": time.time(), "sheets": []}
            )
            session["sheets"].append(summary)

    def stats(self, sid: str) -> dict:
        session = self.get(sid)
        sheets = session["sheets"]
        n = len(sheets)
        if n == 0:
            return {"session_id": sid, "scanned": 0}
        avg_conf = sum(s.get("avg_confidence", 0) for s in sheets) / n
        needs_retake = sum(1 for s in sheets if s.get("retake_recommended", False))
        total_flagged = sum(s.get("flagged_count", 0) for s in sheets)
        return {
            "session_id": sid,
            "scanned": n,
            "avg_confidence": round(avg_conf, 1),
            "sheets_needing_retake": needs_retake,
            "total_flagged_questions": total_flagged,
            "sheets": sheets,
        }

mobile_sessions = MobileSessionStore()

# JSON sanitizer for numpy types
def _json_safe(obj):
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(v) for v in obj]
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, np.floating):
        return float(obj)
    if isinstance(obj, np.ndarray):
        return _json_safe(obj.tolist())
    return obj


# ── Helpers ──────────────────────────────────────────────────────────────────

def _check_session(session_id: str) -> Dict[str, Any]:
    try:
        return sess_store.load_session(session_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")


def _validate_ext(filename: str):
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTS:
        raise HTTPException(status_code=400, detail=f"File type not allowed: {ext}")


async def _save_upload(upload: UploadFile, dest: Path):
    dest.parent.mkdir(parents=True, exist_ok=True)
    async with aiofiles.open(dest, "wb") as f:
        content = await upload.read()
        await f.write(content)


# ── Routes ───────────────────────────────────────────────────────────────────

@app.post("/api/session")
def create_session():
    s = sess_store.create_session()
    return {"session_id": s["session_id"]}


@app.post("/api/session/{session_id}/answer-key")
async def upload_answer_key(session_id: str, file: UploadFile = File(...)):
    _check_session(session_id)
    _validate_ext(file.filename)

    ext = Path(file.filename).suffix.lower()
    dest = sess_store._session_dir(session_id) / f"answer_key_raw{ext}"
    await _save_upload(file, dest)

    try:
        if TEMPLATE_AVAILABLE:
            # Use template-based scanner from omr-scanner-fixed
            rows, meta = scan_sheet(str(dest), _template)
            OPTION_MAP = {1: "A", 2: "B", 3: "C", 4: "D"}
            answer_key = {}
            for r in rows:
                q_num = r.get("Question")
                answer = r.get("Answer")
                answer_key[q_num] = OPTION_MAP.get(answer, "") if answer is not None else ""
        elif OMR_MODULES_AVAILABLE:
            # Fallback to adaptive detector
            answers_raw, flags, _, _conf = detect_bubbles(str(dest))
            OPTION_MAP = {1: "A", 2: "B", 3: "C", 4: "D"}
            answer_key = {}
            for q in range(1, 41):
                opt = answers_raw.get(q)
                answer_key[q] = OPTION_MAP.get(opt, "") if opt is not None else ""
        else:
            raise HTTPException(status_code=503, detail="No OMR detection modules available")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Answer key scan failed: {e}")

    # Save to session AND persistent store
    session_fmt = {f"q{k}": v for k, v in answer_key.items()}
    sess_store.set_answer_key(session_id, session_fmt)
    answer_key_store.save(answer_key)

    return {"answers": session_fmt, "saved_permanently": True}


@app.get("/api/answer-key")
def get_global_answer_key():
    """Get the permanently stored answer key."""
    key = answer_key_store.load()
    if not key:
        raise HTTPException(status_code=404, detail="No answer key saved yet")
    return {"answers": {f"q{k}": v for k, v in key.items()}}


@app.post("/api/session/{session_id}/answer-key/use-saved")
def use_saved_answer_key(session_id: str):
    """Apply the permanently saved answer key to this session."""
    _check_session(session_id)
    key = answer_key_store.load()
    if not key:
        raise HTTPException(status_code=404, detail="No saved answer key found. Please upload one first.")
    session_fmt = {f"q{k}": v for k, v in key.items()}
    sess_store.set_answer_key(session_id, session_fmt)
    return {"answers": session_fmt, "message": "Saved answer key applied to this session"}


@app.get("/api/session/{session_id}/answer-key")
def get_answer_key(session_id: str):
    s = _check_session(session_id)
    if not s.get("answer_key"):
        raise HTTPException(status_code=404, detail="No answer key uploaded yet")
    return {"answers": s["answer_key"]}


@app.post("/api/session/{session_id}/answer-key/manual")
def submit_manual_answer_key(session_id: str, body: Dict[str, Any]):
    _check_session(session_id)
    answers = body.get("answers", {})
    if not answers or len(answers) != 40:
        raise HTTPException(status_code=400, detail="Must provide exactly 40 answers")
    sess_store.set_answer_key(session_id, answers)
    return {"answers": answers}


@app.post("/api/answer-key/save-manual")
def save_manual_answer_key_globally(body: Dict[str, Any]):
    """Save manually entered answer key to persistent store."""
    answers = body.get("answers", {})
    if not answers or len(answers) != 40:
        raise HTTPException(status_code=400, detail="Must provide exactly 40 answers")
    # Convert from session format {q1: A, ...} to store format {1: A, ...}
    store_format = {int(k.replace("q", "")): v for k, v in answers.items()}
    answer_key_store.save(store_format)
    return {"message": "Answer key saved permanently"}


@app.post("/api/session/{session_id}/sheets")
async def upload_sheets(session_id: str, files: list[UploadFile] = File(...),
                        names: str = ""):
    """Upload student OMR sheets. Optional 'names' param: comma-separated student names."""
    if not OMR_MODULES_AVAILABLE:
        raise HTTPException(status_code=503, detail="OMR detection modules not available due to dependency issues")
    
    s = _check_session(session_id)

    if not s.get("answer_key"):
        raise HTTPException(status_code=400, detail="Upload answer key before uploading sheets")

    existing_sheets = s.get("sheets", [])
    if len(existing_sheets) + len(files) > MAX_SHEETS:
        raise HTTPException(
            status_code=400,
            detail=f"This session already has {len(existing_sheets)} sheets. Maximum {MAX_SHEETS} per session."
        )

    name_list = [n.strip() for n in names.split(",")] if names.strip() else []

    # Offset sheet numbering by how many sheets already exist in the session
    # so a second batch never gets IDs that collide with the first batch.
    existing_count = len(s.get("sheets", []))

    sheet_metas = []
    for i, upload in enumerate(files, start=1):
        _validate_ext(upload.filename)
        sheet_id = f"{existing_count + i:03d}"
        dest = sess_store.sheet_upload_path(session_id, sheet_id, upload.filename)
        await _save_upload(upload, dest)

        student_name = name_list[i - 1] if i - 1 < len(name_list) else None

        meta = {
            "sheet_id": sheet_id,
            "filename": upload.filename,
            "raw_path": str(dest),
            "status": "QUEUED",
            "student_id": student_name,   # pre-fill with provided name if any
            "score": None,
            "error": None,
        }
        sess_store.add_sheet(session_id, meta)
        sheet_metas.append(meta)

    enqueue_sheets(session_id, sheet_metas)
    return {"enqueued": len(sheet_metas), "sheet_ids": [m["sheet_id"] for m in sheet_metas]}


@app.get("/api/session/{session_id}/status")
def get_status(session_id: str):
    s = _check_session(session_id)
    return {
        "session_id": session_id,
        "answer_key_ready": s.get("answer_key") is not None,
        "sheets": [
            {
                "sheet_id": sh["sheet_id"],
                "filename": sh["filename"],
                "status": sh["status"],
                "student_id": sh.get("student_id"),
                "score": sh.get("score"),
                "error": sh.get("error"),
                "confidence": sh.get("sheet_confidence"),
            }
            for sh in s.get("sheets", [])
        ],
    }


@app.get("/api/session/{session_id}/progress")
async def sse_progress(session_id: str):
    _check_session(session_id)
    q: asyncio.Queue = asyncio.Queue(maxsize=100)
    subscribe_sse(session_id, q)

    async def event_stream():
        try:
            while True:
                try:
                    event = await asyncio.wait_for(q.get(), timeout=20.0)
                    yield f"data: {json.dumps(event)}\n\n"
                    if event.get("type") == "BATCH_COMPLETE":
                        break
                except asyncio.TimeoutError:
                    # Keepalive every 20s — prevents proxy/browser from closing connection
                    yield "data: {\"type\":\"KEEPALIVE\"}\n\n"
        finally:
            unsubscribe_sse(session_id, q)

    return StreamingResponse(
        event_stream(), media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/session/{session_id}/results")
def get_results(session_id: str):
    s = _check_session(session_id)
    sheets_out = []
    for sh in s.get("sheets", []):
        raw = sess_store.load_detection(session_id, sh["sheet_id"])
        sheets_out.append({**sh, "detection": raw})
    return {
        "session_id": session_id,
        "answer_key": s.get("answer_key"),
        "sheets": sheets_out,
    }


@app.get("/api/session/{session_id}/report")
def get_report(session_id: str):
    s = _check_session(session_id)
    pdf_path = sess_store.report_path(session_id)
    sheets_with_detection = []
    for sh in s.get("sheets", []):
        raw = sess_store.load_detection(session_id, sh["sheet_id"])
        sheets_with_detection.append({**sh, "detection": raw or {}})
    try:
        generate_report(
            session_id=session_id,
            answer_key=s.get("answer_key") or {},
            sheets=sheets_with_detection,
            output_path=str(pdf_path),
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Report generation failed: {e}")
    return FileResponse(path=str(pdf_path), media_type="application/pdf",
                        filename=f"omr_report_{session_id[:8]}.pdf")


@app.get("/api/summary-report")
def get_summary_report():
    """Generate a PDF with all students across all sessions."""
    all_results = results_store.load_all()
    if not all_results:
        raise HTTPException(status_code=404, detail="No results stored yet")
    out_path = "data/summary_report.pdf"
    try:
        generate_summary_report(all_results, out_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Summary report failed: {e}")
    return FileResponse(path=out_path, media_type="application/pdf",
                        filename="omr_summary_report.pdf")


@app.get("/api/all-results")
def get_all_results():
    """Return all stored results."""
    return {"results": results_store.load_all()}


@app.get("/api/export/excel")
def export_excel():
    """Download all results as an Excel file."""
    import tempfile
    from omr.omr_scanner import batch_to_excel
    all_results = results_store.load_all()
    if not all_results:
        raise HTTPException(status_code=404, detail="No results to export")

    # Build a list of per-question answers from stored results
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    header = (["#", "Student ID", "File"] +
              [f"Q{i}" for i in range(1, 41)] +
              ["Intelligence/10", "Science/10", "Social/10", "Math/10", "Total/40", "Date"])
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    yellow = PatternFill("solid", start_color="FFFF00")
    green  = PatternFill("solid", start_color="C6EFCE")
    red    = PatternFill("solid", start_color="FFC7CE")

    for idx, r in enumerate(all_results, 1):
        q_map = {q["q_no"]: q for q in r.get("per_question", [])}
        sc = r.get("section_scores", {})
        row = (
            [idx, r.get("student_id", ""), r.get("filename", "")] +
            [q_map.get(q, {}).get("marked", "") for q in range(1, 41)] +
            [sc.get("intelligence", 0), sc.get("science", 0),
             sc.get("social", 0), sc.get("math", 0), sc.get("total", 0),
             r.get("timestamp", "")[:10]]
        )
        ws.append(row)
        row_num = ws.max_row
        for q in range(1, 41):
            qd = q_map.get(q, {})
            col = q + 3  # offset: #, student_id, file = 3 cols
            if qd.get("is_correct"):
                ws.cell(row=row_num, column=col).fill = green
            elif qd.get("marked"):
                ws.cell(row=row_num, column=col).fill = red

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 10
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return FileResponse(path=tmp.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename="omr_results.xlsx")


@app.delete("/api/session/{session_id}")
def delete_session(session_id: str):
    """Delete a session and all its data."""
    import shutil
    session_dir = sess_store._session_dir(session_id)
    if not session_dir.exists():
        raise HTTPException(status_code=404, detail="Session not found")
    shutil.rmtree(session_dir)
    return {"deleted": session_id}


@app.get("/api/session/{session_id}/detection/{sheet_id}")
def get_sheet_detection(session_id: str, sheet_id: str):
    """Get detection data including all marked answers for a sheet"""
    _check_session(session_id)
    raw = sess_store.load_detection(session_id, sheet_id)
    if raw is None:
        raise HTTPException(status_code=404, detail="Detection data not found")
    return raw


@app.get("/api/session/{session_id}/sheet/{sheet_id}/raw")
def get_raw_detection(session_id: str, sheet_id: str):
    _check_session(session_id)
    raw = sess_store.load_detection(session_id, sheet_id)
    if raw is None:
        raise HTTPException(status_code=404, detail="Detection data not found for this sheet")
    return raw


@app.get("/api/debug/grid-check")
async def debug_grid_check(session_id: str):
    """
    Debug endpoint: re-runs detection on the stored answer-key image and returns
    per-question bubble intensities + threshold so you can verify grid alignment.
    GET /api/debug/grid-check?session_id=<id>
    """
    import glob as _glob
    session_dir = sess_store._session_dir(session_id)
    # Find the raw answer key file
    candidates = list(session_dir.glob("answer_key_raw.*"))
    if not candidates:
        raise HTTPException(status_code=404, detail="No answer key image found for this session")
    img_path = str(candidates[0])

    try:
        # Use the updated OMR scanner for debug analysis
        answers, flags, raw_results, _conf = detect_bubbles(img_path)
        
        # Convert to format expected by frontend
        debug_questions = []
        for q_num in range(1, 41):
            bubble_intensities = raw_results.get(q_num, {})
            detected_answer = answers.get(q_num)
            flag = flags.get(q_num)
            
            # Calculate confidence based on gap between darkest and second darkest
            if bubble_intensities:
                sorted_vals = sorted(bubble_intensities.values())
                gap = sorted_vals[1] - sorted_vals[0] if len(sorted_vals) >= 2 else 0
                confidence = min(gap / 50.0, 1.0) if detected_answer else 0.0
            else:
                confidence = 0.0
                
            debug_questions.append({
                "q_no": q_num,
                "detected": detected_answer,
                "confidence": confidence,
                "intensities": bubble_intensities,
                "is_blank": detected_answer is None,
                "is_multi": False,  # Multi-marking would be detected in flags
                "flag": flag
            })
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return {
        "method": "adaptive_omr_scanner",
        "detection_info": "Uses Hough Circle Transform with adaptive grid calibration",
        "questions": debug_questions,
    }


# ── Mobile Scan Endpoints ─────────────────────────────────────────────────────

@app.post("/api/mobile/session")
def create_mobile_session():
    """Create a mobile scan session."""
    return {"session_id": mobile_sessions.create()}


@app.get("/api/mobile/session/{session_id}/stats")
def get_mobile_session_stats(session_id: str):
    """Get mobile session statistics."""
    return _json_safe(mobile_sessions.stats(session_id))


@app.post("/api/mobile/scan")
async def mobile_scan(
    image: UploadFile = File(...),
    session_id: Optional[str] = Form(None),
    sheet_label: Optional[str] = Form(None),
):
    """
    Mobile scan endpoint - processes a pre-validated capture from the mobile frontend.
    Uses template-based scanner if available, otherwise falls back to adaptive detector.
    """
    if image.content_type not in ("image/jpeg", "image/png", "image/webp"):
        raise HTTPException(400, f"Unsupported content type: {image.content_type}")

    sheet_id = uuid.uuid4().hex[:10]
    suffix = Path(image.filename or "sheet.jpg").suffix or ".jpg"
    dest = MOBILE_UPLOAD_DIR / f"{sheet_id}{suffix}"
    data = await image.read()
    if len(data) < 1024:
        raise HTTPException(400, "Uploaded image is empty or too small")
    dest.write_bytes(data)

    try:
        if TEMPLATE_AVAILABLE:
            # Use template-based scanner
            rows, meta = scan_sheet(str(dest), _template, debug_dir=str(MOBILE_UPLOAD_DIR / "debug"))
            
            # Build summary for mobile frontend
            confidences = [r.get("Confidence", 0) for r in rows]
            avg_conf = sum(confidences) / len(confidences) if confidences else 0.0
            
            flagged = [
                {
                    "question": r.get("Question"),
                    "status": r.get("Status"),
                    "confidence": r.get("Confidence"),
                    "notes": r.get("Notes"),
                }
                for r in rows
                if r.get("Status") in ("REVIEW", "MULTI") or r.get("Confidence", 0) < 50.0
            ]
            
            retake_recommended = avg_conf < 55.0 or len(flagged) > 10
            
            summary = {
                "avg_confidence": round(avg_conf, 1),
                "flagged_count": len(flagged),
                "flagged_questions": flagged,
                "retake_recommended": retake_recommended,
                "sheet_notes": meta.get("sheet_notes", []),
                "align_quality": meta.get("align_quality", {}),
                "image_quality": meta.get("image_quality", {}),
            }
            
            result = _json_safe({
                "sheet_id": sheet_id,
                "sheet_label": sheet_label,
                "questions": rows,
                **summary,
            })
        else:
            # Fallback to adaptive detector
            answers, flags, raw_results, confidence = detect_bubbles(str(dest))
            
            # Convert to mobile-friendly format
            questions = []
            for q_num in range(1, 41):
                questions.append({
                    "Question": q_num,
                    "Answer": answers.get(q_num),
                    "Status": flags.get(q_num, "OK"),
                    "Confidence": confidence,
                    "Notes": "",
                })
            
            avg_conf = confidence
            flagged_count = sum(1 for f in flags.values() if f in ("REVIEW", "MULTI"))
            retake_recommended = avg_conf < 55.0 or flagged_count > 10
            
            result = _json_safe({
                "sheet_id": sheet_id,
                "sheet_label": sheet_label,
                "questions": questions,
                "avg_confidence": round(avg_conf, 1),
                "flagged_count": flagged_count,
                "flagged_questions": [],
                "retake_recommended": retake_recommended,
            })
            
    except Exception as exc:
        raise HTTPException(422, f"Could not process sheet: {exc}") from exc

    if session_id:
        mobile_sessions.add_result(
            session_id,
            _json_safe({
                "sheet_id": sheet_id,
                "sheet_label": sheet_label,
                "avg_confidence": result.get("avg_confidence", 0),
                "flagged_count": result.get("flagged_count", 0),
                "retake_recommended": result.get("retake_recommended", False),
            }),
        )

    return JSONResponse(result)


@app.get("/api/mobile/config")
def mobile_config():
    """Expose template configuration for mobile frontend."""
    if TEMPLATE_AVAILABLE:
        return {
            "canon_w": _template["canon_w"],
            "canon_h": _template["canon_h"],
            "target_aspect": _template["canon_w"] / _template["canon_h"],
            "questions": N_QUESTIONS,
            "options_per_question": 4,
            "template_available": True,
        }
    else:
        return {
            "questions": 40,
            "options_per_question": 4,
            "template_available": False,
        }


@app.get("/api/mobile/health")
def mobile_health():
    """Health check for mobile frontend."""
    return {
        "status": "ok",
        "template_available": TEMPLATE_AVAILABLE,
        "questions": N_QUESTIONS if TEMPLATE_AVAILABLE else 40,
    }


# ── Static File Serving ─────────────────────────────────────────────────────────

# Serve mobile frontend
if MOBILE_FRONTEND_DIR.exists():
    app.mount("/mobile", StaticFiles(directory=str(MOBILE_FRONTEND_DIR), html=True), name="mobile")
