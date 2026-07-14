"""
Advanced OMR Scanner - Production Ready v2.1
Features:
- Two-pass verification system (intensity + filled-pixel validation)
- Multi-metric bubble analysis (mean intensity + fill ratio)
- Adaptive preprocessing for varying image quality
- Automatic perspective correction
- Multi-stage circle detection with fallback strategies
- Per-row adaptive thresholding with confidence scoring
- Robust handling of light marks, partial fills, and circled bubbles
- Smart block splitting resistant to middle-column interference
- Enhanced debug logging and visualization
"""
import cv2
import numpy as np
import json
import os
import sys
from typing import Tuple, Dict, Optional, List

# Debug logging
DEBUG_ENABLED = False

# --- Template-based reader dependencies -------------------------------
# The template-based reader (detect_bubbles) warps every photo into a
# fixed canonical space (align.py) and then samples 160 fixed bubble
# positions from template.json. This is far more reliable than blind
# HoughCircle+k-means grid guessing (kept below as _detect_bubbles_legacy
# and used only if alignment fails), because the bubble grid no longer has
# to be re-discovered from scratch on every (possibly heavily filled)
# sheet - only the outer border does.
_MODULE_DIR = os.path.dirname(os.path.abspath(__file__))
_TEMPLATE_PATH = os.path.join(_MODULE_DIR, "template.json")

try:
    from .align import align_sheet, load_image, CANON_W, CANON_H
except Exception:  # pragma: no cover - direct-script / non-package execution
    sys.path.insert(0, _MODULE_DIR)
    try:
        from align import align_sheet, load_image, CANON_W, CANON_H  # type: ignore
    except Exception:
        align_sheet = None  # type: ignore
        load_image = None  # type: ignore
        CANON_W, CANON_H = 1400, 2200

_TEMPLATE_CACHE = None


def _load_template():
    """Load and cache template.json (160 bubble coords in canonical space)."""
    global _TEMPLATE_CACHE
    if _TEMPLATE_CACHE is None:
        with open(_TEMPLATE_PATH) as f:
            _TEMPLATE_CACHE = json.load(f)
    return _TEMPLATE_CACHE

def log(msg: str):
    """Debug logging function"""
    if DEBUG_ENABLED:
        print(f"[OMR-DEBUG] {msg}")


def _try_hough(gray_blur, min_r, max_r, param2):
    circles = cv2.HoughCircles(
        gray_blur, cv2.HOUGH_GRADIENT, dp=1, minDist=18,
        param1=50, param2=param2, minRadius=min_r, maxRadius=max_r
    )
    return circles


def _preprocess_image(gray):
    """Advanced preprocessing: handles shadows, uneven lighting, low contrast."""
    # Normalize contrast
    if gray.max() > gray.min():
        gray = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
    
    # Remove shadows with morphological background subtraction
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (15, 15))
    background = cv2.morphologyEx(gray, cv2.MORPH_CLOSE, kernel)
    gray = cv2.subtract(background, gray)
    gray = 255 - gray
    
    # Adaptive histogram equalization
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    gray = clahe.apply(gray)
    
    return gray


def _detect_circles(gray):
    """Multi-stage circle detection with preprocessing and fallback strategies."""
    h, w = gray.shape
    r_est = max(8, min(30, w // 80))
    min_r = max(6, r_est - 6)
    max_r = r_est + 10
    
    # Stage 1: Standard detection on median-blurred image
    blur = cv2.medianBlur(gray, 5)
    for param2 in [22, 18, 14, 10, 8]:
        circles = _try_hough(blur, min_r, max_r, param2)
        if circles is not None and len(circles[0]) >= 8:
            return circles, blur
    
    # Stage 2: Enhanced preprocessing for difficult images
    enhanced = _preprocess_image(gray)
    blur2 = cv2.medianBlur(enhanced, 5)
    for param2 in [20, 16, 12, 8]:
        circles = _try_hough(blur2, min_r, max_r, param2)
        if circles is not None and len(circles[0]) >= 8:
            return circles, blur2
    
    # Stage 3: Bilateral filter for noise reduction while preserving edges
    bilateral = cv2.bilateralFilter(gray, 9, 75, 75)
    blur3 = cv2.medianBlur(bilateral, 3)
    for param2 in [18, 14, 10, 7]:
        circles = _try_hough(blur3, min_r, max_r, param2)
        if circles is not None and len(circles[0]) >= 8:
            return circles, blur3

    # Stage 4: Wider radius range + aggressive CLAHE for mobile/WhatsApp photos
    clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))
    eq = clahe.apply(gray)
    _, bw = cv2.threshold(eq, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    blur4 = cv2.medianBlur(bw, 3)
    r_lo = max(4, r_est - 8)
    r_hi = r_est + 14
    for param2 in [15, 10, 7, 5]:
        circles = _try_hough(blur4, r_lo, r_hi, param2)
        if circles is not None and len(circles[0]) >= 8:
            return circles, blur4
    
    return None, blur


def _safe_kmeans(data, k, criteria, attempts=10):
    n = len(data)
    if n < k:
        return None, None
    _, labels, centers = cv2.kmeans(data, k, None, criteria, attempts, cv2.KMEANS_PP_CENTERS)
    return labels, centers


def _compute_row_y(all_pts):
    """
    Compute the 20 shared row y-positions from ALL detected bubble points
    on the sheet (both option blocks pooled together). Pooling roughly
    doubles the points per row cluster, making row positions far more stable.
    """
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)
    all_ys = all_pts[:, 1].reshape(-1, 1).astype(np.float32)
    n_rows_k = min(20, len(all_ys))
    row_labels, row_centers = _safe_kmeans(all_ys, n_rows_k, criteria)

    if row_labels is None:
        y_min, y_max = float(all_ys.min()), float(all_ys.max())
        row_y_sorted = np.linspace(y_min, y_max, 20).tolist()
    else:
        cluster_y = sorted(float(c) for c in row_centers.flatten())
        if len(cluster_y) < 20:
            diffs = np.diff(cluster_y)
            diffs = diffs[diffs > 3]
            pitch = float(np.median(diffs)) if len(diffs) else 30.0
            row_y_sorted = list(cluster_y)
            while len(row_y_sorted) < 20:
                row_y_sorted.append(row_y_sorted[-1] + pitch)
        else:
            row_y_sorted = cluster_y

    # Snap any row deviating from the linear trend back onto the fitted line
    row_idx_arr = np.arange(20)
    row_y_arr = np.array(row_y_sorted, dtype=np.float64)
    diffs = np.diff(row_y_arr)
    diffs = diffs[diffs > 3]
    est_pitch = float(np.median(diffs)) if len(diffs) else 30.0
    for _ in range(2):
        ay, by_ = np.polyfit(row_idx_arr, row_y_arr, 1)
        fitted = ay * row_idx_arr + by_
        resid = row_y_arr - fitted
        bad = np.abs(resid) > max(est_pitch * 0.4, 8)
        if not bad.any():
            break
        row_y_arr[bad] = fitted[bad]
    return row_y_arr.tolist()


def _build_grid(block_pts, gray, radius_est, row_y_sorted):
    """
    Build a 20x4 grid of (x,y) sample positions for one answer block,
    given shared row y-positions from _compute_row_y.
    Uses per-column linear regression (x = a*row + b) to track perspective
    drift across rows, so angled/tilted photos are handled correctly.
    """
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)
    h, w = gray.shape

    n_cols = min(4, len(block_pts))
    bxs = block_pts[:, 0].reshape(-1, 1).astype(np.float32)
    col_labels, col_centers = _safe_kmeans(bxs, n_cols, criteria)
    if col_labels is None or n_cols < 4:
        x_min_b, x_max_b = float(bxs.min()), float(bxs.max())
        col_centers_list = np.linspace(x_min_b, x_max_b, 4).tolist()
        col_labels = np.zeros((len(block_pts), 1), dtype=np.int32)
        cluster_to_opt = {0: 0}
    else:
        # Sort column centers left to right
        col_centers_flat = col_centers.flatten()
        cluster_order = np.argsort(col_centers_flat)
        col_centers_list = sorted(col_centers_flat.tolist())
        
        # Map each cluster ID to its position (0=leftmost, 3=rightmost)
        # cluster_order tells us which cluster ID is in which position
        # We want: cluster_id -> option_index (0-3 from left to right)
        cluster_to_opt = {int(cluster_order[i]): i for i in range(len(cluster_order))}

    # Group detected points by column (option index)
    col_pts_by_opt = {opt_idx: [] for opt_idx in range(4)}
    for i in range(len(block_pts)):
        opt_idx = cluster_to_opt.get(int(col_labels[i][0]), -1)
        if 0 <= opt_idx < 4:
            col_pts_by_opt[opt_idx].append(block_pts[i])

    # For each column, fit a linear model x(row_y) = a*row_y + b so that
    # sampling tracks perspective drift across rows instead of using a fixed x.
    row_y_arr = np.array(row_y_sorted, dtype=np.float64)

    col_grid = {}
    for opt_idx in range(4):
        pts_list = col_pts_by_opt[opt_idx]
        if len(pts_list) >= 2:
            px = np.array([p[0] for p in pts_list], dtype=np.float64)
            py = np.array([p[1] for p in pts_list], dtype=np.float64)
            # find which row each detected point belongs to (nearest row_y)
            row_indices = np.array([
                int(np.argmin(np.abs(row_y_arr - yp))) for yp in py
            ])
            if len(np.unique(row_indices)) >= 2:
                a, b = np.polyfit(row_indices, px, 1)
            else:
                a, b = 0.0, float(np.median(px))
        else:
            # fallback: use column center from k-means
            a, b = 0.0, col_centers_list[opt_idx] if opt_idx < len(col_centers_list) else 0.0

        grid_col = []
        for row_idx in range(20):
            gx = int(round(a * row_idx + b))
            gy = int(round(row_y_sorted[row_idx]))
            gx = max(radius_est, min(w - radius_est - 1, gx))
            gy = max(radius_est, min(h - radius_est - 1, gy))
            grid_col.append((gx, gy))
        col_grid[opt_idx] = grid_col

    return col_grid


def _auto_rotate_image(img, gray):
    """Auto-rotate image if detected as upside-down or severely tilted."""
    # Detect edges
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)
    lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=100, minLineLength=100, maxLineGap=10)
    
    if lines is not None and len(lines) > 5:
        angles = []
        for line in lines[:50]:  # Check first 50 lines
            x1, y1, x2, y2 = line[0]
            angle = np.arctan2(y2 - y1, x2 - x1) * 180 / np.pi
            angles.append(angle)
        
        median_angle = np.median(angles)
        
        # If significantly tilted, rotate
        if abs(median_angle) > 2 and abs(median_angle) < 45:
            h, w = img.shape[:2]
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, median_angle, 1.0)
            img = cv2.warpAffine(img, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    return img, gray


# ---------------------------------------------------------------------
# Template-based reader (PRIMARY). Aligns the sheet to canonical space,
# then reads 160 fixed bubble positions with a small local darkest-disk
# search that absorbs residual perspective/scale drift, and decides each
# question by RELATIVE darkness (the marked bubble is much darker than its
# 3 siblings) rather than a brittle global cut.
# ---------------------------------------------------------------------
N_QUESTIONS = 40
N_OPTIONS = 4

# Reading tunables (canonical space, darkness measured as 255-mean on the
# illumination-normalised warp, so 0 == paper-white, high == solid mark).
_MIN_FILL_DARK = 48.0        # a bubble must be at least this dark to count as a mark
_STRONG_FILL_DARK = 82.0     # unambiguously filled regardless of siblings
_REL_MARGIN = 20.0           # darkest must beat the mean of the other 3 by this
_MIN_GAP = 9.0               # darkest must beat the 2nd darkest by this
_MULTI_GAP = 9.0             # 2nd bubble this close to darkest -> possible double mark


def _normalise_illumination(warped_bgr):
    """Flatten uneven phone-camera lighting so a fixed darkness scale is
    meaningful across the whole sheet (divide by a heavily-blurred copy)."""
    gray = cv2.cvtColor(warped_bgr, cv2.COLOR_BGR2GRAY)
    bg = cv2.GaussianBlur(gray, (0, 0), sigmaX=40)
    norm = cv2.divide(gray.astype(np.float32), bg.astype(np.float32) + 1e-6) * 160.0
    return np.clip(norm, 0, 255).astype(np.uint8)


def _best_darkness(norm, cx, cy, sample_r, search, mask):
    """Darkness of the darkest disk found within +/-search px of (cx,cy).
    The local search absorbs the small residual drift left after the
    perspective warp, so a slightly mis-located template point still lands
    on the real bubble. Returns (darkness, x, y)."""
    h, w = norm.shape
    best_d, bx, by = -1.0, cx, cy
    for dy in range(-search, search + 1, 3):
        y = cy + dy
        if y - sample_r < 0 or y + sample_r + 1 > h:
            continue
        for dx in range(-search, search + 1, 3):
            x = cx + dx
            if x - sample_r < 0 or x + sample_r + 1 > w:
                continue
            patch = norm[y - sample_r:y + sample_r + 1, x - sample_r:x + sample_r + 1]
            vals = patch[mask == 255]
            d = 255.0 - float(vals.mean())
            if d > best_d:
                best_d, bx, by = d, x, y
    return best_d, bx, by


def detect_bubbles(img_path, debug_out=None):
    """
    Main detection function.

    PRIMARY path: perspective-align the photographed sheet into a fixed
    canonical space and read the 160 template bubble positions. Falls back
    to the legacy HoughCircle+k-means detector only if alignment machinery
    is unavailable or the warp is clearly untrustworthy.

    Returns:
        answers: Dict[int, Optional[int]] - question -> selected option (1-4) or None
        flags: Dict[int, str] - questions with detection warnings
        results: Dict[int, Dict[int, float]] - per-bubble mean intensity (low == dark)
        confidence_scores: Dict[int, int] - 0-100 confidence per question
    """
    if align_sheet is None or load_image is None:
        return _detect_bubbles_legacy(img_path, debug_out)

    try:
        tmpl = _load_template()
        img = load_image(img_path)
        warped, quality = align_sheet(img, out_size=(CANON_W, CANON_H))
    except Exception as e:
        log(f"Template align failed ({e}); using legacy detector")
        return _detect_bubbles_legacy(img_path, debug_out)

    # An untrustworthy border (whole-frame fallback) means the template
    # coordinates cannot be trusted; the legacy self-calibrating detector
    # copes better with an un-warped raw photo in that case.
    if quality.get("border_method") == "full_frame_fallback":
        log("Alignment fell back to full frame; using legacy detector")
        return _detect_bubbles_legacy(img_path, debug_out)

    radius = int(tmpl.get("radius", 35))
    sample_r = max(4, int(radius * 0.62))
    search = max(4, int(radius * 0.6))
    mask = np.zeros((2 * sample_r + 1, 2 * sample_r + 1), np.uint8)
    cv2.circle(mask, (sample_r, sample_r), sample_r, 255, -1)

    norm = _normalise_illumination(warped)
    bubbles = tmpl["bubbles"]

    # Sample every bubble (darkness + refined centre).
    dark = {}       # (q,o) -> darkness
    centre = {}     # (q,o) -> (x,y)
    for q in range(1, N_QUESTIONS + 1):
        for o in range(1, N_OPTIONS + 1):
            b = bubbles[f"{q}_{o}"]
            cx, cy = int(round(b["x"])), int(round(b["y"]))
            d, bx, by = _best_darkness(norm, cx, cy, sample_r, search, mask)
            dark[(q, o)] = d
            centre[(q, o)] = (bx, by)

    answers: Dict[int, Optional[int]] = {}
    flags: Dict[int, str] = {}
    results: Dict[int, Dict[int, float]] = {}
    confidence_scores: Dict[int, int] = {}

    for q in range(1, N_QUESTIONS + 1):
        opt_d = [(dark[(q, o)], o) for o in range(1, N_OPTIONS + 1)]
        # store mean intensity (255 - darkness) for API /raw compatibility
        results[q] = {o: max(0.0, 255.0 - dark[(q, o)]) for o in range(1, N_OPTIONS + 1)}

        opt_d.sort(reverse=True)
        d1, o1 = opt_d[0]
        d2, o2 = opt_d[1]
        others_mean = float(np.mean([opt_d[1][0], opt_d[2][0], opt_d[3][0]]))
        gap = d1 - d2

        strong = d1 >= _STRONG_FILL_DARK and gap >= (_MIN_GAP - 3)
        relative = (d1 >= _MIN_FILL_DARK and (d1 - others_mean) >= _REL_MARGIN and gap >= _MIN_GAP)
        is_filled = strong or relative

        if is_filled:
            answers[q] = o1
            # confidence from how dominant / dark the chosen mark is
            gap_conf = 45.0 * min(1.0, gap / 45.0)
            dark_conf = 35.0 * min(1.0, d1 / 140.0)
            confidence_scores[q] = int(min(100, 20 + gap_conf + dark_conf))

            # possible double-mark: another bubble almost as dark and also a real mark
            if d2 >= _MIN_FILL_DARK and gap < _MULTI_GAP:
                flags[q] = "multi_mark"
                confidence_scores[q] = max(25, confidence_scores[q] - 40)
            elif confidence_scores[q] < 45:
                flags[q] = "low_confidence"
        else:
            answers[q] = None
            confidence_scores[q] = 0
            # everything dark & similar -> smudged/over-filled row worth review
            if d1 >= _MIN_FILL_DARK and others_mean >= _MIN_FILL_DARK and gap < _MULTI_GAP:
                flags[q] = "row_smudged"

    if debug_out:
        dbg = cv2.cvtColor(norm, cv2.COLOR_GRAY2BGR)
        for q in range(1, N_QUESTIONS + 1):
            marked = answers.get(q)
            for o in range(1, N_OPTIONS + 1):
                x, y = centre[(q, o)]
                filled = (marked == o)
                col = (0, 200, 0) if filled else (0, 0, 220)
                cv2.circle(dbg, (x, y), radius, col, 3 if filled else 1)
                cv2.putText(dbg, f"{dark[(q, o)]:.0f}", (x - 18, y - radius - 3),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 120, 0), 1, cv2.LINE_AA)
        cv2.imwrite(debug_out, dbg)

    return answers, flags, results, confidence_scores


def _detect_bubbles_legacy(img_path, debug_out=None):
    """
    Legacy self-calibrating detector (HoughCircle + k-means grid). Retained
    as a fallback for when perspective alignment is unavailable/untrusted.

    Returns:
        answers: Dict[int, Optional[int]] - Question number -> selected option (1-4) or None
        flags: Dict[int, str] - Questions with detection warnings
        results: Dict[int, Dict[int, float]] - Raw intensity values for all bubbles
    """
    img = cv2.imread(img_path)
    if img is None:
        raise RuntimeError(f"Could not read image: {img_path}")
    orig = img.copy()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Auto-rotate if needed
    img, gray = _auto_rotate_image(img, gray)
    h, w = gray.shape

    raw_circles, gray_blur = _detect_circles(gray)
    if raw_circles is None:
        raise RuntimeError(
            f"No circles found in {img_path}. "
            "Check that the image shows a clear OMR sheet with visible bubbles."
        )
    circles = np.round(raw_circles[0]).astype(int)

    circles = [c for c in circles if 50 < c[0] < w - 50 and 200 < c[1] < h - 30]
    if len(circles) < 8:
        raise RuntimeError(f"Too few circles detected ({len(circles)}) after filtering edges.")

    # filter to circles near median radius
    radii = np.array([c[2] for c in circles])
    r_med = np.median(radii)
    circles = [c for c in circles if abs(c[2] - r_med) <= max(2, r_med * 0.35)]
    if len(circles) < 8:
        raise RuntimeError(f"Too few circles remain ({len(circles)}) after radius filtering.")

    # drop header debris (stamp/logo) using largest-gap strategy
    ys_all = sorted(c[1] for c in circles)
    n_total = len(ys_all)
    window = min(50, n_total)
    cut_y = None
    best_gap = 0
    for i in range(1, window):
        gap = ys_all[i] - ys_all[i - 1]
        if gap > 60 and gap > best_gap:
            # Only cut if enough circles remain after the cut (at least 60% of total)
            remaining = sum(1 for y in ys_all if y >= ys_all[i])
            if remaining >= max(8, n_total * 0.6):
                best_gap = gap
                cut_y = ys_all[i]
    if cut_y is not None:
        circles = [c for c in circles if c[1] >= cut_y]
    if len(circles) < 8:
        raise RuntimeError(f"Too few circles remain ({len(circles)}) after outlier removal.")

    pts = np.array([[c[0], c[1]] for c in circles])
    radius_est = int(np.median([c[2] for c in circles]))
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    # Split left/right answer blocks using the image X midpoint.
    # K-means on 2 clusters is unreliable here because the sheet has a
    # "question number" column printed in the middle (between the two answer
    # grids). That column's circles pull the right-block centroid leftward,
    # causing column offsets. A hard midpoint split is much more stable for
    # this fixed two-block layout.
    x_mid = w / 2.0
    block_labels_arr = np.where(pts[:, 0] < x_mid, 0, 1).reshape(-1, 1).astype(np.int32)
    block_labels = block_labels_arr
    # left block = 0, right block = 1
    left_block_id = 0

    results = {}
    bubble_details = {}  # Store detailed metrics for each bubble
    debug_img = orig.copy()
    
    # Adaptive sample radius based on detected bubble size
    sample_r = max(int(radius_est * 0.60), 7)
    
    # Apply advanced preprocessing for better fill detection
    # Otsu thresholding to separate filled from unfilled
    _, gray_thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # Compute shared row y-positions from all points pooled across both blocks
    row_y_sorted = _compute_row_y(pts)

    for block_id, q_offset in [(left_block_id, 0), (1 - left_block_id, 20)]:
        mask = block_labels.flatten() == block_id
        block_pts = pts[mask]

        if len(block_pts) < 4:
            for row_idx in range(20):
                q_num = row_idx + 1 + q_offset
                results[q_num] = {1: 255, 2: 255, 3: 255, 4: 255}
                # Initialize bubble_details for skipped questions
                bubble_details[q_num] = {
                    1: {'mean': 255, 'filled_ratio': 0.0},
                    2: {'mean': 255, 'filled_ratio': 0.0},
                    3: {'mean': 255, 'filled_ratio': 0.0},
                    4: {'mean': 255, 'filled_ratio': 0.0}
                }
            continue

        col_grid = _build_grid(block_pts, gray, radius_est, row_y_sorted)

        for row_idx in range(20):
            q_num = row_idx + 1 + q_offset
            row_vals = {}
            bubble_details[q_num] = {}
            
            for opt_idx in range(4):
                x, y = col_grid[opt_idx][row_idx]
                mask_c = np.zeros(gray.shape, dtype=np.uint8)
                cv2.circle(mask_c, (x, y), sample_r, 255, -1)
                
                # Multi-metric bubble analysis
                mean_val = cv2.mean(gray, mask=mask_c)[0]
                
                # Calculate filled pixel percentage using thresholded image
                filled_pixels = cv2.countNonZero(cv2.bitwise_and(255 - gray_thresh, mask_c))
                total_pixels = cv2.countNonZero(mask_c)
                filled_ratio = filled_pixels / (total_pixels + 1e-6)
                
                # Store detailed metrics
                bubble_details[q_num][opt_idx + 1] = {
                    'mean': mean_val,
                    'filled_ratio': filled_ratio
                }
                
                row_vals[opt_idx + 1] = mean_val
                
                # Enhanced debug visualization
                if debug_out:
                    color = (0, 255, 0) if filled_ratio < 0.15 else (0, 165, 255)
                    cv2.circle(debug_img, (x, y), radius_est, color, 2)
                    cv2.circle(debug_img, (x, y), 2, (0, 0, 255), -1)
                    label = f"{q_num}.{opt_idx + 1}"
                    cv2.putText(debug_img, label, (x - radius_est, y - radius_est - 3),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.35, (255, 0, 0), 1, cv2.LINE_AA)
            results[q_num] = row_vals

    if debug_out:
        cv2.imwrite(debug_out, debug_img)

    # --- TWO-PASS VERIFICATION SYSTEM ---
    # Pass 1: Intensity-based detection
    # Pass 2: Filled-pixel-ratio verification
    
    all_vals = [v for opts in results.values() for v in opts.values()]
    blank_baseline = float(np.median(all_vals))
    q1_val = float(np.percentile(all_vals, 25))
    q3_val = float(np.percentile(all_vals, 75))
    
    # Dynamic thresholds - RELAXED for better detection
    darkness_factor = 1.0 - (blank_baseline / 255.0) * 0.3
    global_fill_threshold = blank_baseline * (0.88 + darkness_factor * 0.05)  # Increased from 0.78
    absolute_dark_threshold = min(q1_val * 1.6, blank_baseline * 0.72)  # Increased from 1.4 and 0.60

    answers = {}
    flags = {}
    confidence_scores = {}
    
    log(f"Blank baseline: {blank_baseline:.1f}, Q1: {q1_val:.1f}, Threshold: {global_fill_threshold:.1f}")
    
    for q in sorted(results.keys()):
        opts = results[q]
        sorted_opts = sorted(opts.items(), key=lambda x: x[1])
        darkest_opt, darkest_val = sorted_opts[0]
        second_opt, second_val = sorted_opts[1]
        third_val = sorted_opts[2][1]
        brightest_val = sorted_opts[3][1]

        # Get filled ratios for verification
        darkest_filled_ratio = bubble_details[q][darkest_opt]['filled_ratio']
        second_filled_ratio = bubble_details[q][second_opt]['filled_ratio']
        
        # Calculate decision metrics
        row_mean = float(np.mean([v for _, v in sorted_opts]))
        row_std = float(np.std([v for _, v in sorted_opts]))
        gap_to_second = second_val - darkest_val
        relative_gap = gap_to_second / (row_mean + 1e-6)
        darkness_ratio = darkest_val / (blank_baseline + 1e-6)
        
        # MULTI-CRITERIA DECISION (Pass 1: Intensity) - RELAXED
        pass1_obvious = darkest_val < absolute_dark_threshold
        pass1_relative = relative_gap > 0.08 and gap_to_second > 3  # Reduced from 0.10 and 5
        pass1_below_global = darkest_val < global_fill_threshold
        
        # VERIFICATION (Pass 2: Filled pixels) - RELAXED
        pass2_filled = darkest_filled_ratio > 0.12  # Reduced from 0.18
        pass2_distinct = darkest_filled_ratio > second_filled_ratio * 1.3  # Reduced from 1.5
        
        # COMBINED DECISION - More lenient: Either pass can confirm
        is_filled_pass1 = (pass1_obvious or (pass1_below_global and pass1_relative))
        is_filled_pass2 = (pass2_filled and pass2_distinct)
        
        # Final decision: Either pass OR weak evidence from both
        is_filled = (is_filled_pass1 or is_filled_pass2) or \
                   (pass1_below_global and darkest_filled_ratio > 0.10) or \
                   (pass2_filled and darkest_val < global_fill_threshold * 1.05)
        
        # Confidence calculation - IMPROVED scoring
        confidence = 0
        if is_filled:
            # More generous confidence calculation
            intensity_conf = 45 * (1.0 - min(1.0, darkness_ratio))
            gap_conf = 35 * min(1.0, relative_gap / 0.15)  # Easier to get full points
            fill_conf = 20 * min(1.0, darkest_filled_ratio / 0.25)  # Lower bar
            confidence = min(100, int(intensity_conf + gap_conf + fill_conf + 15))  # +15 base bonus
        
        # Decision logic - LESS FLAGGING
        if is_filled:
            answers[q] = darkest_opt
            confidence_scores[q] = confidence
            
            # Multi-mark detection - stricter to reduce false positives
            second_is_filled = (second_val < global_fill_threshold * 0.90 and 
                              second_filled_ratio > 0.18 and
                              relative_gap < 0.08)  # Very close marks
            
            if second_is_filled:
                flags[q] = "multi_mark"
                confidence_scores[q] = max(30, confidence - 35)  # Less penalty
                log(f"Q{q}: Multi-mark detected (opt {darkest_opt} & {second_opt})")
            elif confidence < 45:  # Only flag if very low confidence (was 60)
                flags[q] = "low_confidence"
                log(f"Q{q}: Low confidence {confidence}% (opt {darkest_opt})")
        else:
            # Check for problematic cases - only flag truly problematic rows
            all_dark = brightest_val < global_fill_threshold * 0.85  # Stricter threshold
            all_similar = row_std < 3  # Must be very similar
            all_filled = all([bubble_details[q][i+1]['filled_ratio'] > 0.20 for i in range(4)])  # Higher bar
            
            if (all_dark and all_similar) or all_filled:
                flags[q] = "row_smudged"
                log(f"Q{q}: Row appears smudged/dirty")
            # Don't flag "no_clear_mark" - just leave blank without flagging
            
            answers[q] = None
            confidence_scores[q] = 0

    return answers, flags, results, confidence_scores


def score_sheet(answers, answer_key):
    correct = {}
    score = 0
    for q, key_opt in answer_key.items():
        marked = answers.get(q)
        is_correct = marked == key_opt
        correct[q] = is_correct
        if is_correct:
            score += 1
    return score, correct


def batch_to_excel(image_paths, out_xlsx, answer_key=None, roll_numbers=None):
    """Generate Excel report with color-coded confidence indicators."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    header = ["Sheet #", "Roll No"] + [f"Q{i}" for i in range(1, 41)]
    if answer_key:
        header += ["Score", "Flagged Questions", "Avg Confidence"]
    else:
        header += ["Flagged Questions", "Avg Confidence"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    yellow = PatternFill("solid", start_color="FFFF00")  # Low confidence
    orange = PatternFill("solid", start_color="FFA500")  # Multi-mark
    red = PatternFill("solid", start_color="FF6B6B")     # No mark

    for idx, path in enumerate(image_paths, start=1):
        try:
            answers, flags, _, confidence_scores = detect_bubbles(path)
            roll = roll_numbers[idx - 1] if roll_numbers and idx <= len(roll_numbers) else ""
            row = [idx, roll] + [answers.get(q, "") for q in range(1, 41)]
            
            avg_confidence = int(np.mean([confidence_scores.get(q, 0) for q in range(1, 41)]))
            
            if answer_key:
                score, _ = score_sheet(answers, answer_key)
                row += [score, ", ".join(f"Q{q}" for q in sorted(flags)), avg_confidence]
            else:
                row += [", ".join(f"Q{q}" for q in sorted(flags)), avg_confidence]
            ws.append(row)

            row_num = ws.max_row
            for q, flag_type in flags.items():
                col_idx = 2 + q
                if flag_type == "multi_mark":
                    ws.cell(row=row_num, column=col_idx).fill = orange
                elif flag_type == "low_confidence":
                    ws.cell(row=row_num, column=col_idx).fill = yellow
                elif flag_type in ["no_clear_mark", "row_smudged"]:
                    ws.cell(row=row_num, column=col_idx).fill = red
        except Exception as e:
            # Add error row
            row = [idx, roll_numbers[idx-1] if roll_numbers and idx <= len(roll_numbers) else ""] + ["ERROR"] * 40
            row += [0, str(e), 0] if answer_key else [str(e), 0]
            ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 10

    wb.save(out_xlsx)


def print_simple_answers(answers, total_questions=40):
    """Print 'Q{N}: Option X' for every question. Unmarked prints 'Not marked'."""
    for q in range(1, total_questions + 1):
        opt = answers.get(q)
        if opt is None:
            print(f"Q{q}: Not marked")
        else:
            print(f"Q{q}: {opt}")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--batch":
        out_xlsx = sys.argv[2]
        images = sys.argv[3:]
        batch_to_excel(images, out_xlsx)
        print(f"Saved results for {len(images)} sheet(s) to {out_xlsx}")
    else:
        img_path = sys.argv[1]
        debug_out = sys.argv[2] if len(sys.argv) > 2 else None
        answers, flags, raw, confidence = detect_bubbles(img_path, debug_out)
        print_simple_answers(answers)
        
        # Print confidence summary
        flagged = [q for q in flags if flags[q] != "no_clear_mark" or answers[q] is not None]
        if flagged:
            print(f"\nWarnings on questions: {', '.join(map(str, sorted(flagged)))}")
        avg_conf = int(np.mean([confidence.get(q, 0) for q in range(1, 41)]))
        print(f"Average confidence: {avg_conf}%")
