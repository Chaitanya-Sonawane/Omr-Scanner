"""
Batch OMR scanner — processes every sheet photo in a folder and produces one
Excel workbook with:
  - a "Results" sheet: one row per scanned photo, Q1-Q40 detected answers,
    sheet-level confidence, a review flag, and a live SCORE formula that
    reads against the answer key (so re-editing the key recalculates scores
    automatically, no re-scanning needed)
  - an "AnswerKey" sheet: one editable row of correct answers (fill this in
    before / after generating the report — the Score column recalculates
    either way once you open the file in Excel or LibreOffice)
  - a "Flagged" sheet: just the sheets that need a human look, so you don't
    have to scroll through all 50 to find the ones that need attention

USAGE:
    python batch_scan.py /path/to/sheets_folder /path/to/output_report.xlsx

    Sheets folder should contain .jpg/.jpeg/.png photos, one per student.
    Answer key defaults to all blank ("") — fill in AnswerKey!C2:AR2 in the
    output file with the 40 correct options (A/B/C/D) and re-open/recalculate
    to see scores.

REQUIRES: adaptive_detector.py in the same directory (or on PYTHONPATH).
"""
import sys
import os
import glob
import cv2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .omr_scanner import detect_bubbles

N_QUESTIONS = 40
IMG_EXTENSIONS = (".jpg", ".jpeg", ".png", ".JPG", ".JPEG", ".PNG")

HEADER_FILL = PatternFill("solid", start_color="2F5597", end_color="2F5597")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
FLAG_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
FLAG_FONT = Font(name="Arial", color="9C0006")
LOWCONF_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
NORMAL_FONT = Font(name="Arial")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

# Per-question confidence below this shades the cell as "check by hand"
LOW_CONF_CELL_THRESHOLD = 0.35


def scan_folder(folder_path):
    """Run detect_adaptive on every image in folder_path. Returns list of
    (filename, SheetDetection) sorted by filename, plus a list of
    (filename, error_message) for files that failed to process."""
    paths = sorted(
        p for p in glob.glob(os.path.join(folder_path, "*"))
        if p.endswith(IMG_EXTENSIONS)
    )
    if not paths:
        raise SystemExit(f"No image files found in {folder_path}")

    results, errors = [], []
    for path in paths:
        fname = os.path.basename(path)
        try:
            img = cv2.imread(path)
            if img is None:
                errors.append((fname, "Could not read image file (corrupt or unsupported format)"))
                print(f"  {fname}: FAILED (unreadable file)")
                continue
            
            # Use the integrated OMR scanner
            answers, flags, raw_results = detect_bubbles(path)
            
            # Convert to format expected by build_workbook
            # Create a simple result object compatible with the existing code
            class SimpleResult:
                def __init__(self, answers, flags, raw_results):
                    self.answers = answers
                    self.flags = flags
                    self.raw_results = raw_results
                    
                    # Calculate sheet confidence from flags
                    total_questions = 40
                    flagged_questions = len(flags)
                    self.sheet_confidence = max(0.0, 1.0 - (flagged_questions / total_questions))
                    
                    # Set flagged_for_review based on confidence threshold
                    self.flagged_for_review = self.sheet_confidence < 0.7
                    
                    # Create question objects
                    self.questions = []
                    option_map = {1: "A", 2: "B", 3: "C", 4: "D"}
                    
                    for q_num in range(1, 41):
                        detected_opt = answers.get(q_num)
                        detected_answer = option_map.get(detected_opt, "") if detected_opt else ""
                        
                        # Calculate question confidence
                        if q_num in raw_results and detected_opt:
                            intensities = list(raw_results[q_num].values())
                            if len(intensities) >= 2:
                                intensities.sort()
                                gap = intensities[1] - intensities[0]  # gap between darkest and second darkest
                                confidence = min(gap / 50.0, 1.0)  # normalize gap to 0-1
                            else:
                                confidence = 0.5
                        else:
                            confidence = 0.0
                            
                        # Create question object
                        q_obj = type('Question', (), {
                            'q_no': q_num,
                            'detected_answer': detected_answer,
                            'confidence': confidence
                        })()
                        self.questions.append(q_obj)
                    
                def get_detected_answers(self):
                    option_map = {1: "A", 2: "B", 3: "C", 4: "D"}
                    return {q: option_map.get(ans, "") if ans else "" 
                           for q, ans in self.answers.items()}
            
            result = SimpleResult(answers, flags, raw_results)
            results.append((fname, result))
            print(f"  {fname}: OK (confidence {result.sheet_confidence:.2f})")
        except Exception as e:
            errors.append((fname, str(e)))
            print(f"  {fname}: FAILED ({e})")
    return results, errors


def build_workbook(results, errors, out_path):
    wb = Workbook()

    # ---- AnswerKey sheet ----
    ak = wb.active
    ak.title = "AnswerKey"
    ak["A1"] = "Row"
    ak["A1"].font = HEADER_FONT
    ak["A1"].fill = HEADER_FILL
    ak["A2"] = "Correct Answer"
    ak["A2"].font = Font(name="Arial", bold=True)
    for q in range(1, N_QUESTIONS + 1):
        col = get_column_letter(q + 2)  # C.. onward (A=label, B spacer)
        ak[f"{col}1"] = f"Q{q}"
        ak[f"{col}1"].font = HEADER_FONT
        ak[f"{col}1"].fill = HEADER_FILL
        ak[f"{col}1"].alignment = Alignment(horizontal="center")
        ak[f"{col}2"] = ""  # fill in correct option here: A/B/C/D
        ak[f"{col}2"].font = Font(name="Arial", color="0000FF", bold=True)  # blue = input cell
        ak[f"{col}2"].alignment = Alignment(horizontal="center")
        ak.column_dimensions[col].width = 5
    ak["A3"] = "Fill in row 2 (blue cells) with the correct option letter for each question, then reopen/recalculate."
    ak["A3"].font = Font(name="Arial", italic=True, size=9)
    ak.column_dimensions["A"].width = 16

    # ---- Results sheet ----
    ws = wb.create_sheet("Results")
    headers = ["File", "Sheet Confidence", "Flagged for Review"] + [f"Q{q}" for q in range(1, N_QUESTIONS + 1)] + ["Score"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    q_start_col = 4  # column D = Q1
    score_col = q_start_col + N_QUESTIONS  # one after the last question column

    for r, (fname, result) in enumerate(results, start=2):
        ws.cell(row=r, column=1, value=fname).font = NORMAL_FONT
        conf_cell = ws.cell(row=r, column=2, value=result.sheet_confidence)
        conf_cell.font = NORMAL_FONT
        flag_cell = ws.cell(row=r, column=3, value="YES" if result.flagged_for_review else "")
        flag_cell.font = FLAG_FONT if result.flagged_for_review else NORMAL_FONT
        if result.flagged_for_review:
            flag_cell.fill = FLAG_FILL

        for q_obj in result.questions:
            col = q_start_col + q_obj.q_no - 1
            ans = q_obj.detected_answer if q_obj.detected_answer else "BLANK"
            cell = ws.cell(row=r, column=col, value=ans)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            if q_obj.confidence < LOW_CONF_CELL_THRESHOLD and ans not in ("BLANK",):
                cell.fill = LOWCONF_FILL

        # Live formula: count how many of this row's Q1..Q40 match AnswerKey row 2
        first_q_col = get_column_letter(q_start_col)
        last_q_col = get_column_letter(q_start_col + N_QUESTIONS - 1)
        ak_first = get_column_letter(3)
        ak_last = get_column_letter(3 + N_QUESTIONS - 1)
        formula = (
            f"=SUMPRODUCT((${first_q_col}{r}:${last_q_col}{r}="
            f"AnswerKey!${ak_first}$2:${ak_last}$2)*"
            f"(AnswerKey!${ak_first}$2:${ak_last}$2<>\"\"))"
        )
        ws.cell(row=r, column=score_col, value=formula).font = NORMAL_FONT

    for r, (fname, err) in enumerate(errors, start=len(results) + 2):
        ws.cell(row=r, column=1, value=fname).font = NORMAL_FONT
        ws.cell(row=r, column=2, value="ERROR").font = FLAG_FONT
        ws.cell(row=r, column=2).fill = FLAG_FILL
        ws.cell(row=r, column=3, value=err[:200]).font = FLAG_FONT

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for q in range(N_QUESTIONS):
        ws.column_dimensions[get_column_letter(q_start_col + q)].width = 7
    ws.column_dimensions[get_column_letter(score_col)].width = 9
    ws.freeze_panes = "D2"

    # ---- Flagged sheet: quick filtered view ----
    fl = wb.create_sheet("Flagged")
    fl["A1"] = "Sheets needing manual review"
    fl["A1"].font = Font(name="Arial", bold=True, size=12)
    for c, h in enumerate(["File", "Sheet Confidence", "Reason"], start=1):
        cell = fl.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    r = 3
    for fname, result in results:
        low_conf_qs = [q.q_no for q in result.questions if q.confidence < LOW_CONF_CELL_THRESHOLD]
        if result.flagged_for_review or low_conf_qs:
            reason = []
            if result.flagged_for_review:
                reason.append("low overall confidence")
            if low_conf_qs:
                reason.append(f"faint/ambiguous marks on Q{','.join(map(str, low_conf_qs))}")
            fl.cell(row=r, column=1, value=fname).font = NORMAL_FONT
            fl.cell(row=r, column=2, value=result.sheet_confidence).font = NORMAL_FONT
            fl.cell(row=r, column=3, value="; ".join(reason)).font = NORMAL_FONT
            r += 1
    for fname, err in errors:
        fl.cell(row=r, column=1, value=fname).font = FLAG_FONT
        fl.cell(row=r, column=2, value="ERROR").font = FLAG_FONT
        fl.cell(row=r, column=3, value=err[:200]).font = FLAG_FONT
        r += 1
    fl.column_dimensions["A"].width = 34
    fl.column_dimensions["B"].width = 14
    fl.column_dimensions["C"].width = 60

    wb.save(out_path)


def main():
    if len(sys.argv) < 3:
        print("Usage: python batch_scan.py <sheets_folder> <output_report.xlsx>")
        sys.exit(1)
    folder_path, out_path = sys.argv[1], sys.argv[2]
    print(f"Scanning sheets in {folder_path} ...")
    results, errors = scan_folder(folder_path)
    print(f"\nProcessed {len(results)} sheets, {len(errors)} failed.")
    build_workbook(results, errors, out_path)
    print(f"Report written to {out_path}")


if __name__ == "__main__":
    main()
