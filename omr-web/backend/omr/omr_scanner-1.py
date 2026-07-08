import cv2
import numpy as np
import json
import sys


def _try_hough(gray_blur, min_r, max_r, param2):
    circles = cv2.HoughCircles(
        gray_blur, cv2.HOUGH_GRADIENT, dp=1, minDist=18,
        param1=50, param2=param2, minRadius=min_r, maxRadius=max_r
    )
    return circles


def _detect_circles(gray):
    h, w = gray.shape
    r_est = max(8, min(30, w // 80))
    min_r = max(6, r_est - 6)
    max_r = r_est + 10
    blur = cv2.medianBlur(gray, 5)
    for param2 in [22, 18, 14, 10]:
        circles = _try_hough(blur, min_r, max_r, param2)
        if circles is not None and len(circles[0]) >= 8:
            return circles, blur
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    blur2 = cv2.medianBlur(enhanced, 5)
    for param2 in [18, 12]:
        circles = _try_hough(blur2, min_r, max_r, param2)
        if circles is not None and len(circles[0]) >= 8:
            return circles, blur2
    return None, blur


def _safe_kmeans(data, k, criteria, attempts=10):
    n = len(data)
    if n < k:
        return None, None
    _, labels, centers = cv2.kmeans(data, k, None, criteria, attempts, cv2.KMEANS_PP_CENTERS)
    return labels, centers


def _compute_row_y(all_pts):
    """
    Compute the 20 shared row y-positions from ALL detected bubble points
    on the sheet (both option blocks pooled together). Both blocks sit on
    the exact same printed horizontal lines, so pooling roughly doubles
    the points feeding each row's cluster (up to 8 per row instead of 4),
    which makes row positions far less likely to be thrown off by a
    single missing/extra detection in just one block.
    """
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)
    all_ys = all_pts[:, 1].reshape(-1, 1).astype(np.float32)
    n_rows_k = min(20, len(all_ys))
    row_labels, row_centers = _safe_kmeans(all_ys, n_rows_k, criteria)

    if row_labels is None:
        y_min, y_max = float(all_ys.min()), float(all_ys.max())
        row_y_sorted = np.linspace(y_min, y_max, 20).tolist()
    else:
        cluster_y = sorted(float(c) for c in row_centers.flatten())
        if len(cluster_y) < 20:
            diffs = np.diff(cluster_y)
            diffs = diffs[diffs > 3]
            pitch = float(np.median(diffs)) if len(diffs) else 30.0
            row_y_sorted = list(cluster_y)
            while len(row_y_sorted) < 20:
                row_y_sorted.append(row_y_sorted[-1] + pitch)
        else:
            row_y_sorted = cluster_y

    # Row positions should increase ~linearly (constant pitch, plus mild
    # tilt/perspective). KMeans occasionally misplaces an edge row (too few
    # points assigned to it); fit a line through all 20 rows and snap any
    # row whose position deviates too far from the fitted trend back onto it.
    row_idx_arr = np.arange(20)
    row_y_arr = np.array(row_y_sorted, dtype=np.float64)
    diffs = np.diff(row_y_arr)
    diffs = diffs[diffs > 3]
    est_pitch = float(np.median(diffs)) if len(diffs) else 30.0
    for _ in range(2):
        ay, by_ = np.polyfit(row_idx_arr, row_y_arr, 1)
        fitted = ay * row_idx_arr + by_
        resid = row_y_arr - fitted
        bad = np.abs(resid) > max(est_pitch * 0.4, 8)
        if not bad.any():
            break
        row_y_arr[bad] = fitted[bad]
    return row_y_arr.tolist()


def _build_grid(block_pts, gray, radius_est, row_y_sorted):
    """
    Build a 20x4 grid of (x,y) sample positions for one answer block,
    given shared row y-positions (see _compute_row_y). Only the x-center
    of each of the 4 option columns is computed per-block here.
    """
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)
    h, w = gray.shape

    # cluster into 4 columns (x-position is stable across rows)
    n_cols = min(4, len(block_pts))
    bxs = block_pts[:, 0].reshape(-1, 1).astype(np.float32)
    col_labels, col_centers = _safe_kmeans(bxs, n_cols, criteria)
    if col_labels is None or n_cols < 4:
        x_min_b, x_max_b = float(bxs.min()), float(bxs.max())
        col_centers_list = np.linspace(x_min_b, x_max_b, 4).tolist()
        col_labels = np.zeros((len(block_pts), 1), dtype=np.int32)
        cluster_to_opt = {0: 0}
    else:
        col_centers_list = sorted(col_centers.flatten().tolist())
        cluster_order = np.argsort(col_centers.flatten())
        cluster_to_opt = {int(cluster_order[i]): i for i in range(len(cluster_order))}

    # refine each column's x-center as the median x of its assigned points
    col_pts_by_opt = {opt_idx: [] for opt_idx in range(4)}
    for i in range(len(block_pts)):
        opt_idx = cluster_to_opt.get(int(col_labels[i][0]), -1)
        if 0 <= opt_idx < 4:
            col_pts_by_opt[opt_idx].append(block_pts[i])
    col_x_center = {}
    for opt_idx in range(4):
        pts_list = col_pts_by_opt[opt_idx]
        col_x_center[opt_idx] = (float(np.median([p[0] for p in pts_list]))
                                  if pts_list else col_centers_list[opt_idx])

    # build final grid: shared y per row, fixed x per column
    col_grid = {}
    for opt_idx in range(4):
        cx = col_x_center[opt_idx]
        grid_col = []
        for row_idx in range(20):
            gx = int(round(cx))
            gy = int(round(row_y_sorted[row_idx]))
            gx = max(radius_est, min(w - radius_est - 1, gx))
            gy = max(radius_est, min(h - radius_est - 1, gy))
            grid_col.append((gx, gy))
        col_grid[opt_idx] = grid_col

    return col_grid



def detect_bubbles(img_path, debug_out=None):
    img = cv2.imread(img_path)
    if img is None:
        raise RuntimeError(f"Could not read image: {img_path}")
    orig = img.copy()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape

    raw_circles, gray_blur = _detect_circles(gray)
    if raw_circles is None:
        raise RuntimeError(
            f"No circles found in {img_path}. "
            "Check that the image shows a clear OMR sheet with visible bubbles."
        )
    circles = np.round(raw_circles[0]).astype(int)

    circles = [c for c in circles if 50 < c[0] < w - 50 and 200 < c[1] < h - 30]
    if len(circles) < 8:
        raise RuntimeError(f"Too few circles detected ({len(circles)}) after filtering edges.")

    # filter to circles near median radius (removes text/border/stamp noise)
    radii = np.array([c[2] for c in circles])
    r_med = np.median(radii)
    circles = [c for c in circles if abs(c[2] - r_med) <= max(2, r_med * 0.35)]
    if len(circles) < 8:
        raise RuntimeError(f"Too few circles remain ({len(circles)}) after radius filtering.")

    # drop stray outliers above the answer grid (e.g. the circular
    # institution logo/stamp printed in the header, which can match the
    # expected bubble radius). Header debris is always a small cluster near
    # the very top of the sheet, so scan a generous early window (not just
    # the first ~15 points, which can miss a larger debris cluster) and cut
    # at the single largest gap found there. Picking the largest gap (rather
    # than aiming for an exact expected count) avoids over-trimming real
    # grid rows when the total circle count is naturally off from 160.
    ys_all = sorted(c[1] for c in circles)
    n_total = len(ys_all)
    window = min(50, n_total)
    cut_y = None
    best_gap = 0
    for i in range(1, window):
        gap = ys_all[i] - ys_all[i - 1]
        if gap > 60 and gap > best_gap:
            best_gap = gap
            cut_y = ys_all[i]
    if cut_y is not None:
        circles = [c for c in circles if c[1] >= cut_y]
    if len(circles) < 8:
        raise RuntimeError(f"Too few circles remain ({len(circles)}) after outlier removal.")

    pts = np.array([[c[0], c[1]] for c in circles])
    radius_est = int(np.median([c[2] for c in circles]))
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    xs = pts[:, 0].reshape(-1, 1).astype(np.float32)
    block_labels, block_centers = _safe_kmeans(xs, 2, criteria)
    if block_labels is None:
        block_labels = np.zeros((len(pts), 1), dtype=np.int32)
        block_centers = np.array([[xs.mean()], [xs.mean()]], dtype=np.float32)

    left_block_id = int(np.argmin(block_centers.flatten()))

    results = {}
    debug_img = orig.copy()
    # Sample only the inner ~55% of each bubble's radius. A tighter sample
    # area is less likely to pick up ink bleeding in from an adjacent row
    # or column when the grid position is off by a few pixels.
    sample_r = max(int(radius_est * 0.55), 6)

    for block_id, q_offset in [(left_block_id, 0), (1 - left_block_id, 20)]:
        mask = block_labels.flatten() == block_id
        block_pts = pts[mask]

        if len(block_pts) < 4:
            for row_idx in range(20):
                results[row_idx + 1 + q_offset] = {1: 255, 2: 255, 3: 255, 4: 255}
            continue

        row_y_sorted = _compute_row_y(block_pts)
        col_grid = _build_grid(block_pts, gray, radius_est, row_y_sorted)

        for row_idx in range(20):
            q_num = row_idx + 1 + q_offset
            row_vals = {}
            for opt_idx in range(4):
                x, y = col_grid[opt_idx][row_idx]
                mask_c = np.zeros(gray.shape, dtype=np.uint8)
                cv2.circle(mask_c, (x, y), sample_r, 255, -1)
                mean_val = cv2.mean(gray, mask=mask_c)[0]
                row_vals[opt_idx + 1] = mean_val
                cv2.circle(debug_img, (x, y), radius_est, (0, 255, 0), 2)
                cv2.circle(debug_img, (x, y), 2, (0, 0, 255), -1)
                # Label with "question.option" so the sampled x/y position can be
                # visually cross-checked against the printed "1 2 3 4" columns.
                label = f"{q_num}.{opt_idx + 1}"
                cv2.putText(debug_img, label, (x - radius_est, y - radius_est - 3),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.35, (255, 0, 0), 1, cv2.LINE_AA)
            results[q_num] = row_vals

    if debug_out:
        cv2.imwrite(debug_out, debug_img)

    # ---- adaptive blank baseline ----
    # Across a 40-question sheet, most of the 160 sampled bubbles are unmarked.
    # Their mean brightness (the "blank baseline") varies with paper/lighting/
    # camera conditions, so a fixed absolute threshold doesn't generalize
    # across photos. Instead, take the baseline directly from this sheet.
    all_vals = [v for opts in results.values() for v in opts.values()]
    blank_baseline = float(np.median(all_vals))
    # a genuinely filled bubble is much darker than the blank baseline
    fill_threshold = blank_baseline * 0.72

    answers = {}
    flags = {}
    for q in sorted(results.keys()):
        opts = results[q]
        sorted_opts = sorted(opts.items(), key=lambda x: x[1])
        darkest_opt, darkest_val = sorted_opts[0]
        second_opt, second_val = sorted_opts[1]

        if darkest_val < fill_threshold:
            answers[q] = darkest_opt
            # still record ambiguity internally (not surfaced in the simple output)
            if second_val < fill_threshold:
                flags[q] = "multi_mark"
        else:
            answers[q] = None
            flags[q] = "no_clear_mark"

    return answers, flags, results


def score_sheet(answers, answer_key):
    correct = {}
    score = 0
    for q, key_opt in answer_key.items():
        marked = answers.get(q)
        is_correct = marked == key_opt
        correct[q] = is_correct
        if is_correct:
            score += 1
    return score, correct


def batch_to_excel(image_paths, out_xlsx, answer_key=None, roll_numbers=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    header = ["Sheet #", "Roll No"] + [f"Q{i}" for i in range(1, 41)]
    if answer_key:
        header += ["Score", "Flagged (low confidence)"]
    else:
        header += ["Flagged (low confidence)"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    yellow = PatternFill("solid", start_color="FFFF00")

    for idx, path in enumerate(image_paths, start=1):
        answers, flags, _ = detect_bubbles(path)
        roll = roll_numbers[idx - 1] if roll_numbers else ""
        row = [idx, roll] + [answers.get(q, "") for q in range(1, 41)]
        if answer_key:
            score, _ = score_sheet(answers, answer_key)
            row += [score, ", ".join(f"Q{q}" for q in sorted(flags))]
        else:
            row += [", ".join(f"Q{q}" for q in sorted(flags))]
        ws.append(row)

        row_num = ws.max_row
        for q in flags:
            col_idx = 2 + q
            ws.cell(row=row_num, column=col_idx).fill = yellow

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 10

    wb.save(out_xlsx)


def print_simple_answers(answers, total_questions=40):
    """Print just 'Question N: Option X' for every question, one per line.
    Unmarked questions print as 'Not marked'."""
    for q in range(1, total_questions + 1):
        opt = answers.get(q)
        if opt is None:
            print(f"Q{q}: Not marked")
        else:
            print(f"Q{q}: {opt}")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--batch":
        out_xlsx = sys.argv[2]
        images = sys.argv[3:]
        batch_to_excel(images, out_xlsx)
        print(f"Saved results for {len(images)} sheet(s) to {out_xlsx}")
    else:
        img_path = sys.argv[1]
        debug_out = sys.argv[2] if len(sys.argv) > 2 else None
        answers, flags, raw = detect_bubbles(img_path, debug_out)
        print_simple_answers(answers)
