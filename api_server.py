#!/usr/bin/env python3
"""
FastAPI server for OMR Scanner Web Interface
Connects the Next.js UI to the Python OMR scanner backend
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from typing import List, Optional, Dict
import cv2
import numpy as np
import json
import uuid
import os
import tempfile
import asyncio
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# --- OMR pipeline imports ---------------------------------------------
# Prefer the backend package (omr-web/backend/omr/) which has the
# adaptive/self-calibrating scanner + preprocessor + enhancer. Fall back to
# the root-level omr_scanner.py only if the backend package can't be found,
# so the API still runs (with the weaker scanner) rather than crashing.
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'omr-web', 'backend'))

try:
    import omr.omr_scanner as omr_scanner
    print("✅ Using backend scanner: omr-web/backend/omr/omr_scanner.py")
except ImportError as e:
    print(f"⚠️  Backend scanner not found ({e}), falling back to root omr_scanner.py")
    import omr_scanner  # root-level fallback

app = FastAPI(title="OMR Scanner API")

# CORS Configuration
# Allow localhost for dev + any Netlify domain for production
# Set ALLOWED_ORIGINS env var on Render to your exact Netlify URL if needed
_raw_origins = os.environ.get(
    "ALLOWED_ORIGINS",
    "http://localhost:3000,http://127.0.0.1:3000"
)
_allow_origins = [o.strip() for o in _raw_origins.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allow_origins,
    allow_origin_regex=r"https://.*\.netlify\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory storage (use database in production)
sessions = {}
results_store = {}

# Persistent answer key storage
# Use /tmp on Render (ephemeral but writable); fall back to local dir in dev
_DATA_DIR = os.environ.get("DATA_DIR", os.path.join(os.path.dirname(__file__), "data"))
os.makedirs(_DATA_DIR, exist_ok=True)
_SAVED_KEY_FILE = os.path.join(_DATA_DIR, "saved_answer_key.json")
_saved_answer_key: dict = {"answers": {}}

# Load saved key from disk on startup
if os.path.isfile(_SAVED_KEY_FILE):
    try:
        with open(_SAVED_KEY_FILE) as f:
            _saved_answer_key["answers"] = json.load(f)
        print(f"✅ Loaded saved answer key ({len(_saved_answer_key['answers'])} questions)")
    except Exception as e:
        print(f"⚠️  Could not load saved answer key: {e}")


def normalize_answers(answers: dict) -> dict:
    """
    Normalize answer key to consistent format: {"1": "A", "2": "B", ...}
    Accepts any of these input formats from frontend or scanner:
      - {"q1": "A", "q2": "B"}   (frontend manual entry)
      - {"1": "A", "2": "B"}     (already normalized)
      - {"1": 1, "2": 2}         (scanner int options)
      - {"1": "1", "2": "2"}     (scanner str options)
    """
    OPT_MAP = {"1": "A", "2": "B", "3": "C", "4": "D",
               1: "A", 2: "B", 3: "C", 4: "D"}
    normalized = {}
    for k, v in answers.items():
        # Strip leading 'q' from key
        q_num = str(k).lstrip('q').lstrip('Q')
        # Convert numeric option to letter
        if v in OPT_MAP:
            normalized[q_num] = OPT_MAP[v]
        else:
            normalized[q_num] = str(v).upper()
    return normalized


def _persist_answer_key(answers: dict):
    """Save normalized answer key to disk."""
    _saved_answer_key["answers"] = answers
    try:
        with open(_SAVED_KEY_FILE, "w") as f:
            json.dump(answers, f)
    except Exception as e:
        print(f"⚠️  Could not persist answer key: {e}")


def process_omr_image(image_path: str) -> tuple:
    """
    Preprocess (resize to 800×1040, CLAHE) then scan with OMR scanner.
    Preprocessing dramatically reduces CPU time on large phone-camera photos.
    Returns (answers, flags, raw_data)
    """
    # Try to use the preprocessor from the backend package
    try:
        import sys as _sys
        _sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'omr-web', 'backend'))
        from omr.preprocessor import preprocess
        import cv2 as _cv2
        processed = preprocess(image_path)   # float32 grayscale 800×1040
        # Write processed image to a temp file so detect_bubbles can read it
        with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as tmp:
            tmp_path = tmp.name
        _cv2.imwrite(tmp_path, processed.astype('uint8'))
        try:
            answers, flags, raw, _conf = omr_scanner.detect_bubbles(tmp_path)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    except Exception:
        # Fallback: run scanner directly on the original image
        answers, flags, raw, _conf = omr_scanner.detect_bubbles(image_path)
    return answers, flags, raw


class SessionData:
    def __init__(self, session_id: str):
        self.session_id = session_id
        self.answer_key = None
        self.sheets = []
        self.results = []
        self.status = "created"
        self.created_at = datetime.now()


@app.get("/")
async def root():
    return {"message": "OMR Scanner API", "version": "1.0", "status": "running"}

@app.get("/health")
async def health():
    return {"status": "ok"}


@app.post("/api/session")
async def create_session():
    """Create a new OMR processing session"""
    session_id = f"sess_{uuid.uuid4().hex[:8]}"
    sessions[session_id] = SessionData(session_id)
    return {"session_id": session_id, "status": "created"}


@app.post("/api/session/{session_id}/answer-key/manual")
async def set_answer_key_manual(session_id: str, answer_key: Dict):
    """Set answer key manually and persist it globally"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = sessions[session_id]
    answers = normalize_answers(answer_key.get("answers", {}))
    session.answer_key = answers
    session.status = "answer_key_set"

    # Auto-save globally
    _persist_answer_key(answers)

    return {"status": "success", "message": "Answer key set successfully", "answers": answers}


@app.post("/api/session/{session_id}/answer-key")
async def set_answer_key_image(
    session_id: str,
    file: UploadFile = File(...)
):
    """Extract answer key from uploaded OMR image"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Save uploaded file temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        # Process the image using your OMR scanner
        image = cv2.imread(tmp_path)
        if image is None:
            raise HTTPException(status_code=400, detail="Invalid image file")
        
        # Call your OMR scanner function with enhancement
        answers, flags, raw = process_omr_image(tmp_path)
        
        if answers is None:
            raise HTTPException(status_code=400, detail="Could not detect bubbles in image")
        
        # Convert and normalize answers to A/B/C/D with numeric keys
        answer_key = normalize_answers({str(q): opt for q, opt in answers.items() if opt is not None})

        session = sessions[session_id]
        session.answer_key = answer_key
        session.status = "answer_key_set"

        # Auto-save globally
        _persist_answer_key(answer_key)

        return {"status": "success", "answers": answer_key, "flags": flags}
    
    finally:
        # Clean up temp file
        os.unlink(tmp_path)


@app.post("/api/session/{session_id}/sheets")
async def upload_sheets(
    session_id: str,
    files: List[UploadFile] = File(...),
    names: Optional[str] = Form(None)
):
    """Upload student answer sheets for processing"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = sessions[session_id]
    
    if not session.answer_key:
        raise HTTPException(status_code=400, detail="Answer key not set")
    
    # Parse student names if provided
    name_list = names.split(',') if names else []
    
    # Process each uploaded sheet
    sheet_ids = []
    for idx, file in enumerate(files):
        sheet_id = f"sheet_{uuid.uuid4().hex[:8]}"
        
        # Save file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        student_name = name_list[idx] if idx < len(name_list) and name_list[idx] else f"Student_{len(session.sheets) + 1}"
        
        sheet_info = {
            "id": sheet_id,
            "filename": file.filename,
            "path": tmp_path,
            "student_id": student_name,
            "name": student_name,
            "status": "QUEUED"
        }
        
        session.sheets.append(sheet_info)
        sheet_ids.append(sheet_id)
    
    # Automatically start processing
    session.status = "processing"
    
    return {
        "status": "success",
        "sheet_ids": sheet_ids,
        "total_sheets": len(sheet_ids)
    }


@app.get("/api/session/{session_id}/progress")
async def process_sheets(session_id: str):
    """Process all uploaded sheets (Server-Sent Events for progress)"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = sessions[session_id]
    
    async def event_generator():
        """Stream processing progress"""
        total = len(session.sheets)
        
        if total == 0:
            yield f"data: {json.dumps({'type': 'BATCH_COMPLETE', 'totalProcessed': 0})}\n\n"
            return
        
        for idx, sheet in enumerate(session.sheets):
            # Skip already processed sheets (also guards against browsers
            # auto-reconnecting the SSE stream and reprocessing everything)
            if sheet["status"] in ("DONE", "ERROR"):
                continue
            
            # Use the sheet's real, persistent id (the same one returned by
            # POST /sheets in sheet_ids). Previously this was a positional
            # zfill(3) index recomputed from session.sheets, which only
            # matched the frontend's own numbering for a session's very
            # first upload batch -- any later batch (or the sheets list
            # growing between calls) produced ids the frontend never
            # assigned to any row, so those events had nowhere to land and
            # the rows sat stuck on QUEUED forever.
            sheet_id = sheet["id"]
                
            # Update status to PROCESSING
            sheet["status"] = "PROCESSING"
            yield f"data: {json.dumps({'type': 'PROCESSING', 'sheet_id': sheet_id, 'status': 'PROCESSING', 'student_id': sheet.get('name')})}\n\n"
            
            try:
                # Process the sheet with enhancement
                answers, flags, raw = process_omr_image(sheet["path"])
                
                if answers is None:
                    sheet["status"] = "ERROR"
                    sheet["error"] = "Could not detect bubbles"
                    yield f"data: {json.dumps({'type': 'ERROR', 'sheet_id': sheet_id, 'status': 'ERROR', 'error': 'Detection failed'})}\n\n"
                    continue
                
                # Compare with answer key — both normalized to A/B/C/D
                student_answers = {}
                OPT_MAP = {"1": "A", "2": "B", "3": "C", "4": "D"}
                for q, opt in answers.items():
                    if opt is not None:
                        student_answers[str(q)] = OPT_MAP.get(str(opt), str(opt))
                correct_answers = session.answer_key
                
                score = 0
                total_questions = len(correct_answers)
                detailed_answers = {}
                
                for q_num, correct_ans in correct_answers.items():
                    student_ans = student_answers.get(str(q_num), "")
                    is_correct = student_ans == correct_ans
                    if is_correct:
                        score += 1
                    detailed_answers[q_num] = {
                        "marked": student_ans,
                        "correct": correct_ans,
                        "is_correct": is_correct
                    }
                
                # Calculate confidence (1.0 if no flags, lower if there are issues)
                confidence = 1.0 - (len(flags) * 0.1) if flags else 1.0
                confidence = max(0.0, min(1.0, confidence))
                
                # Store result
                result_data = {
                    "id": sheet["student_id"],
                    "name": sheet["name"],
                    "filename": sheet["filename"],
                    "score": score,
                    "total": total_questions,
                    "percentage": (score / total_questions * 100) if total_questions > 0 else 0,
                    "answers": detailed_answers,
                    "flags": flags,
                    "confidence": confidence,
                    "timestamp": datetime.now().isoformat()
                }
                
                session.results.append(result_data)
                sheet["status"] = "DONE"
                sheet["result"] = result_data
                
                # Send DONE event with all required fields
                yield f"data: {json.dumps({'type': 'DONE', 'sheet_id': sheet_id, 'status': 'DONE', 'score': {'correct': score, 'total': total_questions}, 'confidence': confidence, 'student_id': sheet.get('name')})}\n\n"
                
            except Exception as e:
                sheet["status"] = "ERROR"
                sheet["error"] = str(e)
                yield f"data: {json.dumps({'type': 'ERROR', 'sheet_id': sheet_id, 'status': 'ERROR', 'error': str(e)})}\n\n"
            
            # Small delay for UI
            await asyncio.sleep(0.3)
        
        # Mark session as complete
        session.status = "complete"
        
        # Send completion
        yield f"data: {json.dumps({'type': 'BATCH_COMPLETE', 'totalProcessed': len(session.results)})}\n\n"
    
    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/api/session/{session_id}/results")
async def get_results(session_id: str):
    """Get all results for a session"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = sessions[session_id]
    
    # Calculate statistics
    if not session.results:
        return {"results": [], "statistics": {}, "status": session.status}
    
    scores = [r["score"] for r in session.results]
    total_questions = session.results[0]["total"]
    passing_score = total_questions // 2  # 50% passing
    
    stats = {
        "total_students": len(session.results),
        "average_score": sum(scores) / len(scores),
        "highest_score": max(scores),
        "lowest_score": min(scores),
        "pass_count": len([s for s in scores if s >= passing_score]),
        "fail_count": len([s for s in scores if s < passing_score]),
        "pass_rate": (len([s for s in scores if s >= passing_score]) / len(scores)) * 100
    }
    
    return {
        "results": session.results,
        "statistics": stats,
        "total_questions": total_questions,
        "status": session.status
    }


@app.get("/api/session/{session_id}/status")
async def get_status(session_id: str):
    """Get session status"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = sessions[session_id]
    
    return {
        "status": session.status,
        "total_sheets": len(session.sheets),
        "processed": len([s for s in session.sheets if s["status"] == "DONE"]),
        "queued": len([s for s in session.sheets if s["status"] == "QUEUED"]),
        "processing": len([s for s in session.sheets if s["status"] == "PROCESSING"]),
        "errors": len([s for s in session.sheets if s["status"] == "ERROR"])
    }


@app.get("/api/session/{session_id}/answer-key")
async def get_answer_key(session_id: str):
    """Get the answer key for a session"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = sessions[session_id]
    
    if not session.answer_key:
        raise HTTPException(status_code=404, detail="Answer key not set")
    
    return {"answers": session.answer_key}


@app.get("/api/answer-key")
async def check_saved_answer_key():
    """Check if there's a saved answer key — returns frontend-compatible format"""
    if _saved_answer_key["answers"]:
        # Return with q-prefixed keys for the frontend
        frontend_answers = {f"q{k}": v for k, v in _saved_answer_key["answers"].items()}
        return {"answers": frontend_answers}
    raise HTTPException(status_code=404, detail="No saved answer key")


@app.post("/api/answer-key/save-manual")
async def save_answer_key_global(payload: Dict):
    """Persist the answer key globally so it survives new sessions"""
    answers = normalize_answers(payload.get("answers", {}))
    if not answers:
        raise HTTPException(status_code=400, detail="No answers provided")
    _persist_answer_key(answers)
    return {"status": "saved", "total_questions": len(answers)}


@app.post("/api/session/{session_id}/answer-key/use-saved")
async def use_saved_answer_key(session_id: str):
    """Load the globally saved answer key into this session"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    if not _saved_answer_key["answers"]:
        raise HTTPException(status_code=404, detail="No saved answer key available")
    session = sessions[session_id]
    session.answer_key = _saved_answer_key["answers"]
    session.status = "answer_key_set"
    # Return q-prefixed keys for the frontend state
    frontend_answers = {f"q{k}": v for k, v in _saved_answer_key["answers"].items()}
    return {"status": "success", "answers": frontend_answers}


@app.get("/api/session/{session_id}/sheet/{sheet_id}/raw")
async def get_raw_detection(session_id: str, sheet_id: str):
    """Get raw detection data for a specific sheet"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = sessions[session_id]
    sheet = next((s for s in session.sheets if s["id"] == sheet_id), None)
    
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    
    if "result" not in sheet:
        raise HTTPException(status_code=400, detail="Sheet not processed yet")
    
    return sheet["result"]


@app.get("/api/session/{session_id}/detection/{sheet_id}")
async def get_detection(session_id: str, sheet_id: str):
    """Get per-question detection data for SheetAnswers component"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = sessions[session_id]
    sheet = next((s for s in session.sheets if s["id"] == sheet_id), None)

    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    if "result" not in sheet:
        raise HTTPException(status_code=400, detail="Sheet not processed yet")

    result = sheet["result"]

    # Build per-question list for the frontend SheetAnswers component
    questions = []
    for q_num, detail in result.get("answers", {}).items():
        questions.append({
            "q_no": int(q_num),
            "marked": detail.get("marked", ""),
            "correct": detail.get("correct", ""),
            "is_correct": detail.get("is_correct", False),
        })
    questions.sort(key=lambda x: x["q_no"])

    return {
        "student_id": result.get("name", sheet.get("name", "")),
        "filename": result.get("filename", sheet.get("filename", "")),
        "total_score": result.get("score", 0),
        "out_of": result.get("total", 40),
        "questions": questions,
    }


@app.get("/api/results")
async def get_all_results():
    """Get results from all sessions"""
    all_results = []
    for session in sessions.values():
        all_results.extend(session.results)
    
    return {"results": all_results, "total": len(all_results)}


@app.get("/api/session/{session_id}/export/excel")
async def export_excel(session_id: str):
    """Export results as Excel file"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = sessions[session_id]
    
    if not session.results:
        raise HTTPException(status_code=400, detail="No results to export")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OMR Results"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    headers = ["Student ID", "Name", "Score", "Total", "Percentage", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    total_questions = session.results[0]["total"]
    passing_score = total_questions // 2
    
    for row, result in enumerate(session.results, 2):
        score = result["score"]
        percentage = (score / total_questions) * 100
        status = "Pass" if score >= passing_score else "Fail"
        
        ws.cell(row, 1, result["id"])
        ws.cell(row, 2, result["name"])
        ws.cell(row, 3, score)
        ws.cell(row, 4, total_questions)
        ws.cell(row, 5, f"{percentage:.1f}%")
        
        status_cell = ws.cell(row, 6, status)
        if status == "Pass":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as downloadable file
    filename = f"omr_results_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/session/{session_id}/export/pdf")
async def export_pdf(session_id: str):
    """Export results as PDF — one page per student + summary page"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = sessions[session_id]

    if not session.results:
        raise HTTPException(status_code=400, detail="No results to export")

    from reportlab.platypus import PageBreak, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    SUBJECTS = [
        ("Intelligence",    1,  10),
        ("Science",        11,  20),
        ("Social Science", 21,  30),
        ("Mathematics",    31,  40),
    ]

    def subject_score(answers: dict, start: int, end: int) -> tuple:
        correct = sum(
            1 for q in range(start, end + 1)
            if answers.get(str(q), {}).get("is_correct", False)
        )
        return correct, 10

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=0.6*inch, rightMargin=0.6*inch,
        topMargin=0.6*inch, bottomMargin=0.6*inch
    )
    styles = getSampleStyleSheet()
    elements = []

    BLUE   = colors.HexColor('#2563EB')
    LBLUE  = colors.HexColor('#DBEAFE')
    GREEN  = colors.HexColor('#16A34A')
    RED    = colors.HexColor('#DC2626')
    LGREY  = colors.HexColor('#F3F4F6')
    GOLD   = colors.HexColor('#F59E0B')

    title_style  = styles['Title']
    normal_style = styles['Normal']

    # ── Per-student pages ──────────────────────────────────────────────
    for result in session.results:
        name     = result.get("name", "Unknown")
        answers  = result.get("answers", {})
        total_sc = result.get("score", 0)
        total_q  = result.get("total", 40)

        # Header
        elements.append(Paragraph(f"<b>{name}</b>", title_style))
        elements.append(Paragraph(
            f"Session: {session_id} &nbsp;&nbsp; Date: {datetime.now().strftime('%Y-%m-%d')}",
            normal_style
        ))
        elements.append(HRFlowable(width="100%", thickness=1, color=BLUE))
        elements.append(Spacer(1, 0.15*inch))

        # Subject score summary
        subj_data = [["Subject", "Score", "Out of"]]
        for subj, s, e in SUBJECTS:
            sc, mx = subject_score(answers, s, e)
            subj_data.append([subj, str(sc), str(mx)])
        subj_data.append(["TOTAL", str(total_sc), str(total_q)])

        subj_table = Table(subj_data, colWidths=[2.8*inch, 1*inch, 1*inch])
        subj_table.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0),  BLUE),
            ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('BACKGROUND',   (0, -1), (-1, -1), LBLUE),
            ('FONTNAME',     (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN',        (1, 0), (-1, -1),  'CENTER'),
            ('GRID',         (0, 0), (-1, -1),  0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, LGREY]),
        ]))
        elements.append(subj_table)
        elements.append(Spacer(1, 0.2*inch))

        # Per-question response table (10 cols wide)
        elements.append(Paragraph("<b>Question-wise Responses</b>", normal_style))
        elements.append(Spacer(1, 0.08*inch))

        for subj, s, e in SUBJECTS:
            elements.append(Paragraph(f"<b>{subj} (Q{s}–Q{e})</b>", normal_style))
            header_row = ["Q"] + [str(q) for q in range(s, e+1)]
            marked_row = ["Marked"]
            correct_row = ["Correct"]
            result_row  = ["Result"]

            for q in range(s, e+1):
                d = answers.get(str(q), {})
                marked_row.append(d.get("marked", "-") or "-")
                correct_row.append(d.get("correct", "-") or "-")
                result_row.append("✓" if d.get("is_correct") else "✗")

            q_table = Table(
                [header_row, marked_row, correct_row, result_row],
                colWidths=[0.85*inch] + [0.55*inch]*10
            )
            # Build cell colors for result row
            result_styles = [
                ('BACKGROUND', (0, 0), (-1, 0), BLUE),
                ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
                ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (0, -1), LGREY),
                ('FONTNAME',   (0, 1), (0, -1), 'Helvetica-Bold'),
                ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
                ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE',   (0, 0), (-1, -1), 8),
            ]
            for col_i, q in enumerate(range(s, e+1), start=1):
                d = answers.get(str(q), {})
                cell_color = GREEN if d.get("is_correct") else RED
                result_styles.append(('TEXTCOLOR', (col_i, 3), (col_i, 3), cell_color))
                result_styles.append(('FONTNAME',  (col_i, 3), (col_i, 3), 'Helvetica-Bold'))

            q_table.setStyle(TableStyle(result_styles))
            elements.append(q_table)
            elements.append(Spacer(1, 0.12*inch))

        elements.append(PageBreak())

    # ── Summary page ──────────────────────────────────────────────────
    elements.append(Paragraph("<b>Session Results Summary</b>", title_style))
    elements.append(Paragraph(
        f"Session: {session_id} &nbsp;&nbsp; Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        normal_style
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=BLUE))
    elements.append(Spacer(1, 0.2*inch))

    # Full results table
    hdr = ["Name", "Intel\n/10", "Sci\n/10", "Soc.Sci\n/10", "Maths\n/10", "Total\n/40"]
    summary_data = [hdr]

    for result in session.results:
        answers = result.get("answers", {})
        row = [result.get("name", "")]
        for _, s, e in SUBJECTS:
            sc, _ = subject_score(answers, s, e)
            row.append(str(sc))
        row.append(str(result.get("score", 0)))
        summary_data.append(row)

    sum_table = Table(summary_data, colWidths=[2.2*inch, 0.8*inch, 0.7*inch, 0.9*inch, 0.8*inch, 0.8*inch])
    sum_table.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0),  BLUE),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('ALIGN',          (1, 0), (-1, -1), 'CENTER'),
        ('GRID',           (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LGREY]),
        ('FONTSIZE',       (0, 0), (-1, -1), 9),
    ]))
    elements.append(sum_table)
    elements.append(Spacer(1, 0.3*inch))

    # Top 3 rankers
    sorted_results = sorted(session.results, key=lambda r: r.get("score", 0), reverse=True)
    top3 = sorted_results[:3]
    medals = ["🥇 1st", "🥈 2nd", "🥉 3rd"]

    elements.append(Paragraph("<b>Top 3 Rankers</b>", normal_style))
    elements.append(Spacer(1, 0.1*inch))

    rank_data = [["Rank", "Name", "Total Score"]]
    for i, r in enumerate(top3):
        rank_data.append([medals[i], r.get("name", ""), f"{r.get('score', 0)}/40"])

    rank_table = Table(rank_data, colWidths=[1.2*inch, 3*inch, 1.5*inch])
    rank_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0),  GOLD),
        ('TEXTCOLOR',  (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',   (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#FEF9C3'), colors.HexColor('#FFF7ED'), colors.HexColor('#F0FDF4')]),
    ]))
    elements.append(rank_table)

    doc.build(elements)
    output.seek(0)

    filename = f"omr_results_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/export/all-students/pdf")
async def export_all_students_pdf():
    """Single-page table of all students across all sessions with subject scores"""
    from reportlab.platypus import PageBreak, HRFlowable

    all_results = []
    for session in sessions.values():
        all_results.extend(session.results)

    if not all_results:
        raise HTTPException(status_code=400, detail="No results available")

    SUBJECTS = [
        ("Intelligence",    1,  10),
        ("Science",        11,  20),
        ("Social Science", 21,  30),
        ("Mathematics",    31,  40),
    ]

    def subject_score(answers, start, end):
        return sum(
            1 for q in range(start, end + 1)
            if answers.get(str(q), {}).get("is_correct", False)
        )

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=0.5*inch, rightMargin=0.5*inch,
        topMargin=0.6*inch, bottomMargin=0.6*inch
    )
    styles = getSampleStyleSheet()
    elements = []

    BLUE  = colors.HexColor('#2563EB')
    LGREY = colors.HexColor('#F3F4F6')
    GOLD  = colors.HexColor('#F59E0B')

    elements.append(Paragraph("<b>All Students Results</b>", styles['Title']))
    elements.append(Paragraph(
        f"Total students: {len(all_results)} &nbsp;&nbsp; Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        styles['Normal']
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=BLUE))
    elements.append(Spacer(1, 0.2*inch))

    # Table header
    header = ["#", "Name", "Intel\n/10", "Sci\n/10", "Soc.Sci\n/10", "Maths\n/10", "Total\n/40"]
    table_data = [header]

    sorted_results = sorted(all_results, key=lambda r: r.get("score", 0), reverse=True)

    for i, result in enumerate(sorted_results, 1):
        answers = result.get("answers", {})
        row = [
            str(i),
            result.get("name", ""),
        ]
        for _, s, e in SUBJECTS:
            row.append(str(subject_score(answers, s, e)))
        row.append(str(result.get("score", 0)))
        table_data.append(row)

    col_widths = [0.4*inch, 2.4*inch, 0.8*inch, 0.7*inch, 0.9*inch, 0.8*inch, 0.8*inch]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

    tbl_style = [
        ('BACKGROUND',     (0, 0), (-1, 0),  BLUE),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN',          (1, 0), (1, -1),  'LEFT'),
        ('GRID',           (0, 0), (-1, -1), 0.4, colors.grey),
        ('FONTSIZE',       (0, 0), (-1, -1), 8.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LGREY]),
    ]
    # Gold highlight top 3
    for rank_row in range(1, min(4, len(table_data))):
        tbl_style.append(('BACKGROUND', (0, rank_row), (-1, rank_row), colors.HexColor('#FEF9C3')))
        tbl_style.append(('FONTNAME',   (0, rank_row), (-1, rank_row), 'Helvetica-Bold'))

    tbl.setStyle(TableStyle(tbl_style))
    elements.append(tbl)

    # Top 3 rankers footer
    elements.append(Spacer(1, 0.3*inch))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph("<b>Top 3 Rankers</b>", styles['Normal']))
    elements.append(Spacer(1, 0.08*inch))

    medals = ["🥇 1st", "🥈 2nd", "🥉 3rd"]
    rank_data = [["Rank", "Name", "Total /40"]]
    for i, r in enumerate(sorted_results[:3]):
        rank_data.append([medals[i], r.get("name", ""), f"{r.get('score', 0)}/40"])

    rank_tbl = Table(rank_data, colWidths=[1*inch, 3*inch, 1.2*inch])
    rank_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0),  GOLD),
        ('TEXTCOLOR',  (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',   (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
            colors.HexColor('#FEF9C3'),
            colors.HexColor('#FFF7ED'),
            colors.HexColor('#F0FDF4')
        ]),
    ]))
    elements.append(rank_tbl)

    doc.build(elements)
    output.seek(0)

    filename = f"all_students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    print(f"🚀 Starting OMR Scanner API Server on port {port}...")
    uvicorn.run("api_server:app", host="0.0.0.0", port=port, reload=False)
