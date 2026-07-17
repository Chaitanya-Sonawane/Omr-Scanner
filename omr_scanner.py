"""
OMR Scanner v4.0
Detects filled bubbles in 40-question, 4-option OMR answer sheets.

Improvements over v3:
  - preprocess_for_omr() now returns (gray, quality) tuple — quality checked
    before processing; bad images raise a descriptive RuntimeError early
  - Sheet validation: aspect-ratio check, minimum coverage, convexity score
  - High-resolution warping at fixed canonical size (1200x1600) for consistent
    circle detection regardless of input resolution
  - Adaptive thresholding block-size scaled to warped image dimensions
  - Contour-based circularity validation on every detected bubble candidate
    (rejects dust spots, print artefacts, partial circles at sheet edges)
  - Fill-percentage analysis with per-question dynamic baseline calibration
  - Robust row-Y estimation: DBSCAN-style gap outlier rejection before K-means
  - Multi-mark detection tightened: requires both dark intensity AND distinct
    fill-ratio separation between top-2 candidates
  - Stability buffer API: accepts a sequence of frames and only returns a
    result when N consecutive frames agree (eliminates motion blur captures)
  - Perspective correction uses full homography from all 4 sheet corners
  - Morphological closing on ink mask fills partially-erased bubble centres
"""

import cv2
import numpy as np
import sys
from typing import Dict, List, Optional, Tuple

from enhancer import preprocess_for_omr, validate_image_quality, ImageQuality

DEBUG_ENABLED = False

# Canonical output size after perspective warp — large enough for reliable
# circle detection, small enough to process quickly.
WARP_W, WARP_H = 1200, 1600


def _log(msg: str) -> None:
    if DEBUG_ENABLED:
        print(f"[OMR] {msg}")


# ---------------------------------------------------------------------------
# Sheet detection & validation
# ---------------------------------------------------------------------------

def _order_corners(pts: np.ndarray) -> np.ndarray:
    """Order 4 points as top-left, top-right, bottom-right, bottom-left."""
    pts = np.array(pts, dtype=np.float32).reshape(4, 2)
    s = pts.sum(axis=1)
    d = np.diff(pts, axis=1).ravel()
    tl = pts[int(np.argmin(s))]
    br = pts[int(np.argmax(s))]
    tr = pts[int(np.argmin(d))]
    bl = pts[int(np.argmax(d))]
    return np.array([tl, tr, br, bl], dtype=np.float32)


def _quad_aspect_ok(quad: np.ndarray, tol: float = 0.45) -> bool:
    """
    Check that the detected quad has an aspect ratio consistent with A4/Letter
    paper (portrait ~0.71 or landscape ~1.41).  Rejects wildly-skewed quads
    that are really shadows or table edges.
    """
    ordered = _order_corners(quad)
    tl, tr, br, bl = ordered
    w = float((np.linalg.norm(tr - tl) + np.linalg.norm(br - bl)) / 2)
    h = float((np.linalg.norm(bl - tl) + np.linalg.norm(br - tr)) / 2)
    if w < 1 or h < 1:
        return False
    ratio = h / w
    # Accept both portrait (h/w ~1.41) and landscape (h/w ~0.71) with tolerance
    return abs(ratio - 1.41) < tol or abs(ratio - 0.71) < tol


def _convexity_score(contour: np.ndarray) -> float:
    """Ratio of contour area to convex hull area — 1.0 for a perfect convex shape."""
    hull_area = cv2.contourArea(cv2.convexHull(contour))
    if hull_area < 1:
        return 0.0
    return cv2.contourArea(contour) / hull_area


def _detect_sheet_quad(gray: np.ndarray) -> Optional[np.ndarray]:
    """
    Find the OMR sheet as the largest convex 4-point contour.

    Validation added over v3:
      - Aspect ratio must be consistent with A4/Letter paper
      - Convexity score > 0.85 (rejects L-shaped or irregular blobs)
      - Sheet must cover at least 25 % of the frame
    """
    h, w = gray.shape
    img_area = float(h * w)

    proc = cv2.GaussianBlur(gray, (5, 5), 0)
    _, otsu = cv2.threshold(proc, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    edges = cv2.Canny(proc, 50, 150)
    combined = cv2.bitwise_or(otsu, edges)
    kernel = np.ones((7, 7), np.uint8)
    combined = cv2.morphologyEx(combined, cv2.MORPH_CLOSE, kernel, iterations=2)

    contours, _ = cv2.findContours(combined, cv2.RETR_EXTERNAL,
                                   cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return None

    for c in sorted(contours, key=cv2.contourArea, reverse=True)[:5]:
        area = cv2.contourArea(c)
        if area < img_area * 0.25 or area > img_area * 0.999:
            continue
        if _convexity_score(c) < 0.85:
            continue

        hull = cv2.convexHull(c)
        peri = cv2.arcLength(hull, True)
        quad = None
        for eps in (0.02, 0.03, 0.05, 0.08):
            approx = cv2.approxPolyDP(hull, eps * peri, True)
            if len(approx) == 4 and cv2.isContourConvex(approx):
                quad = approx.reshape(4, 2).astype(np.float32)
                break
        if quad is None:
            box = cv2.boxPoints(cv2.minAreaRect(c))
            if cv2.contourArea(box.astype(np.float32)) >= img_area * 0.25:
                quad = box.astype(np.float32)
        if quad is not None and _quad_aspect_ok(quad):
            return quad
    return None


def _warp_to_sheet(
    img: np.ndarray,
    gray: np.ndarray,
) -> Tuple[np.ndarray, np.ndarray, bool]:
    """
    Detect the sheet quad and warp both BGR and grayscale to the fixed
    canonical size (WARP_W x WARP_H) using a full perspective homography.

    Fixed canonical size (v4 change) ensures circle radius estimates and
    morphological kernel sizes are consistent regardless of input resolution.
    """
    detect_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if img.ndim == 3 else img
    quad = _detect_sheet_quad(detect_gray)
    if quad is None:
        return img, gray, False

    ordered = _order_corners(quad)
    tl, tr, br, bl = ordered

    # Sanity: warped region must be at least 100x100
    out_w = int(max(np.linalg.norm(br - bl), np.linalg.norm(tr - tl)))
    out_h = int(max(np.linalg.norm(tr - br), np.linalg.norm(tl - bl)))
    if out_w < 100 or out_h < 100:
        return img, gray, False

    # Always warp to canonical size for consistent downstream parameters
    dst = np.array([[0, 0], [WARP_W - 1, 0],
                    [WARP_W - 1, WARP_H - 1], [0, WARP_H - 1]], dtype=np.float32)
    M = cv2.getPerspectiveTransform(ordered, dst)
    img_w  = cv2.warpPerspective(img,  M, (WARP_W, WARP_H), flags=cv2.INTER_CUBIC)
    gray_w = cv2.warpPerspective(gray, M, (WARP_W, WARP_H), flags=cv2.INTER_CUBIC)
    _log(f"Sheet warped to canonical {WARP_W}x{WARP_H}")
    return img_w, gray_w, True


# ---------------------------------------------------------------------------
# Rotation correction
# ---------------------------------------------------------------------------

def _auto_rotate(img: np.ndarray, gray: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    """
    Correct small skew (<= 15 deg) using dominant Hough line angle.
    Shi-Tomasi corners focus the edge map on meaningful sheet structure.
    """
    corners = cv2.goodFeaturesToTrack(gray, maxCorners=300, qualityLevel=0.01,
                                      minDistance=10)
    mask = np.zeros_like(gray)
    if corners is not None:
        for pt in corners:
            x, y = pt.ravel().astype(int)
            cv2.circle(mask, (x, y), 12, 255, -1)
    else:
        mask[:] = 255

    edges = cv2.bitwise_and(cv2.Canny(gray, 40, 120), mask)
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180,
                             threshold=80, minLineLength=80, maxLineGap=15)
    if lines is None or len(lines) < 5:
        return img, gray

    angles = [
        np.degrees(np.arctan2(l[0][3] - l[0][1], l[0][2] - l[0][0]))
        for l in lines[:100]
        if abs(np.degrees(np.arctan2(l[0][3] - l[0][1], l[0][2] - l[0][0]))) <= 15
    ]
    if not angles:
        return img, gray

    skew = float(np.median(angles))
    if abs(skew) < 0.4:
        return img, gray

    h, w = img.shape[:2]
    M = cv2.getRotationMatrix2D((w / 2, h / 2), skew, 1.0)
    img_r = cv2.warpAffine(img, M, (w, h), flags=cv2.INTER_CUBIC,
                            borderMode=cv2.BORDER_REPLICATE)
    _log(f"Auto-rotated {skew:.2f} deg")
    return img_r, cv2.cvtColor(img_r, cv2.COLOR_BGR2GRAY)


# ---------------------------------------------------------------------------
# Ink / filled-pixel map (v4: morphological closing added)
# ---------------------------------------------------------------------------

def _build_ink_mask(gray: np.ndarray) -> np.ndarray:
    """
    Robust binary map of filled/dark pixels.

    v4 additions over v3:
    - Morphological CLOSING after the union mask fills the hollow centres of
      partially-erased bubbles (erasing leaves a ring, not a disc; closing
      reconnects the ring interior so fill-ratio is not underestimated).
    - The opening kernel is now sized proportionally to the image so it scales
      correctly with the canonical warp size.
    """
    h, w = gray.shape
    block = max(31, (min(h, w) // 12) | 1)

    adaptive_dark = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, block, 10,
    )
    _, otsu = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    otsu_dark = cv2.bitwise_not(otsu)

    ink = cv2.bitwise_or(adaptive_dark, otsu_dark)

    # Opening: remove dust/speckle (small isolated noise)
    open_k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    ink = cv2.morphologyEx(ink, cv2.MORPH_OPEN, open_k, iterations=1)

    # Closing: reconnect the hollow centres of partially-erased marks
    close_k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    ink = cv2.morphologyEx(ink, cv2.MORPH_CLOSE, close_k, iterations=1)

    return ink


# ---------------------------------------------------------------------------
# Circularity / contour validation for detected bubbles
# ---------------------------------------------------------------------------

def _circularity(contour: np.ndarray) -> float:
    """4*pi*area / perimeter^2 — 1.0 for a perfect circle."""
    area = cv2.contourArea(contour)
    peri = cv2.arcLength(contour, True)
    if peri < 1:
        return 0.0
    return float(4 * np.pi * area / (peri * peri))


def _validate_bubble_contours(
    gray: np.ndarray,
    cx: int, cy: int, radius: int,
) -> Tuple[float, bool]:
    """
    Extract the contour(s) inside a bubble ROI and check circularity.

    Returns (best_circularity, is_valid).
    A genuine bubble has circularity > 0.65. Dust spots are irregular
    (< 0.3) and straight print lines are very low (< 0.1).
    """
    r = max(radius, 4)
    x1 = max(0, cx - r)
    y1 = max(0, cy - r)
    x2 = min(gray.shape[1], cx + r)
    y2 = min(gray.shape[0], cy + r)
    roi = gray[y1:y2, x1:x2]
    if roi.size == 0:
        return 0.0, False

    _, bw = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    cnts, _ = cv2.findContours(bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not cnts:
        return 0.0, True   # blank bubble — still valid, just unfilled

    best = max((_circularity(c) for c in cnts), default=0.0)
    return best, best > 0.45   # threshold: genuine circle vs noise


# ---------------------------------------------------------------------------
# Circle detection (Hough, 4-stage fallback)
# ---------------------------------------------------------------------------

def _try_hough(img: np.ndarray, min_r: int, max_r: int, param2: float,
               method: int = cv2.HOUGH_GRADIENT) -> Optional[np.ndarray]:
    return cv2.HoughCircles(
        img, method, dp=1, minDist=16,
        param1=50, param2=param2,
        minRadius=min_r, maxRadius=max_r,
    )


def _detect_circles(gray: np.ndarray) -> Tuple[Optional[np.ndarray], np.ndarray]:
    """
    Multi-stage Hough circle detection with four fallback strategies.
    After the canonical warp (1200x1600) the radius range is now deterministic,
    improving consistency across different input resolutions.
    """
    h, w = gray.shape
    r_est  = max(8, min(30, w // 80))
    min_r  = max(5, r_est - 7)
    max_r  = r_est + 12
    blur   = cv2.medianBlur(gray, 5)

    # Stage 1: HOUGH_GRADIENT_ALT (sub-pixel, OpenCV >= 4.5.1)
    try:
        alt = cv2.HOUGH_GRADIENT_ALT
        for p2 in [0.85, 0.75, 0.65, 0.55]:
            c = _try_hough(blur, min_r, max_r, p2, alt)
            if c is not None and len(c[0]) >= 8:
                return c, blur
    except AttributeError:
        pass

    # Stage 2: Classic HOUGH_GRADIENT, decreasing threshold
    for p2 in [22, 18, 14, 10, 8]:
        c = _try_hough(blur, min_r, max_r, p2)
        if c is not None and len(c[0]) >= 8:
            return c, blur

    # Stage 3: Bilateral filter (preserves edges, reduces noise)
    bilat = cv2.bilateralFilter(gray, 9, 75, 75)
    blur3 = cv2.medianBlur(bilat, 3)
    for p2 in [18, 14, 10, 7]:
        c = _try_hough(blur3, min_r, max_r, p2)
        if c is not None and len(c[0]) >= 8:
            return c, blur3

    # Stage 4: Aggressive CLAHE + Otsu binarisation
    eq   = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8)).apply(gray)
    _, bw = cv2.threshold(eq, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    blur4 = cv2.medianBlur(bw, 3)
    for p2 in [15, 10, 7, 5]:
        c = _try_hough(blur4, max(4, r_est - 9), r_est + 15, p2)
        if c is not None and len(c[0]) >= 8:
            return c, blur4

    return None, blur


# ---------------------------------------------------------------------------
# Grid construction (v4: gap-outlier rejection before K-means)
# ---------------------------------------------------------------------------

def _safe_kmeans(data: np.ndarray, k: int, criteria: tuple, attempts: int = 10):
    if len(data) < k:
        return None, None
    _, labels, centers = cv2.kmeans(data, k, None, criteria, attempts,
                                    cv2.KMEANS_PP_CENTERS)
    return labels, centers


def _reject_row_outliers(ys: np.ndarray) -> np.ndarray:
    """
    Remove Y-values that are clear outliers relative to the expected uniform
    row spacing, before passing to K-means.

    Strategy: compute median gap between sorted Y-values; reject any Y that
    is more than 2.5 * median_gap away from all its neighbours.  This stops
    false-positive circles (logos, stamps, stray ink) from displacing K-means
    row centroids.
    """
    if len(ys) < 4:
        return ys
    sorted_y = np.sort(ys)
    gaps = np.diff(sorted_y.astype(float))
    pos_gaps = gaps[gaps > 2]
    if len(pos_gaps) == 0:
        return ys
    med_gap = float(np.median(pos_gaps))
    if med_gap < 1:
        return ys
    keep = np.ones(len(sorted_y), dtype=bool)
    for i in range(len(sorted_y)):
        dists = np.abs(sorted_y - sorted_y[i])
        dists[i] = np.inf
        min_dist = float(np.min(dists))
        if min_dist > 2.5 * med_gap:
            keep[i] = False
    filtered = sorted_y[keep]
    return filtered if len(filtered) >= 8 else ys


def _compute_row_y(all_pts: np.ndarray) -> List[float]:
    """
    Estimate 20 shared row Y-positions.

    v4: gap-outlier rejection applied to Y-coords before K-means so that
    stray circles (logos/stamps not caught by the header-gap filter) cannot
    displace row centroids.
    """
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 200, 0.1)
    ys_raw = all_pts[:, 1].astype(np.float32)
    ys_clean = _reject_row_outliers(ys_raw)
    ys = ys_clean.reshape(-1, 1)

    labels, centers = _safe_kmeans(ys, min(20, len(ys)), criteria)
    if labels is None:
        return np.linspace(float(ys.min()), float(ys.max()), 20).tolist()

    row_y = sorted(float(c) for c in centers.flatten())
    if len(row_y) < 20:
        diffs = np.diff(row_y)
        pitch = float(np.median(diffs[diffs > 3])) if any(diffs > 3) else 30.0
        while len(row_y) < 20:
            row_y.append(row_y[-1] + pitch)

    # Polyfit outlier snapping (two passes)
    idx = np.arange(20, dtype=np.float64)
    arr = np.array(row_y, dtype=np.float64)
    diffs = np.diff(arr)
    pitch = float(np.median(diffs[diffs > 3])) if any(diffs > 3) else 30.0
    for _ in range(2):
        a, b = np.polyfit(idx, arr, 1)
        fitted = a * idx + b
        bad = np.abs(arr - fitted) > max(pitch * 0.4, 8)
        if not bad.any():
            break
        arr[bad] = fitted[bad]

    return arr.tolist()


def _build_grid(block_pts: np.ndarray, gray: np.ndarray,
                radius_est: int, row_y: List[float]) -> Dict[int, List[Tuple[int, int]]]:
    """
    Build a 20x4 grid of (x, y) sample centres for one answer block.
    Per-column linear regression tracks perspective drift across rows.
    """
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 200, 0.1)
    h, w = gray.shape
    bxs = block_pts[:, 0].reshape(-1, 1).astype(np.float32)
    n_cols = min(4, len(block_pts))
    col_labels, col_centers = _safe_kmeans(bxs, n_cols, criteria)

    if col_labels is None or n_cols < 4:
        col_centers_list = np.linspace(float(bxs.min()), float(bxs.max()), 4).tolist()
        col_labels_arr   = np.zeros((len(block_pts), 1), dtype=np.int32)
        cluster_to_opt   = {0: 0}
    else:
        order = np.argsort(col_centers.flatten())
        col_centers_list = sorted(col_centers.flatten().tolist())
        cluster_to_opt   = {int(order[i]): i for i in range(len(order))}
        col_labels_arr   = col_labels

    col_pts: Dict[int, list] = {i: [] for i in range(4)}
    for i, pt in enumerate(block_pts):
        opt = cluster_to_opt.get(int(col_labels_arr[i][0]), -1)
        if 0 <= opt < 4:
            col_pts[opt].append(pt)

    row_y_arr = np.array(row_y, dtype=np.float64)
    grid: Dict[int, List[Tuple[int, int]]] = {}

    for opt in range(4):
        pts_list = col_pts[opt]
        if len(pts_list) >= 2:
            px = np.array([p[0] for p in pts_list], dtype=np.float64)
            py = np.array([p[1] for p in pts_list], dtype=np.float64)
            ri_of_pt = np.array([int(np.argmin(np.abs(row_y_arr - yp))) for yp in py])
            if len(np.unique(ri_of_pt)) >= 2:
                a, b = np.polyfit(ri_of_pt, px, 1)
            else:
                a, b = 0.0, float(np.median(px))
        else:
            a, b = 0.0, (col_centers_list[opt] if opt < len(col_centers_list) else 0.0)

        col_grid = []
        for ri in range(20):
            gx = int(round(a * ri + b))
            gy = int(round(row_y[ri]))
            gx = max(radius_est, min(w - radius_est - 1, gx))
            gy = max(radius_est, min(h - radius_est - 1, gy))
            col_grid.append((gx, gy))
        grid[opt] = col_grid

    return grid


# ---------------------------------------------------------------------------
# Per-question dynamic threshold calibration
# ---------------------------------------------------------------------------

def _calibrate_thresholds(
    all_intensities: List[float],
) -> Tuple[float, float, float]:
    """
    Derive fill / absolute-dark thresholds from the intensity distribution of
    all bubbles across the sheet.

    v4 change: use the inter-quartile range (IQR) to set a tighter band.
    A sheet full of heavily-marked bubbles has a low 75th percentile; a blank
    sheet has a high one.  The IQR separates "definitely blank" from
    "definitely filled" with a self-calibrating gap.

    Returns (blank_baseline, fill_threshold, absolute_dark).
    """
    blank_baseline = float(np.percentile(all_intensities, 75))
    q15            = float(np.percentile(all_intensities, 15))
    q25            = float(np.percentile(all_intensities, 25))
    iqr            = blank_baseline - q25

    # fill_threshold: anything darker than this is a candidate mark
    fill_threshold = blank_baseline - max(iqr * 0.5, blank_baseline * 0.12)

    # absolute_dark: unambiguously filled regardless of neighbourhood
    absolute_dark  = min(q15 * 1.5, blank_baseline * 0.68)

    _log(f"baseline={blank_baseline:.1f} fill_thresh={fill_threshold:.1f} "
         f"abs_dark={absolute_dark:.1f} iqr={iqr:.1f}")
    return blank_baseline, fill_threshold, absolute_dark


# ---------------------------------------------------------------------------
# Main detection function
# ---------------------------------------------------------------------------

def detect_bubbles(
    img_path: str,
    debug_out: Optional[str] = None,
    validate_quality: bool = False,
) -> Tuple[
    Dict[int, Optional[int]],
    Dict[int, str],
    Dict[int, Dict[int, float]],
    Dict[int, float],
]:
    """
    Detect filled bubbles in a 40-question / 4-option OMR answer sheet.

    Parameters
    ----------
    img_path         : path to input image
    debug_out        : optional path to save annotated debug image
    validate_quality : when True, raise RuntimeError on bad image quality

    Returns
    -------
    answers         : {question: option (1-4) or None}
    flags           : {question: flag-string}
    raw_intensities : {question: {option: mean_gray_value}}
    confidence      : {question: score 0-100}
    """
    # ------------------------------------------------------------------
    # 1. Load + preprocess
    # ------------------------------------------------------------------
    gray, quality = preprocess_for_omr(img_path, validate=validate_quality)

    if validate_quality and quality is not None and not quality.is_acceptable:
        raise RuntimeError(
            f"Image quality check failed for '{img_path}': {quality.reason}"
        )

    img = cv2.imread(img_path)
    if img is None:
        raise RuntimeError(f"Could not read image: {img_path}")

    # ------------------------------------------------------------------
    # 2. Sheet isolation + canonical perspective warp
    # ------------------------------------------------------------------
    img, gray, _warped = _warp_to_sheet(img, gray)
    img, gray = _auto_rotate(img, gray)
    h, w = gray.shape

    # ------------------------------------------------------------------
    # 3. Circle detection
    # ------------------------------------------------------------------
    raw_circles, _ = _detect_circles(gray)
    if raw_circles is None:
        raise RuntimeError(
            f"No circles detected in '{img_path}'. "
            "Ensure the image shows a clear OMR sheet with visible bubbles."
        )

    circles = np.round(raw_circles[0]).astype(int).tolist()

    # Remove circles too close to image borders
    circles = [c for c in circles if 40 < c[0] < w - 40 and 150 < c[1] < h - 20]
    if len(circles) < 8:
        raise RuntimeError(f"Too few circles after border filter ({len(circles)}).")

    # Keep only circles near the median radius
    radii = np.array([c[2] for c in circles])
    r_med = float(np.median(radii))
    circles = [c for c in circles if abs(c[2] - r_med) <= max(2.5, r_med * 0.35)]
    if len(circles) < 8:
        raise RuntimeError(f"Too few circles after radius filter ({len(circles)}).")

    # ------------------------------------------------------------------
    # 4. Circularity validation — reject non-circular detections
    # ------------------------------------------------------------------
    valid_circles = []
    for c in circles:
        circ, ok = _validate_bubble_contours(gray, c[0], c[1], c[2])
        if ok:
            valid_circles.append(c)
        else:
            _log(f"Rejected non-circular detection at ({c[0]},{c[1]}) circ={circ:.2f}")
    if len(valid_circles) >= 8:
        circles = valid_circles
    else:
        _log("Circularity filter would leave too few circles — skipping filter")

    # ------------------------------------------------------------------
    # 5. Header debris removal (largest vertical gap heuristic)
    # ------------------------------------------------------------------
    ys = sorted(c[1] for c in circles)
    n_total = len(ys)
    cut_y, best_gap = None, 0
    for i in range(1, min(60, n_total)):
        gap = ys[i] - ys[i - 1]
        if gap > 55 and gap > best_gap:
            remaining = sum(1 for y in ys if y >= ys[i])
            if remaining >= max(8, n_total * 0.55):
                best_gap = gap
                cut_y = ys[i]
    if cut_y is not None:
        circles = [c for c in circles if c[1] >= cut_y]
    if len(circles) < 8:
        raise RuntimeError(f"Too few circles after header removal ({len(circles)}).")

    pts        = np.array([[c[0], c[1]] for c in circles])
    radius_est = int(np.median([c[2] for c in circles]))

    # Hard X-midpoint split into two answer blocks
    block_ids = np.where(pts[:, 0] < w / 2.0, 0, 1)

    row_y  = _compute_row_y(pts)
    # Sample radius: 60% of detected radius gives clean fill measurements
    # without spilling into adjacent bubbles
    sample_r = max(int(radius_est * 0.60), 6)

    # ------------------------------------------------------------------
    # 6. Ink mask (adaptive + Otsu union, opening + closing)
    # ------------------------------------------------------------------
    ink_mask = _build_ink_mask(gray)

    # ------------------------------------------------------------------
    # 7. Measure every bubble
    # ------------------------------------------------------------------
    raw_intensities: Dict[int, Dict[int, float]] = {}
    bubble_details:  Dict[int, Dict[int, dict]]  = {}
    debug_img = img.copy()

    for block_id, q_offset in [(0, 0), (1, 20)]:
        block_pts = pts[block_ids == block_id]

        if len(block_pts) < 4:
            for ri in range(20):
                q = ri + 1 + q_offset
                raw_intensities[q] = {o: 255.0 for o in range(1, 5)}
                bubble_details[q]  = {o: {'mean': 255.0, 'filled_ratio': 0.0,
                                          'circularity': 0.0}
                                      for o in range(1, 5)}
            continue

        col_grid = _build_grid(block_pts, gray, radius_est, row_y)

        for ri in range(20):
            q = ri + 1 + q_offset
            raw_intensities[q] = {}
            bubble_details[q]  = {}

            for opt_idx in range(4):
                cx, cy = col_grid[opt_idx][ri]
                circ_mask = np.zeros(gray.shape, dtype=np.uint8)
                cv2.circle(circ_mask, (cx, cy), sample_r, 255, -1)

                mean_val = float(cv2.mean(gray, mask=circ_mask)[0])

                dark_px  = cv2.countNonZero(cv2.bitwise_and(ink_mask, circ_mask))
                total_px = cv2.countNonZero(circ_mask)
                fill_ratio = dark_px / (total_px + 1e-9)

                # Circularity of the ink blob inside this bubble
                circ_val, _ = _validate_bubble_contours(gray, cx, cy, sample_r)

                opt = opt_idx + 1
                raw_intensities[q][opt] = mean_val
                bubble_details[q][opt] = {
                    'mean': mean_val,
                    'filled_ratio': fill_ratio,
                    'circularity': circ_val,
                }

                if debug_out:
                    colour = (0, 200, 0) if fill_ratio < 0.15 else (0, 100, 255)
                    cv2.circle(debug_img, (cx, cy), radius_est, colour, 2)
                    cv2.putText(debug_img, f"{q}.{opt}",
                                (cx - radius_est, cy - radius_est - 2),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.3, (200, 0, 0), 1,
                                cv2.LINE_AA)

    if debug_out:
        cv2.imwrite(debug_out, debug_img)

    # ------------------------------------------------------------------
    # 8. Dynamic threshold calibration
    # ------------------------------------------------------------------
    all_vals = [v for row in raw_intensities.values() for v in row.values()]
    blank_baseline, fill_threshold, absolute_dark = _calibrate_thresholds(all_vals)

    # ------------------------------------------------------------------
    # 9. Two-pass decision with tightened multi-mark logic
    # ------------------------------------------------------------------
    answers:    Dict[int, Optional[int]] = {}
    flags:      Dict[int, str]           = {}
    confidence: Dict[int, float]         = {}

    for q in sorted(raw_intensities):
        opts   = raw_intensities[q]
        dets   = bubble_details[q]
        ranked = sorted(opts.items(), key=lambda x: x[1])

        darkest_opt, darkest_val = ranked[0]
        second_opt,  second_val  = ranked[1]
        brightest_val            = ranked[3][1]

        darkest_fill  = dets[darkest_opt]['filled_ratio']
        second_fill   = dets[second_opt]['filled_ratio']
        darkest_circ  = dets[darkest_opt]['circularity']

        row_mean = float(np.mean([v for _, v in ranked]))
        row_std  = float(np.std([v for _, v in ranked]))
        gap      = second_val - darkest_val
        rel_gap  = gap / (row_mean + 1e-9)

        # Pass 1 — intensity evidence
        p1_obvious  = darkest_val < absolute_dark
        p1_relative = rel_gap > 0.10 and gap > 4
        p1_below    = darkest_val < fill_threshold
        pass1 = p1_obvious or (p1_below and p1_relative)

        # Pass 2 — fill-ratio evidence
        p2_filled   = darkest_fill > 0.15
        p2_distinct = darkest_fill > second_fill * 1.4
        pass2 = p2_filled and p2_distinct

        # Circularity guard: ink blob must be reasonably circular.
        # If circularity is very low it is likely a dust streak or smudge.
        circ_ok = darkest_circ > 0.35 or darkest_fill < 0.10

        is_filled = circ_ok and ((pass1 and pass2) or p1_obvious or
                                  (p2_filled and p1_below))

        if is_filled:
            darkness_score = max(0.0, 1.0 - darkest_val / (blank_baseline + 1e-9))
            gap_score      = min(1.0, rel_gap / 0.20)
            fill_score     = min(1.0, darkest_fill / 0.30)
            circ_score     = min(1.0, darkest_circ / 0.80)
            # Circularity contributes 10 % of the confidence score
            conf = int(min(100,
                           darkness_score * 40 +
                           gap_score      * 30 +
                           fill_score     * 20 +
                           circ_score     * 10))

            answers[q]    = darkest_opt
            confidence[q] = float(conf)

            # Tightened multi-mark: BOTH intensity AND fill must agree
            second_intensity_dark = (second_val < fill_threshold * 0.92
                                     and rel_gap < 0.07)
            second_fill_dark      = second_fill > 0.15
            second_fill_close     = darkest_fill < second_fill * 1.6

            if second_intensity_dark and second_fill_dark and second_fill_close:
                flags[q]      = "multi_mark"
                confidence[q] = float(max(25, conf - 30))
                _log(f"Q{q}: multi_mark (opts {darkest_opt} & {second_opt})")
            elif conf < 40:
                flags[q] = "low_confidence"
                _log(f"Q{q}: low_confidence ({conf}%)")
        else:
            answers[q]    = None
            confidence[q] = 0.0

            all_similar = row_std < 4
            all_dark    = brightest_val < fill_threshold * 0.88
            all_filled  = all(bubble_details[q][o]['filled_ratio'] > 0.18
                              for o in range(1, 5))
            if (all_dark and all_similar) or all_filled:
                flags[q] = "row_smudged"
                _log(f"Q{q}: row_smudged")
            else:
                flags[q] = "no_clear_mark"
                _log(f"Q{q}: no_clear_mark")

    return answers, flags, raw_intensities, confidence


# ---------------------------------------------------------------------------
# Stability-buffer API (multi-frame capture validation)
# ---------------------------------------------------------------------------

class StabilityBuffer:
    """
    Collects the last N detection results from consecutive frames and signals
    when the sheet has been stable long enough to trust the capture.

    Typical use (live camera loop):

        buf = StabilityBuffer(n_frames=5, agreement_threshold=0.90)
        while cap.isOpened():
            ret, frame = cap.read()
            cv2.imwrite("_tmp.jpg", frame)
            try:
                ans, flags, _, conf = detect_bubbles("_tmp.jpg")
            except RuntimeError:
                buf.reset()
                continue
            if buf.push(ans, conf):
                final_ans = buf.best_result()
                break
    """

    def __init__(self, n_frames: int = 5, agreement_threshold: float = 0.90):
        self.n_frames            = n_frames
        self.agreement_threshold = agreement_threshold
        self._history: List[Tuple[Dict, Dict]] = []   # (answers, confidence)

    def reset(self) -> None:
        self._history.clear()

    def push(
        self,
        answers: Dict[int, Optional[int]],
        confidence: Dict[int, float],
    ) -> bool:
        """
        Add one frame result.  Returns True when the buffer is full and all
        frames agree on >= agreement_threshold of questions.
        """
        self._history.append((answers, confidence))
        if len(self._history) > self.n_frames:
            self._history.pop(0)
        if len(self._history) < self.n_frames:
            return False
        return self._is_stable()

    def _is_stable(self) -> bool:
        reference = self._history[0][0]
        matches   = 0
        total     = 0
        for ans, _ in self._history[1:]:
            for q in range(1, 41):
                total   += 1
                matches += int(ans.get(q) == reference.get(q))
        return (matches / (total + 1e-9)) >= self.agreement_threshold

    def best_result(self) -> Tuple[Dict, Dict]:
        """
        Return the (answers, confidence) pair with the highest average
        confidence from the stable buffer.
        """
        return max(
            self._history,
            key=lambda t: float(np.mean(list(t[1].values()))) if t[1] else 0.0,
        )


# ---------------------------------------------------------------------------
# Scoring & batch utilities
# ---------------------------------------------------------------------------

def score_sheet(
    answers: Dict[int, Optional[int]],
    answer_key: Dict[int, int],
) -> Tuple[int, Dict[int, bool]]:
    """Compare answers to the answer key. Returns (score, {q: is_correct})."""
    correct: Dict[int, bool] = {}
    score = 0
    for q, key_opt in answer_key.items():
        ok = answers.get(q) == key_opt
        correct[q] = ok
        if ok:
            score += 1
    return score, correct


def batch_to_excel(
    image_paths: List[str],
    out_xlsx: str,
    answer_key: Optional[Dict[int, int]] = None,
    roll_numbers: Optional[List[str]] = None,
) -> None:
    """Generate an Excel report with colour-coded confidence indicators."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    header = ["Sheet #", "Roll No"] + [f"Q{i}" for i in range(1, 41)]
    header += (["Score", "Flagged", "Avg Confidence"] if answer_key
                else ["Flagged", "Avg Confidence"])
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fill_yellow = PatternFill("solid", start_color="FFFF00")
    fill_orange = PatternFill("solid", start_color="FFA500")
    fill_red    = PatternFill("solid", start_color="FF6B6B")

    for idx, path in enumerate(image_paths, start=1):
        roll = (roll_numbers[idx - 1] if roll_numbers and idx <= len(roll_numbers) else "")
        try:
            ans, flg, _, conf = detect_bubbles(path)
            row = [idx, roll] + [ans.get(q, "") for q in range(1, 41)]
            avg_conf = int(np.mean([conf.get(q, 0) for q in range(1, 41)]))
            if answer_key:
                sc, _ = score_sheet(ans, answer_key)
                row += [sc, ", ".join(f"Q{q}" for q in sorted(flg)), avg_conf]
            else:
                row += [", ".join(f"Q{q}" for q in sorted(flg)), avg_conf]
            ws.append(row)
            rn = ws.max_row
            for q, ftype in flg.items():
                col = 2 + q
                if ftype == "multi_mark":
                    ws.cell(row=rn, column=col).fill = fill_orange
                elif ftype == "low_confidence":
                    ws.cell(row=rn, column=col).fill = fill_yellow
                elif ftype in ("no_clear_mark", "row_smudged"):
                    ws.cell(row=rn, column=col).fill = fill_red
        except Exception as exc:
            row = [idx, roll] + ["ERROR"] * 40
            row += ([0, str(exc), 0] if answer_key else [str(exc), 0])
            ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 10
    wb.save(out_xlsx)


def print_answers(answers: Dict[int, Optional[int]], total_questions: int = 40) -> None:
    for q in range(1, total_questions + 1):
        opt = answers.get(q)
        print(f"Q{q}: {'Option ' + str(opt) if opt else 'Not marked'}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--batch":
        _out  = sys.argv[2]
        _imgs = sys.argv[3:]
        batch_to_excel(_imgs, _out)
        print(f"Saved results for {len(_imgs)} sheet(s) to {_out}")
    else:
        _path = sys.argv[1] if len(sys.argv) > 1 else None
        if not _path:
            print("Usage: python omr_scanner.py <image> [debug_out.jpg]")
            sys.exit(1)
        _debug = sys.argv[2] if len(sys.argv) > 2 else None
        _ans, _flg, _, _conf = detect_bubbles(_path, _debug, validate_quality=True)
        print_answers(_ans)
        if _flg:
            print(f"\nFlags: {', '.join(f'Q{q}={v}' for q, v in sorted(_flg.items()))}")
        print(f"Avg confidence: {int(np.mean([_conf.get(q, 0) for q in range(1, 41)]))}%")
